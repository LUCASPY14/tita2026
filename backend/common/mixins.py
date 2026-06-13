"""
Mixins reutilizables para vistas de la API
"""

import csv
import io
from django.http import HttpResponse


class ExportCSVMixin:
    """
    Agrega exportación CSV y Excel al endpoint list.

    ?formato=csv  → descarga CSV (UTF-8 BOM, compatible con Excel latino)
    ?formato=xlsx → descarga Excel (.xlsx) con estilos básicos

    Uso en el ViewSet:
        export_fields = [("Encabezado", "attr_o_callable"), ...]
        export_filename = "nombre_sin_extension"

    El atributo puede ser:
        - string "campo"          → str(obj.campo)
        - string "campo__sub"     → dot-walk: obj.campo.sub
        - callable lambda obj: …  → resultado libre
    """

    export_fields: list = []
    export_filename: str = "export"

    def _get_value(self, obj, attr):
        if callable(attr):
            return attr(obj)
        parts = attr.split("__")
        val = obj
        for part in parts:
            if val is None:
                return ""
            val = getattr(val, part, "")
        return val if val is not None else ""

    def list(self, request, *args, **kwargs):
        fmt = request.query_params.get("formato")
        if fmt == "csv":
            return self._export_csv(request, *args, **kwargs)
        if fmt == "xlsx":
            return self._export_xlsx(request, *args, **kwargs)
        return super().list(request, *args, **kwargs)

    def _get_rows(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        headers = [h for h, _ in self.export_fields]
        attrs = [a for _, a in self.export_fields]
        rows = [[self._get_value(obj, a) for a in attrs] for obj in queryset]
        return headers, rows

    def _export_csv(self, request, *args, **kwargs):
        headers, rows = self._get_rows(request, *args, **kwargs)
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = (
            f'attachment; filename="{self.export_filename}.csv"'
        )
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response

    def _export_xlsx(self, request, *args, **kwargs):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse(
                "openpyxl no está instalado.", status=500, content_type="text/plain"
            )

        headers, rows = self._get_rows(request, *args, **kwargs)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.export_filename[:31]  # Excel limit

        # Header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="1E40AF")  # blue-800
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Data rows (alternate row shading)
        fill_alt = PatternFill(fill_type="solid", fgColor="EFF6FF")  # blue-50
        for i, row in enumerate(rows, start=2):
            ws.append([str(v) if v is not None else "" for v in row])
            if i % 2 == 0:
                for cell in ws[i]:
                    cell.fill = fill_alt

        # Auto column widths (capped at 50)
        for col_idx, col in enumerate(ws.columns, 1):
            width = max(
                (len(str(cell.value or "")) for cell in col), default=8
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 4, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{self.export_filename}.xlsx"'
        )
        return response
