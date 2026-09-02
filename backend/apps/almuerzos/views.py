"""
Views para la app almuerzos
"""

import logging
from datetime import date

logger = logging.getLogger(__name__)
from decimal import Decimal

from django.db import models, transaction

import csv

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.response import Response
from rest_framework.views import APIView

from rest_framework import serializers as drf_serializers

from common.permissions import (
    IsAdmin, IsAdminOrReadOnly, IsCajeroOrAdmin, IsCajeroCobradorOrAdmin,
    IsStaffOrClienteWeb, IsStaffUser,
)
from common.utils.medios_pago import resolver_medio_pago
from common.throttling import SensitiveEndpointThrottle
from apps.usuarios.auditoria import registrar_auditoria

import django_filters
from django_filters.rest_framework import DjangoFilterBackend


class _EstadoInFilter(django_filters.BaseInFilter, django_filters.CharFilter):
    """Permite filtrar por varios estados a la vez: ?estado=PENDIENTE,PARCIAL"""

from .models import (
    Alergeno,
    CuentaAlmuerzoMensual,
    DetalleMenuDiario,
    MenuDiario,
    MovimientoSaldoAlmuerzo,
    PagoCuentaAlmuerzo,
    PlanAlmuerzo,
    PrecioAlmuerzo,
    ProductoAlergeno,
    RecargaSaldoAlmuerzo,
    RegistroConsumoAlmuerzo,
    SaldoAlmuerzo,
    SuscripcionAlmuerzo,
    TipoAlmuerzo,
)
from .serializers import (
    AlergenoSerializer,
    CuentaAlmuerzoMensualSerializer,
    DetalleMenuDiarioSerializer,
    MenuDiarioSerializer,
    MovimientoSaldoAlmuerzoSerializer,
    PagoCuentaAlmuerzoSerializer,
    PlanAlmuerzoSerializer,
    PrecioAlmuerzoSerializer,
    ProductoAlergenoSerializer,
    RecargaSaldoAlmuerzoSerializer,
    RegistroConsumoAlmuerzoSerializer,
    SaldoAlmuerzoSerializer,
    SuscripcionAlmuerzoSerializer,
    TipoAlmuerzoSerializer,
)
from .filters import RegistroConsumoFilter
from .services import AlmuerzoService
from .validators import validar_limite_registros_diarios, validar_restricciones_alergenicas


# ==============================================================================
# HELPERS
# ==============================================================================

def get_precio_almuerzo_activo(fecha=None):
    """Retorna el PrecioAlmuerzo vigente para la fecha dada."""
    if fecha is None:
        fecha = date.today()
    return (
        PrecioAlmuerzo.objects.filter(
            fecha_inicio_vigencia__lte=fecha,
            activo=True,
        )
        .filter(
            models.Q(fecha_fin_vigencia__isnull=True) |
            models.Q(fecha_fin_vigencia__gte=fecha)
        )
        .order_by("-fecha_inicio_vigencia")
        .first()
    )


# ==============================================================================
# PRECIO ALMUERZO
# ==============================================================================

class PrecioAlmuerzoViewSet(viewsets.ModelViewSet):
    queryset = PrecioAlmuerzo.objects.all()
    serializer_class = PrecioAlmuerzoSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ["activo"]
    ordering = ["-fecha_inicio_vigencia"]

    @action(detail=False, methods=["get"], url_path="precio-actual")
    def precio_actual(self, request):
        precio = get_precio_almuerzo_activo()
        if precio:
            return Response(PrecioAlmuerzoSerializer(precio).data)
        return Response(
            {"error": "No hay un precio de almuerzo vigente configurado. Configure uno en el admin."},
            status=status.HTTP_404_NOT_FOUND,
        )


# ==============================================================================
# TIPO ALMUERZO
# ==============================================================================

class TipoAlmuerzoViewSet(viewsets.ModelViewSet):
    queryset = TipoAlmuerzo.objects.all()
    serializer_class = TipoAlmuerzoSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ["activo"]
    search_fields = ["nombre"]


# ==============================================================================
# PLAN ALMUERZO
# ==============================================================================

class PlanAlmuerzoViewSet(viewsets.ModelViewSet):
    queryset = PlanAlmuerzo.objects.all()
    serializer_class = PlanAlmuerzoSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ["activo", "tipo"]
    search_fields = ["nombre"]


# ==============================================================================
# SUSCRIPCION ALMUERZO
# ==============================================================================

class SuscripcionAlmuerzoViewSet(viewsets.ModelViewSet):
    serializer_class = SuscripcionAlmuerzoSerializer
    permission_classes = [IsStaffOrClienteWeb]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ["estado", "hijo", "plan"]
    ordering = ["-fecha_inicio"]

    def get_queryset(self):
        qs = SuscripcionAlmuerzo.objects.select_related("hijo", "plan").all()
        user = self.request.user
        if user.rol == "CLIENTE_WEB":
            if not user.cliente:
                return qs.none()
            return qs.filter(hijo__cliente_responsable=user.cliente)
        return qs


# ==============================================================================
# REGISTRO CONSUMO ALMUERZO
# ==============================================================================

class RegistroConsumoAlmuerzoViewSet(viewsets.ModelViewSet):
    """
    ViewSet para registrar consumos de almuerzo.

    REGLA DE NEGOCIO:
    - La tarjeta se usa SOLO como identificacion de acceso al comedor.
    - NO se descuenta saldo de la tarjeta.
    - Maximo 2 registros por alumno por dia, pero se factura como 1 ALMUERZO por dia:
        1er registro del dia: ya_cobrado=True  -> se agrega el costo a la cuenta mensual
        2do registro del dia: ya_cobrado=False -> costo=0, solo trazabilidad
    - Limite de credito mensual: si el plan tiene limite_credito_mensual, se bloquea
      cuando el monto acumulado en CuentaAlmuerzoMensual alcanza ese tope.
    """

    queryset = RegistroConsumoAlmuerzo.objects.select_related(
        "hijo", "suscripcion", "tipo_almuerzo", "nro_tarjeta"
    ).all()
    serializer_class = RegistroConsumoAlmuerzoSerializer
    permission_classes = [IsStaffOrClienteWeb]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

    def get_permissions(self):
        if self.action in ("create", "update", "partial_update", "destroy", "anular"):
            return [IsCajeroOrAdmin()]
        return [IsStaffOrClienteWeb()]

    def destroy(self, request, *args, **kwargs):
        if not (request.user and request.user.is_authenticated and request.user.rol == "ADMIN"):
            return Response(
                {"error": "Solo el Administrador puede eliminar registros de consumo."},
                status=status.HTTP_403_FORBIDDEN,
            )
        registro = self.get_object()
        if registro.estado != RegistroConsumoAlmuerzo.Estado.ANULADO:
            return Response(
                {"error": "Solo se pueden eliminar registros en estado ANULADO. Anulá el registro primero."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        with transaction.atomic():
            # El DELETE dispara el trigger trg_sync_cuenta_almuerzo, que
            # recalcula cantidad_almuerzos/monto_total desde los registros
            # reales restantes — no hay que restar nada acá.
            cuenta = CuentaAlmuerzoMensual.objects.select_for_update().filter(
                hijo=registro.hijo,
                anio=registro.fecha_consumo.year,
                mes=registro.fecha_consumo.month,
            ).first()
            registro.delete()
            if cuenta and registro.ya_cobrado and registro.costo_almuerzo:
                cuenta.refresh_from_db(fields=["monto_total", "monto_pagado", "estado", "fecha_pago"])
                cuenta.actualizar_estado()
        registrar_auditoria(
            request=request,
            operacion="ELIMINAR_REGISTRO_ALMUERZO",
            tabla="almuerzos_registroconsumoalmuerzo",
            descripcion=f"Registro ANULADO eliminado — hijo={registro.hijo_id} fecha={registro.fecha_consumo}",
        )
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(
        detail=True, methods=["post"], url_path="anular",
        throttle_classes=[SensitiveEndpointThrottle],
    )
    def anular(self, request, pk=None):
        """
        Anula un registro de consumo en estado REGISTRADO: revierte el saldo
        de almuerzo debitado (si estaba cobrado) y resincroniza la cuenta
        mensual del alumno. `estado` es read-only en el serializer a propósito
        (un PATCH/PUT genérico no debe poder mover el estado) — esta es la
        única vía habilitada para anular.
        """
        registro = self.get_object()
        if registro.estado != RegistroConsumoAlmuerzo.Estado.REGISTRADO:
            # ValidationError (no un Response directo) para que
            # common.exceptions.custom_exception_handler la normalice a
            # {"detail": ...} — es lo que el frontend sabe leer.
            raise ValidationError({"error": "Solo se pueden anular registros en estado REGISTRADO."})

        with transaction.atomic():
            registro.estado = RegistroConsumoAlmuerzo.Estado.ANULADO
            registro.save(update_fields=["estado"])
            # El UPDATE de arriba dispara el trigger trg_sync_cuenta_almuerzo,
            # que recalcula cantidad_almuerzos/monto_total sin este registro
            # (su estado ya es ANULADO) — acá solo recalculamos el estado
            # derivado y revertimos el saldo corriente si estaba cobrado.
            if registro.ya_cobrado and registro.costo_almuerzo:
                cuenta = CuentaAlmuerzoMensual.objects.select_for_update().filter(
                    hijo=registro.hijo,
                    anio=registro.fecha_consumo.year,
                    mes=registro.fecha_consumo.month,
                ).first()
                if cuenta:
                    cuenta.refresh_from_db(fields=["monto_total", "monto_pagado", "estado", "fecha_pago"])
                    cuenta.actualizar_estado()
                AlmuerzoService._revertir_saldo_almuerzo(registro)

        registrar_auditoria(
            request=request,
            operacion="ANULAR_REGISTRO_ALMUERZO",
            tabla="almuerzos_registroconsumoalmuerzo",
            id_registro=registro.id,
            descripcion=(
                f"Consumo anulado — hijo={registro.hijo_id} "
                f"fecha={registro.fecha_consumo} costo={registro.costo_almuerzo} Gs."
            ),
        )
        return Response(self.get_serializer(registro).data)

    filterset_class = RegistroConsumoFilter
    search_fields = ["hijo__nombre", "hijo__apellido"]
    ordering = ["-fecha_consumo", "-hora_registro"]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        advertencias = self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        data = dict(serializer.data)
        if advertencias:
            data["advertencias"] = advertencias
        return Response(data, status=status.HTTP_201_CREATED, headers=headers)

    def perform_create(self, serializer):
        registro_data = serializer.validated_data
        hijo = registro_data.get("hijo")
        fecha_consumo = registro_data.get("fecha_consumo")
        nro_tarjeta = registro_data.get("nro_tarjeta")
        tipo_almuerzo = registro_data.get("tipo_almuerzo")
        suscripcion = registro_data.get("suscripcion")

        # Tarjeta requerida como identificacion
        if not nro_tarjeta:
            raise ValidationError({"error": "Debe especificar la tarjeta para registrar el ingreso al almuerzo"})

        # Validar que la tarjeta pertenece al hijo y está activa
        if nro_tarjeta.hijo_id != hijo.pk:
            raise ValidationError({"error": "La tarjeta no pertenece al estudiante indicado."})
        if nro_tarjeta.estado != "ACTIVA":
            raise ValidationError({
                "error": f"La tarjeta está {nro_tarjeta.get_estado_display().lower()} y no puede usarse para ingresar."
            })

        # Validar restricciones alérgénicas del hijo
        forzar = self.request.data.get("forzar_restriccion", False)
        advertencias = validar_restricciones_alergenicas(hijo, forzar=bool(forzar))

        # Validar limite de 2 registros por dia
        es_primer_registro = validar_limite_registros_diarios(hijo, fecha_consumo)

        # Validar suscripcion si se provee
        if suscripcion and suscripcion.estado != SuscripcionAlmuerzo.Estado.ACTIVA:
            raise ValidationError({
                "error": "La suscripcion no esta activa",
                "estado_suscripcion": suscripcion.estado,
            })

        # Determinar costo. El almuerzo es una cuenta corriente: nunca se
        # bloquea el registro por saldo — puede quedar negativo.
        if es_primer_registro:
            precio_obj = get_precio_almuerzo_activo(fecha_consumo)
            if precio_obj:
                costo_calculado = precio_obj.precio_unitario
            elif tipo_almuerzo:
                costo_calculado = tipo_almuerzo.precio_unitario
            else:
                raise ValidationError({
                    "error": "No hay precio de almuerzo configurado. Configure un precio vigente primero."
                })
        else:
            costo_calculado = Decimal("0")

        with transaction.atomic():
            if es_primer_registro:
                # Asegurar la cuenta ANTES de crear el registro: el trigger
                # trg_sync_cuenta_almuerzo (migración 0014) sincroniza
                # cantidad_almuerzos/monto_total al insertar, pero solo si la
                # fila de la cuenta ya existe.
                self._asegurar_cuenta_mensual(hijo, fecha_consumo)

            registro = serializer.save(
                costo_almuerzo=costo_calculado,
                ya_cobrado=es_primer_registro,
                estado=RegistroConsumoAlmuerzo.Estado.REGISTRADO,
                registrado_por=self.request.user,
            )

            if es_primer_registro:
                self._marcar_acreditado(registro)
                AlmuerzoService._debitar_saldo_almuerzo(registro)

        if es_primer_registro:
            AlmuerzoService._notificar_ingreso_comedor(registro)

        return advertencias

    def _asegurar_cuenta_mensual(self, hijo, fecha):
        """Crea la cuenta mensual del alumno si todavía no existe.

        No suma cantidad_almuerzos/monto_total acá: eso lo hace el trigger
        trg_sync_cuenta_almuerzo al insertar el RegistroConsumoAlmuerzo.
        """
        CuentaAlmuerzoMensual.objects.get_or_create(
            hijo=hijo,
            anio=fecha.year,
            mes=fecha.month,
            defaults={
                "cantidad_almuerzos": 0,
                "monto_total": 0,
                "monto_pagado": 0,
                "forma_cobro": CuentaAlmuerzoMensual.FormaCobro.EFECTIVO,
                "estado": CuentaAlmuerzoMensual.Estado.PENDIENTE,
            },
        )

    def _marcar_acreditado(self, registro):
        """Marca el registro como incorporado y recalcula el estado de la
        cuenta (el trigger ya sincronizó cantidad_almuerzos/monto_total)."""
        fecha = registro.fecha_consumo
        cuenta = CuentaAlmuerzoMensual.objects.select_for_update().get(
            hijo=registro.hijo, anio=fecha.year, mes=fecha.month,
        )
        # Recalcular estado: monto_total cambió, puede haber salido de PAGADO
        cuenta.refresh_from_db(fields=["monto_total", "monto_pagado", "estado", "fecha_pago"])
        cuenta.actualizar_estado()

        # Si no se marca, cerrar_cuentas_mes_anterior lo vuelve a sumar al
        # cerrar el mes (el trigger ya lo contó al crearlo).
        registro.marcado_en_cuenta = True
        registro.save(update_fields=["marcado_en_cuenta"])


# ==============================================================================
# CUENTA ALMUERZO MENSUAL
# ==============================================================================

class CuentaAlmuerzoMensualFilter(django_filters.FilterSet):
    estado = _EstadoInFilter(field_name="estado", lookup_expr="in")

    class Meta:
        model = CuentaAlmuerzoMensual
        fields = ["hijo", "anio", "mes", "estado"]


class CuentaAlmuerzoMensualViewSet(viewsets.ModelViewSet):
    queryset = CuentaAlmuerzoMensual.objects.select_related("hijo__grado", "hijo__tarjeta").all()
    serializer_class = CuentaAlmuerzoMensualSerializer
    permission_classes = [IsStaffOrClienteWeb]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = CuentaAlmuerzoMensualFilter
    ordering = ["-anio", "-mes"]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if hasattr(user, "rol") and user.rol == "CLIENTE_WEB":
            qs = qs.filter(hijo__cliente_responsable=user.cliente)
        return qs

    @action(detail=False, methods=["post"], url_path="generar", permission_classes=[IsAdmin])
    def generar(self, request):
        """
        POST /api/almuerzos/cuentas-mensuales/generar/
        Body: {anio: 2026, mes: 5}
        Genera cuentas para todas las suscripciones activas del mes indicado.
        """
        class _Serializer(drf_serializers.Serializer):
            anio = drf_serializers.IntegerField(min_value=2020, max_value=2099)
            mes = drf_serializers.IntegerField(min_value=1, max_value=12)

        serializer = _Serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        anio = serializer.validated_data["anio"]
        mes = serializer.validated_data["mes"]

        suscripciones = SuscripcionAlmuerzo.objects.filter(
            estado=SuscripcionAlmuerzo.Estado.ACTIVA
        ).select_related("hijo")

        creadas = 0
        for suscripcion in suscripciones:
            _, fue_creada = CuentaAlmuerzoMensual.objects.get_or_create(
                hijo=suscripcion.hijo,
                anio=anio,
                mes=mes,
                defaults={
                    "cantidad_almuerzos": 0,
                    "monto_total": 0,
                    "monto_pagado": 0,
                    "forma_cobro": CuentaAlmuerzoMensual.FormaCobro.EFECTIVO,
                    "estado": CuentaAlmuerzoMensual.Estado.PENDIENTE,
                },
            )
            if fue_creada:
                creadas += 1

        return Response({
            "cuentas_creadas": creadas,
            "mes": mes,
            "anio": anio,
        })


# ==============================================================================
# PAGO CUENTA ALMUERZO
# ==============================================================================

class PagoCuentaAlmuerzoViewSet(viewsets.ModelViewSet):
    """
    Solo lectura: CuentaAlmuerzoMensual dejó de ser una cuenta cobrable en
    paralelo a SaldoAlmuerzo (dos sistemas cobrando el mismo almuerzo por
    separado). El pago de almuerzo pasa exclusivamente por
    RecargaSaldoAlmuerzo (/almuerzos/recargas-saldo/); los 41 registros
    históricos de PagoCuentaAlmuerzo se conservan solo para reporte.
    """
    http_method_names = ["get", "head", "options"]
    queryset = PagoCuentaAlmuerzo.objects.select_related("cuenta").all()
    serializer_class = PagoCuentaAlmuerzoSerializer
    permission_classes = [IsCajeroOrAdmin]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ["cuenta"]
    ordering = ["-fecha_pago"]


# ==============================================================================
# SALDO DE ALMUERZO (CUENTA CORRIENTE)
# ==============================================================================

METODOS_CONFIRMACION_INMEDIATA = ("EFECTIVO", "POS DEBITO", "POS CREDITO")

# Mismo tope que Bancard (core/bancard_views.py) — evitar cargas por caja sin límite.
MONTO_MIN_CARGA_CAJA = 5_000
MONTO_MAX_CARGA_CAJA = 5_000_000


class SaldoAlmuerzoViewSet(viewsets.ReadOnlyModelViewSet):
    """Saldo corriente de almuerzo por hijo. Solo lectura — se modifica vía
    RecargaSaldoAlmuerzoViewSet y RegistroConsumoAlmuerzoViewSet."""

    queryset = SaldoAlmuerzo.objects.select_related("hijo__grado", "hijo__cliente_responsable").all()
    serializer_class = SaldoAlmuerzoSerializer
    permission_classes = [IsStaffOrClienteWeb]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ["hijo"]
    ordering_fields = ["saldo_actual", "fecha_actualizacion"]
    ordering = ["-fecha_actualizacion"]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.rol == "CLIENTE_WEB":
            if not user.cliente:
                return qs.none()
            return qs.filter(hijo__cliente_responsable=user.cliente)
        return qs

    @action(detail=True, methods=["get"], url_path="movimientos")
    def movimientos(self, request, pk=None):
        """GET /api/v1/almuerzos/saldos/<id>/movimientos/ — historial del saldo."""
        saldo = self.get_object()
        movimientos = saldo.movimientos.order_by("-fecha")[:100]
        return Response(MovimientoSaldoAlmuerzoSerializer(movimientos, many=True).data)


class RecargaSaldoAlmuerzoViewSet(viewsets.ModelViewSet):
    """Recarga del saldo corriente de almuerzo — cajero, cobrador, admin o portal."""

    queryset = RecargaSaldoAlmuerzo.objects.select_related("hijo", "registrado_por").all()
    serializer_class = RecargaSaldoAlmuerzoSerializer
    permission_classes = [IsCajeroCobradorOrAdmin]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ["hijo", "estado"]
    ordering = ["-fecha_carga"]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        metodo = data.get("metodo_pago", "")

        if metodo in METODOS_CONFIRMACION_INMEDIATA:
            monto_cargado = data["monto_cargado"]
            if monto_cargado < MONTO_MIN_CARGA_CAJA:
                return Response(
                    {"error": f"El monto mínimo de carga es ₲{MONTO_MIN_CARGA_CAJA:,.0f}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if monto_cargado > MONTO_MAX_CARGA_CAJA:
                return Response(
                    {"error": f"El monto máximo de carga es ₲{MONTO_MAX_CARGA_CAJA:,.0f}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            from apps.contabilidad.models import CierreCaja
            cierre_caja = CierreCaja.objects.filter(
                empleado=request.user, estado=CierreCaja.Estado.ABIERTO
            ).select_related("caja").first()
            medio_pago_obj = resolver_medio_pago(metodo)
            nro_factura = (request.data.get("nro_factura") or "").strip()

            with transaction.atomic():
                recarga = AlmuerzoService.recargar_saldo(
                    hijo=data["hijo"],
                    monto=data["monto_cargado"],
                    registrado_por=request.user,
                    metodo_pago=metodo,
                    referencia=data.get("referencia") or "",
                    cierre_caja=cierre_caja,
                    medio_pago_obj=medio_pago_obj,
                )
                if nro_factura:
                    from apps.contabilidad.services import FacturacionService
                    FacturacionService.emitir_para_origen(
                        tipo="RECARGA_ALMUERZO",
                        origen_id=recarga.id,
                        nro_factura=nro_factura,
                    )
            registrar_auditoria(
                request=request,
                operacion="RECARGA_SALDO_ALMUERZO",
                tabla="almuerzos_recargasaldoalmuerzo",
                id_registro=recarga.id,
                descripcion=(
                    f"Recarga {data['monto_cargado']} Gs. en saldo de almuerzo"
                    f" de {data['hijo']} vía {metodo}"
                ),
            )
            out = self.get_serializer(recarga)
            return Response(out.data, status=status.HTTP_201_CREATED)

        if metodo == "CUENTA_CORRIENTE":
            from apps.clientes.models import CuentaCorrienteCliente

            hijo = data["hijo"]
            cliente = hijo.cliente_responsable

            if not cliente.permite_cuenta_corriente:
                return Response(
                    {"error": "El cliente no tiene habilitada la cuenta corriente."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            with transaction.atomic():
                # limite_credito=0 → sin límite; >0 → tope máximo de deuda acumulada.
                if cliente.limite_credito:
                    ultimo_cc = (
                        CuentaCorrienteCliente.objects
                        .filter(cliente=cliente)
                        .select_for_update()
                        .order_by("-id")
                        .first()
                    )
                    saldo_anterior_cc = ultimo_cc.saldo_resultante if ultimo_cc else Decimal("0")
                    if saldo_anterior_cc + data["monto_cargado"] > cliente.limite_credito:
                        return Response(
                            {
                                "error": f"La recarga excede el límite de crédito autorizado (₲{cliente.limite_credito:,.0f}).",
                                "limite_credito": str(cliente.limite_credito),
                                "saldo_deudor": str(saldo_anterior_cc),
                            },
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                recarga = AlmuerzoService.recargar_saldo(
                    hijo=hijo,
                    monto=data["monto_cargado"],
                    registrado_por=request.user,
                    metodo_pago=metodo,
                    referencia=data.get("referencia") or "",
                )
                CuentaCorrienteCliente.objects.create(
                    cliente=cliente,
                    tipo=CuentaCorrienteCliente.Tipo.DEBITO,
                    monto=data["monto_cargado"],
                    descripcion=f"Recarga almuerzo {hijo.nombre_completo}",
                    creado_por=request.user,
                )
            registrar_auditoria(
                request=request,
                operacion="RECARGA_SALDO_ALMUERZO",
                tabla="almuerzos_recargasaldoalmuerzo",
                id_registro=recarga.id,
                descripcion=(
                    f"Recarga {data['monto_cargado']} Gs. en saldo de almuerzo"
                    f" de {hijo} vía cuenta corriente"
                ),
            )
            out = self.get_serializer(recarga)
            return Response(out.data, status=status.HTTP_201_CREATED)

        # Transferencia u otros métodos: queda PENDIENTE para confirmación manual
        recarga = serializer.save(registrado_por=request.user)
        registrar_auditoria(
            request=request,
            operacion="RECARGA_SALDO_ALMUERZO_PENDIENTE",
            tabla="almuerzos_recargasaldoalmuerzo",
            id_registro=recarga.id,
            descripcion=(
                f"Recarga {recarga.monto_cargado} Gs. en saldo de almuerzo"
                f" de hijo={recarga.hijo_id} vía {metodo} — PENDIENTE confirmación"
            ),
        )
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=True, methods=["post"], url_path="confirmar")
    def confirmar(self, request, pk=None):
        """POST /api/v1/almuerzos/recargas-saldo/<id>/confirmar/ — confirma una recarga PENDIENTE."""
        recarga = self.get_object()
        if recarga.estado != RecargaSaldoAlmuerzo.Estado.PENDIENTE:
            return Response(
                {"error": f"Solo se pueden confirmar recargas PENDIENTE. Estado actual: {recarga.estado}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        from apps.contabilidad.models import CierreCaja
        cierre_caja = CierreCaja.objects.filter(
            empleado=request.user, estado=CierreCaja.Estado.ABIERTO
        ).select_related("caja").first()
        medio_pago_obj = resolver_medio_pago(recarga.metodo_pago)
        nro_factura = (request.data.get("nro_factura") or "").strip()
        with transaction.atomic():
            recarga_confirmada = AlmuerzoService.confirmar_recarga(
                recarga=recarga,
                cierre_caja=cierre_caja,
                medio_pago_obj=medio_pago_obj,
            )
            if nro_factura:
                from apps.contabilidad.services import FacturacionService
                FacturacionService.emitir_para_origen(
                    tipo="RECARGA_ALMUERZO",
                    origen_id=recarga_confirmada.id,
                    nro_factura=nro_factura,
                )
        registrar_auditoria(
            request=request,
            operacion="CONFIRMAR_RECARGA_ALMUERZO",
            tabla="almuerzos_recargasaldoalmuerzo",
            id_registro=recarga_confirmada.id,
            descripcion=(
                f"Confirmación recarga {recarga_confirmada.monto_cargado} Gs."
                f" en saldo de almuerzo de hijo={recarga_confirmada.hijo_id}"
            ),
        )
        return Response(self.get_serializer(recarga_confirmada).data)


# ==============================================================================
# ALERGENO
# ==============================================================================

class AlergenoViewSet(viewsets.ModelViewSet):
    queryset = Alergeno.objects.all()
    serializer_class = AlergenoSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ["activo", "severidad"]
    search_fields = ["nombre"]


# ==============================================================================
# PRODUCTO ALERGENO
# ==============================================================================

class ProductoAlergenoViewSet(viewsets.ModelViewSet):
    queryset = ProductoAlergeno.objects.select_related("producto", "alergeno").all()
    serializer_class = ProductoAlergenoSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["producto", "alergeno", "contiene"]


# ==============================================================================
# MENÚ DIARIO
# ==============================================================================

class MenuDiarioViewSet(viewsets.ModelViewSet):
    """
    CRUD de menú del día. Staff puede crear/editar; CLIENTE_WEB puede leer.
    GET /api/almuerzos/menu/?fecha=YYYY-MM-DD
    GET /api/almuerzos/menu/hoy/  → menú del día actual
    """
    queryset = MenuDiario.objects.filter(activo=True)
    serializer_class = MenuDiarioSerializer
    permission_classes = [IsStaffOrClienteWeb]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ["fecha", "activo"]
    ordering = ["-fecha"]

    def perform_create(self, serializer):
        serializer.save(creado_por=self.request.user)

    @action(detail=False, methods=["get"], url_path="hoy")
    def hoy(self, request):
        """Retorna el menú del día actual (fecha de hoy)."""
        menu = MenuDiario.objects.filter(fecha=date.today(), activo=True).first()
        if menu is None:
            return Response({"detail": "No hay menú publicado para hoy."}, status=404)
        return Response(MenuDiarioSerializer(menu).data)


# ==============================================================================
# DETALLE MENÚ DIARIO
# ==============================================================================

class DetalleMenuDiarioViewSet(viewsets.ModelViewSet):
    """
    CRUD de ítems de un menú diario.
    GET  /api/almuerzos/detalle-menu/?menu={id}   → ítems del menú
    GET  /api/almuerzos/detalle-menu/?menu={id}&curso=PLATO_PRINCIPAL
    POST /api/almuerzos/detalle-menu/             → agregar ítem (staff)
    PATCH/DELETE /api/almuerzos/detalle-menu/{id}/
    """
    serializer_class = DetalleMenuDiarioSerializer
    permission_classes = [IsStaffOrClienteWeb]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["menu", "curso", "es_opcional"]

    def get_queryset(self):
        return (
            DetalleMenuDiario.objects
            .select_related("producto", "producto__unidad_medida")
            .order_by("curso", "producto__descripcion")
        )

    def get_permissions(self):
        if self.action in ("list", "retrieve"):
            return [IsStaffOrClienteWeb()]
        return [IsAdminOrReadOnly()]


# ==============================================================================
# REPORTE DE ALMUERZOS
# ==============================================================================

class ReporteAlmuerzosView(APIView):
    """
    GET /api/almuerzos/reportes/?anio=2026&mes=5
    Parámetros opcionales: hijo=<id>, grado=<str>, formato=csv
    Retorna resumen por hijo: cantidad de almuerzos, monto total, pendiente.
    """
    permission_classes = [IsStaffUser]

    def get(self, request):
        from django.http import HttpResponse

        anio = request.query_params.get("anio")
        mes = request.query_params.get("mes")
        hijo_id = request.query_params.get("hijo")
        grado = request.query_params.get("grado")

        if not anio or not mes:
            return Response(
                {"error": "Se requieren los parámetros anio y mes."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        tarjeta_filter = request.query_params.get("tarjeta")

        qs = CuentaAlmuerzoMensual.objects.filter(
            anio=anio, mes=mes
        ).select_related("hijo__grado", "hijo__tarjeta")

        if hijo_id:
            qs = qs.filter(hijo_id=hijo_id)
        if grado:
            qs = qs.filter(hijo__grado__nombre__icontains=grado)
        if tarjeta_filter:
            qs = qs.filter(hijo__tarjeta__nro_tarjeta__icontains=tarjeta_filter)

        filas = []
        for c in qs.order_by("hijo__apellido", "hijo__nombre"):
            tarjeta = getattr(c.hijo, "tarjeta", None)
            filas.append({
                "hijo_id": c.hijo_id,
                "hijo": c.hijo.nombre_completo,
                "grado": c.hijo.grado.nombre if c.hijo.grado else "",
                "nro_tarjeta": tarjeta.nro_tarjeta if tarjeta else "",
                "cantidad_almuerzos": c.cantidad_almuerzos,
                "monto_total": int(c.monto_total),
                "monto_pagado": int(c.monto_pagado),
                "monto_pendiente": int(c.monto_total - c.monto_pagado),
                "estado": c.estado,
            })

        totales = {
            "cantidad_almuerzos": sum(f["cantidad_almuerzos"] for f in filas),
            "monto_total": sum(f["monto_total"] for f in filas),
            "monto_pagado": sum(f["monto_pagado"] for f in filas),
            "monto_pendiente": sum(f["monto_pendiente"] for f in filas),
            "alumnos": len(filas),
            "con_deuda": sum(1 for f in filas if f["monto_pendiente"] > 0),
        }

        if request.query_params.get("formato") == "csv":
            resp = HttpResponse(content_type="text/csv; charset=utf-8-sig")
            resp["Content-Disposition"] = (
                f'attachment; filename="almuerzos_{anio}_{mes}.csv"'
            )
            writer = csv.writer(resp)
            writer.writerow(["REPORTE DE ALMUERZOS", f"{mes}/{anio}"])
            writer.writerow([])
            writer.writerow(["Alumno", "Grado", "Almuerzos", "Total (Gs)", "Pagado (Gs)", "Pendiente (Gs)", "Estado"])
            for f in filas:
                writer.writerow([f["hijo"], f["grado"], f["cantidad_almuerzos"],
                                  f["monto_total"], f["monto_pagado"], f["monto_pendiente"], f["estado"]])
            writer.writerow([])
            writer.writerow(["TOTALES", "", totales["cantidad_almuerzos"],
                              totales["monto_total"], totales["monto_pagado"], totales["monto_pendiente"], ""])
            return resp

        return Response({
            "periodo": {"anio": int(anio), "mes": int(mes)},
            "totales": totales,
            "filas": filas,
        })


class ReporteConsumoGradoView(APIView):
    """
    GET /api/almuerzos/reporte-consumo-grado/?desde=YYYY-MM-DD&hasta=YYYY-MM-DD
    Consumos agrupados por grado: cantidad, tasa de rechazo y distribución horaria.
    Opcional: ?formato=csv
    """
    permission_classes = [IsStaffUser]

    def get(self, request):
        from django.http import HttpResponse
        from django.db.models import Count, Sum, Q
        from django.db.models.functions import ExtractHour

        desde = request.query_params.get("desde")
        hasta = request.query_params.get("hasta")
        if not desde or not hasta:
            return Response(
                {"error": "Se requieren los parámetros desde y hasta (YYYY-MM-DD)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        base_qs = RegistroConsumoAlmuerzo.objects.filter(
            fecha_consumo__gte=desde,
            fecha_consumo__lte=hasta,
        )

        # Agrupación por grado
        por_grado_qs = (
            base_qs
            .values(
                "hijo__grado__nombre",
                "hijo__grado__nivel",
                "hijo__grado__orden",
            )
            .annotate(
                n_consumos=Count("id", filter=Q(estado="REGISTRADO")),
                n_rechazados=Count("id", filter=Q(estado="RECHAZADO")),
                n_anulados=Count("id", filter=Q(estado="ANULADO")),
                monto_total=Sum("costo_almuerzo", filter=Q(estado="REGISTRADO")),
            )
            .order_by("hijo__grado__nivel", "hijo__grado__orden")
        )

        por_grado = []
        for r in por_grado_qs:
            total_registros = (r["n_consumos"] or 0) + (r["n_rechazados"] or 0)
            tasa_rechazo = (
                round(r["n_rechazados"] / total_registros * 100, 1)
                if total_registros > 0 else 0.0
            )
            por_grado.append({
                "grado": r["hijo__grado__nombre"] or "Sin grado",
                "nivel": r["hijo__grado__nivel"],
                "n_consumos": r["n_consumos"] or 0,
                "n_rechazados": r["n_rechazados"] or 0,
                "n_anulados": r["n_anulados"] or 0,
                "tasa_rechazo": tasa_rechazo,
                "monto_total": int(r["monto_total"] or 0),
            })

        # Distribución horaria (solo REGISTRADO)
        horas_qs = (
            base_qs
            .filter(estado="REGISTRADO")
            .annotate(hora=ExtractHour("hora_registro"))
            .values("hora")
            .annotate(n=Count("id"))
            .order_by("hora")
        )
        horarios = [{"hora": r["hora"], "n": r["n"]} for r in horas_qs]

        total_consumos = sum(r["n_consumos"] for r in por_grado)
        total_rechazados = sum(r["n_rechazados"] for r in por_grado)
        total_registros = total_consumos + total_rechazados
        tasa_global = round(total_rechazados / total_registros * 100, 1) if total_registros > 0 else 0.0

        if request.query_params.get("formato") == "csv":
            resp = HttpResponse(content_type="text/csv; charset=utf-8-sig")
            resp["Content-Disposition"] = (
                f'attachment; filename="consumo_grado_{desde}_{hasta}.csv"'
            )
            writer = csv.writer(resp)
            writer.writerow(["CONSUMO POR GRADO", f"{desde} al {hasta}"])
            writer.writerow([])
            writer.writerow(["Grado", "Consumos", "Rechazados", "Anulados", "% Rechazo", "Monto (Gs)"])
            for r in por_grado:
                writer.writerow([r["grado"], r["n_consumos"], r["n_rechazados"],
                                  r["n_anulados"], r["tasa_rechazo"], r["monto_total"]])
            writer.writerow([])
            writer.writerow(["Distribución horaria"])
            writer.writerow(["Hora", "Consumos"])
            for h in horarios:
                writer.writerow([f"{h['hora']:02d}:00", h["n"]])
            return resp

        return Response({
            "periodo": {"desde": desde, "hasta": hasta},
            "resumen": {
                "total_consumos": total_consumos,
                "total_rechazados": total_rechazados,
                "tasa_rechazo_global": tasa_global,
            },
            "por_grado": por_grado,
            "horarios_pico": horarios,
        })


class ReporteCobranzaAlmuerzosView(APIView):
    """
    GET /api/almuerzos/reporte-cobranza/?anio=YYYY
    Cobranza mensual de almuerzos: estado por mes, forma de cobro y tendencia de recupero.
    Opcional: ?formato=csv
    """
    permission_classes = [IsStaffUser]

    def get(self, request):
        from django.http import HttpResponse
        from django.db.models import Count, Sum

        anio_raw = request.query_params.get("anio")
        if not anio_raw:
            return Response(
                {"error": "Se requiere el parámetro anio (YYYY)."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            anio = int(anio_raw)
        except ValueError:
            return Response({"error": "El parámetro anio debe ser un número."}, status=status.HTTP_400_BAD_REQUEST)

        qs = CuentaAlmuerzoMensual.objects.filter(anio=anio).exclude(estado="ANULADO")

        # "Consumido" por mes: sigue viniendo de CuentaAlmuerzoMensual (el
        # trigger trg_sync_cuenta_almuerzo la mantiene sincronizada con los
        # registros reales de comedor).
        consumido_por_mes = {
            r["mes"]: {"n_alumnos": r["n_alumnos"], "monto_total": int(r["monto_total"] or 0)}
            for r in qs.values("mes").annotate(
                n_alumnos=Count("id"), monto_total=Sum("monto_total"),
            )
        }

        # "Cobrado" ya no sale de CuentaAlmuerzoMensual.monto_pagado — desde la
        # cuenta corriente de almuerzo, lo que entra son recargas de saldo, no
        # pagos contra un mes puntual.
        from django.db.models.functions import ExtractMonth
        from .models import RecargaSaldoAlmuerzo
        recargas_qs = RecargaSaldoAlmuerzo.objects.filter(
            estado=RecargaSaldoAlmuerzo.Estado.CONFIRMADA,
            fecha_carga__year=anio,
        )
        cobrado_por_mes = {
            r["mes"]: int(r["total"] or 0)
            for r in (
                recargas_qs
                .annotate(mes=ExtractMonth("fecha_carga"))
                .values("mes")
                .annotate(total=Sum("monto_cargado"))
            )
        }
        por_metodo_qs = recargas_qs.values("metodo_pago").annotate(
            total=Sum("monto_cargado"), n=Count("id"),
        )
        cobrado_por_metodo = {r["metodo_pago"]: int(r["total"] or 0) for r in por_metodo_qs}
        n_recargas_por_metodo = {r["metodo_pago"]: r["n"] for r in por_metodo_qs}

        MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

        meses_con_datos = sorted(set(consumido_por_mes) | set(cobrado_por_mes))
        por_mes = []
        for mes in meses_con_datos:
            mt = consumido_por_mes.get(mes, {}).get("monto_total", 0)
            mc = cobrado_por_mes.get(mes, 0)
            tasa = round(mc / mt * 100, 1) if mt > 0 else 0.0
            por_mes.append({
                "mes": mes,
                "mes_nombre": MESES[mes],
                "n_alumnos": consumido_por_mes.get(mes, {}).get("n_alumnos", 0),
                "monto_total": mt,
                "monto_cobrado": mc,
                "monto_pendiente": mt - mc,
                "tasa_cobro": tasa,
            })

        # Por método de pago de las recargas (reemplaza "por forma de cobro")
        por_forma = [
            {
                "forma_cobro": metodo,
                "n_recargas": n_recargas_por_metodo.get(metodo, 0),
                "monto_cobrado": monto,
            }
            for metodo, monto in sorted(cobrado_por_metodo.items(), key=lambda x: -x[1])
        ]

        monto_anual = sum(m["monto_total"] for m in por_mes)
        cobrado_anual = sum(m["monto_cobrado"] for m in por_mes)
        tasa_anual = round(cobrado_anual / monto_anual * 100, 1) if monto_anual > 0 else 0.0

        # Control de saldo pendiente: familias con saldo de almuerzo negativo,
        # sin importar el mes — es una cuenta corriente, no algo mensual.
        from .models import SaldoAlmuerzo
        saldos_negativos = (
            SaldoAlmuerzo.objects.filter(saldo_actual__lt=0)
            .select_related("hijo__grado", "hijo__cliente_responsable")
            .order_by("saldo_actual")
        )
        saldos_pendientes = [
            {
                "hijo_id": s.hijo_id,
                "hijo": str(s.hijo),
                "grado": s.hijo.grado.nombre if s.hijo.grado else None,
                "saldo_actual": int(s.saldo_actual),
            }
            for s in saldos_negativos
        ]

        formato = request.query_params.get("formato")

        if formato == "csv":
            resp = HttpResponse(content_type="text/csv; charset=utf-8-sig")
            resp["Content-Disposition"] = (
                f'attachment; filename="cobranza_almuerzos_{anio}.csv"'
            )
            writer = csv.writer(resp)
            writer.writerow(["COBRANZA ALMUERZOS", str(anio)])
            writer.writerow([])
            writer.writerow(["Mes", "Alumnos", "Monto Total (Gs)", "Cobrado (Gs)",
                              "Pendiente (Gs)", "% Cobro"])
            for m in por_mes:
                writer.writerow([m["mes_nombre"], m["n_alumnos"], m["monto_total"],
                                  m["monto_cobrado"], m["monto_pendiente"], m["tasa_cobro"]])
            writer.writerow([])
            writer.writerow(["TOTAL ANUAL", "", monto_anual, cobrado_anual,
                              monto_anual - cobrado_anual, tasa_anual])
            if saldos_pendientes:
                writer.writerow([])
                writer.writerow(["SALDOS NEGATIVOS ACTUALES (cuenta corriente, no por mes)"])
                writer.writerow(["Alumno", "Grado", "Saldo (Gs)"])
                for s in saldos_pendientes:
                    writer.writerow([s["hijo"], s["grado"], s["saldo_actual"]])
            return resp

        if formato == "excel":
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Cobranza Almuerzos"

            header_fill = PatternFill("solid", fgColor="1E3A5F")
            header_font = Font(bold=True, color="FFFFFF")
            total_font = Font(bold=True)
            totals_fill = PatternFill("solid", fgColor="E8F0FE")

            ws.append([f"COBRANZA ALMUERZOS — {anio}"])
            ws["A1"].font = Font(bold=True, size=13)
            ws.append([])

            headers = ["Mes", "Alumnos", "Monto Total (Gs)", "Cobrado (Gs)",
                       "Pendiente (Gs)", "% Cobro"]
            ws.append(headers)
            for cell in ws[ws.max_row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for m in por_mes:
                ws.append([m["mes_nombre"], m["n_alumnos"], m["monto_total"],
                            m["monto_cobrado"], m["monto_pendiente"], m["tasa_cobro"]])

            ws.append([])
            total_row = ["TOTAL ANUAL", "", monto_anual, cobrado_anual,
                         monto_anual - cobrado_anual, tasa_anual]
            ws.append(total_row)
            for cell in ws[ws.max_row]:
                cell.font = total_font
                cell.fill = totals_fill

            col_widths = [14, 10, 18, 16, 18, 10]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

            if saldos_pendientes:
                ws.append([])
                ws.append(["SALDOS NEGATIVOS ACTUALES (cuenta corriente, no por mes)"])
                ws["A" + str(ws.max_row)].font = Font(bold=True)
                ws.append(["Alumno", "Grado", "Saldo (Gs)"])
                for cell in ws[ws.max_row]:
                    cell.font = header_font
                    cell.fill = header_fill
                for s in saldos_pendientes:
                    ws.append([s["hijo"], s["grado"], s["saldo_actual"]])

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            resp = HttpResponse(
                buf.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            resp["Content-Disposition"] = f'attachment; filename="cobranza_almuerzos_{anio}.xlsx"'
            return resp

        return Response({
            "anio": anio,
            "resumen": {
                "monto_anual": monto_anual,
                "cobrado_anual": cobrado_anual,
                "pendiente_anual": monto_anual - cobrado_anual,
                "tasa_cobro_anual": tasa_anual,
                "meses_con_datos": len(por_mes),
            },
            "por_mes": por_mes,
            "por_forma_cobro": por_forma,
            "saldos_pendientes": saldos_pendientes,
        })
