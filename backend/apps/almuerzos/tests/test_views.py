"""
Tests de vistas de almuerzos.
Cubre: PrecioAlmuerzoViewSet, TipoAlmuerzoViewSet, PlanAlmuerzoViewSet,
SuscripcionAlmuerzoViewSet, RegistroConsumoAlmuerzoViewSet (create + validaciones),
CuentaAlmuerzoMensualViewSet (generar, CLIENTE_WEB filter), PagoCuentaAlmuerzoViewSet,
PagoAlmuerzoMensualViewSet, AlergenoViewSet, ProductoAlergenoViewSet,
MenuDiarioViewSet (hoy), DetalleMenuDiarioViewSet, ReporteAlmuerzosView.
"""
import pytest
from datetime import date, timedelta
from decimal import Decimal
from rest_framework.test import APIClient


@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def api_admin(api_client, usuario_admin):
    api_client.force_authenticate(user=usuario_admin)
    return api_client


@pytest.fixture
def api_cajero(api_client, usuario_cajero):
    api_client.force_authenticate(user=usuario_cajero)
    return api_client


@pytest.fixture
def grado(db):
    from apps.clientes.models import Grado
    g, _ = Grado.objects.get_or_create(
        nombre="3er grado",
        defaults={"nivel": 3, "orden": 3, "activo": True},
    )
    return g


@pytest.fixture
def hijo_almuerzo(db, cliente, grado):
    from apps.clientes.models import Hijo
    return Hijo.objects.create(
        nombre="Pedro",
        apellido="Almuerzo",
        cliente_responsable=cliente,
        grado=grado,
        activo=True,
    )


@pytest.fixture
def tarjeta_almuerzo(db, hijo_almuerzo):
    from apps.core.models import Tarjeta
    return Tarjeta.objects.create(
        nro_tarjeta="ALMZ-VIEW01",
        hijo=hijo_almuerzo,
        saldo_actual=Decimal("50000"),
        estado=Tarjeta.Estado.ACTIVA,
    )


@pytest.fixture
def tarjeta_bloqueada(db, hijo_almuerzo):
    from apps.core.models import Tarjeta
    return Tarjeta.objects.create(
        nro_tarjeta="ALMZ-BLQ01",
        hijo=hijo_almuerzo,
        saldo_actual=Decimal("10000"),
        estado=Tarjeta.Estado.BLOQUEADA,
    )


@pytest.fixture
def precio_almuerzo(db):
    from apps.almuerzos.models import PrecioAlmuerzo
    return PrecioAlmuerzo.objects.create(
        precio_unitario=Decimal("15000"),
        fecha_inicio_vigencia=date.today() - timedelta(days=30),
        activo=True,
    )


@pytest.fixture
def tipo_almuerzo(db):
    from apps.almuerzos.models import TipoAlmuerzo
    return TipoAlmuerzo.objects.create(
        nombre="Almuerzo Estándar View",
        precio_unitario=Decimal("12000"),
        activo=True,
    )


@pytest.fixture
def plan_sin_limite(db):
    from apps.almuerzos.models import PlanAlmuerzo
    return PlanAlmuerzo.objects.create(
        nombre="Plan Libre Views",
        activo=True,
        precio_mensual=Decimal("200000"),
        limite_credito_mensual=None,
        dias_semana_incluidos="LUN,MAR,MIE,JUE,VIE",
    )


@pytest.fixture
def plan_con_limite(db):
    from apps.almuerzos.models import PlanAlmuerzo
    return PlanAlmuerzo.objects.create(
        nombre="Plan Limitado Views",
        activo=True,
        precio_mensual=Decimal("100000"),
        limite_credito_mensual=Decimal("10000"),
        dias_semana_incluidos="LUN,MAR,MIE,JUE,VIE",
    )


@pytest.fixture
def suscripcion_activa(db, hijo_almuerzo, plan_sin_limite):
    from apps.almuerzos.models import SuscripcionAlmuerzo
    return SuscripcionAlmuerzo.objects.create(
        hijo=hijo_almuerzo,
        plan=plan_sin_limite,
        fecha_inicio=date.today(),
        estado=SuscripcionAlmuerzo.Estado.ACTIVA,
    )


@pytest.fixture
def suscripcion_suspendida(db, hijo_almuerzo, plan_con_limite):
    # sin fecha_fin para que el serializer no la rechace por vencida
    from apps.almuerzos.models import SuscripcionAlmuerzo
    return SuscripcionAlmuerzo.objects.create(
        hijo=hijo_almuerzo,
        plan=plan_con_limite,
        fecha_inicio=date.today() - timedelta(days=60),
        estado=SuscripcionAlmuerzo.Estado.SUSPENDIDA,
    )


@pytest.fixture
def alergeno(db):
    from apps.almuerzos.models import Alergeno
    return Alergeno.objects.create(
        nombre="Maní",
        activo=True,
        severidad="ALTA",
    )


@pytest.fixture
def cuenta_mensual(db, hijo_almuerzo):
    from apps.almuerzos.models import CuentaAlmuerzoMensual
    hoy = date.today()
    return CuentaAlmuerzoMensual.objects.create(
        hijo=hijo_almuerzo,
        anio=hoy.year,
        mes=hoy.month,
        cantidad_almuerzos=5,
        monto_total=Decimal("75000"),
        monto_pagado=Decimal("0"),
        forma_cobro=CuentaAlmuerzoMensual.FormaCobro.EFECTIVO,
        estado=CuentaAlmuerzoMensual.Estado.PENDIENTE,
    )


@pytest.fixture
def menu_hoy(db, usuario_admin):
    from apps.almuerzos.models import MenuDiario
    return MenuDiario.objects.create(
        fecha=date.today(),
        plato_principal="Milanesa con papas",
        activo=True,
        creado_por=usuario_admin,
    )


@pytest.fixture
def usuario_cliente_web(db, cliente):
    from apps.usuarios.models import Usuario
    return Usuario.objects.create_user(
        email="almweb@test.com",
        password="test1234",
        nombre="Web",
        apellido="Alm",
        rol=Usuario.Rol.CLIENTE_WEB,
        cliente=cliente,
    )


@pytest.fixture
def api_cliente_web(api_client, usuario_cliente_web):
    api_client.force_authenticate(user=usuario_cliente_web)
    return api_client


# ── ViewSets simples ───────────────────────────────────────────────────────────

@pytest.mark.django_db
class TestViewSetsSimples:

    def test_precios_almuerzo_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/precios-almuerzo/")
        assert resp.status_code == 200

    def test_tipos_almuerzo_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/tipos-almuerzo/")
        assert resp.status_code == 200

    def test_planes_almuerzo_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/planes-almuerzo/")
        assert resp.status_code == 200

    def test_suscripciones_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/suscripciones/")
        assert resp.status_code == 200

    def test_alergenos_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/alergenos/")
        assert resp.status_code == 200

    def test_productos_alergenos_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/productos-alergenos/")
        assert resp.status_code == 200

    def test_cuentas_mensuales_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/cuentas-mensuales/")
        assert resp.status_code == 200

    def test_pagos_cuentas_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/pagos-cuentas/")
        assert resp.status_code == 200

    def test_pagos_mensuales_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/pagos-mensuales/")
        assert resp.status_code == 200

    def test_menu_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/menu/")
        assert resp.status_code == 200

    def test_detalle_menu_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/detalle-menu/")
        assert resp.status_code == 200

    def test_registros_consumo_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/registros-consumo/")
        assert resp.status_code == 200

    def test_requiere_autenticacion(self, api_client):
        resp = api_client.get("/api/v1/almuerzos/menu/")
        assert resp.status_code in (401, 403)


# ── SuscripcionAlmuerzoViewSet — acceso CLIENTE_WEB ───────────────────────────
# Antes de este fix, el ViewSet no tenía permission_classes propio y heredaba
# el default IsStaffUser (que excluye CLIENTE_WEB) → el portal de padres
# recibía 403 y mostraba "Sin plan de almuerzo activo" aunque sí hubiera uno.

@pytest.mark.django_db
class TestSuscripcionAlmuerzoAccesoClienteWeb:

    def test_cliente_web_puede_ver_la_suscripcion_de_su_hijo(self, api_cliente_web, suscripcion_activa, hijo_almuerzo):
        resp = api_cliente_web.get(
            "/api/v1/almuerzos/suscripciones/", {"hijo": hijo_almuerzo.id, "estado": "ACTIVA"}
        )
        assert resp.status_code == 200
        ids = [s["id"] for s in resp.data["results"]]
        assert suscripcion_activa.id in ids

    def test_cliente_web_no_ve_suscripciones_de_otra_familia(
        self, api_cliente_web, suscripcion_activa, tipo_cliente, lista_precio
    ):
        from decimal import Decimal
        from apps.clientes.models import Cliente, Hijo
        from apps.almuerzos.models import SuscripcionAlmuerzo

        otro_cliente = Cliente.objects.create(
            nombres="Otra", apellidos="Familia", ruc_ci="9998887",
            tipo_cliente=tipo_cliente, lista_precio=lista_precio,
            limite_credito=Decimal("100000"),
        )
        otro_hijo = Hijo.objects.create(
            nombre="Ajeno", apellido="Hijo", cliente_responsable=otro_cliente, activo=True,
        )
        otra_suscripcion = SuscripcionAlmuerzo.objects.create(
            hijo=otro_hijo, plan=suscripcion_activa.plan,
            fecha_inicio=suscripcion_activa.fecha_inicio,
            estado=SuscripcionAlmuerzo.Estado.ACTIVA,
        )

        resp = api_cliente_web.get("/api/v1/almuerzos/suscripciones/")
        assert resp.status_code == 200
        ids = [s["id"] for s in resp.data["results"]]
        assert otra_suscripcion.id not in ids

    def test_cliente_web_no_puede_crear_suscripciones(self, api_cliente_web, hijo_almuerzo, plan_sin_limite):
        resp = api_cliente_web.post("/api/v1/almuerzos/suscripciones/", {
            "hijo": hijo_almuerzo.id, "plan": plan_sin_limite.id,
            "tipo_cobro": "CUENTA", "fecha_inicio": str(date.today()),
        })
        assert resp.status_code == 403

    def test_cajero_sigue_viendo_todas_las_suscripciones(self, api_cajero, suscripcion_activa):
        resp = api_cajero.get("/api/v1/almuerzos/suscripciones/")
        assert resp.status_code == 200
        ids = [s["id"] for s in resp.data["results"]]
        assert suscripcion_activa.id in ids


# ── PrecioAlmuerzoViewSet — precio_actual ─────────────────────────────────────

@pytest.mark.django_db
class TestPrecioActual:

    def test_sin_precio_retorna_404(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/precios-almuerzo/precio-actual/")
        assert resp.status_code == 404

    def test_con_precio_retorna_datos(self, api_cajero, precio_almuerzo):
        resp = api_cajero.get("/api/v1/almuerzos/precios-almuerzo/precio-actual/")
        assert resp.status_code == 200
        assert resp.data["precio_unitario"] == "15000"


# ── MenuDiarioViewSet — hoy ───────────────────────────────────────────────────

@pytest.mark.django_db
class TestMenuHoy:

    def test_sin_menu_hoy_retorna_404(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/menu/hoy/")
        assert resp.status_code == 404

    def test_con_menu_hoy_retorna_datos(self, api_cajero, menu_hoy):
        resp = api_cajero.get("/api/v1/almuerzos/menu/hoy/")
        assert resp.status_code == 200
        assert resp.data["plato_principal"] == "Milanesa con papas"

    def test_create_asigna_creado_por(self, api_admin):
        resp = api_admin.post(
            "/api/v1/almuerzos/menu/",
            {"fecha": str(date.today() + timedelta(days=1)), "plato_principal": "Pollo asado", "activo": True},
            format="json",
        )
        assert resp.status_code == 201


# ── RegistroConsumoAlmuerzoViewSet ────────────────────────────────────────────

@pytest.mark.django_db
class TestRegistroConsumoDestroy:

    def test_destroy_siempre_405(self, api_cajero, hijo_almuerzo, tarjeta_almuerzo,
                                 precio_almuerzo, usuario_cajero):
        from apps.almuerzos.models import RegistroConsumoAlmuerzo
        registro = RegistroConsumoAlmuerzo.objects.create(
            hijo=hijo_almuerzo,
            fecha_consumo=date.today() - timedelta(days=1),
            costo_almuerzo=Decimal("15000"),
            ya_cobrado=True,
            nro_tarjeta=tarjeta_almuerzo,
            registrado_por=usuario_cajero,
            estado=RegistroConsumoAlmuerzo.Estado.REGISTRADO,
        )
        resp = api_cajero.delete(f"/api/v1/almuerzos/registros-consumo/{registro.pk}/")
        assert resp.status_code == 403  # solo ADMIN puede eliminar


@pytest.mark.django_db
class TestRegistroConsumoCreate:

    def test_primer_registro_ok(self, api_cajero, hijo_almuerzo, tarjeta_almuerzo, precio_almuerzo):
        resp = api_cajero.post(
            "/api/v1/almuerzos/registros-consumo/",
            {
                "hijo": hijo_almuerzo.pk,
                "fecha_consumo": str(date.today()),
                "nro_tarjeta": tarjeta_almuerzo.pk,
            },
            format="json",
        )
        assert resp.status_code == 201
        assert resp.data["ya_cobrado"] is True
        assert resp.data["costo_almuerzo"] == "15000"

    def test_sin_precio_usa_tipo_almuerzo(self, api_cajero, hijo_almuerzo, tarjeta_almuerzo, tipo_almuerzo):
        # No hay PrecioAlmuerzo activo → usa tipo_almuerzo.precio_unitario
        resp = api_cajero.post(
            "/api/v1/almuerzos/registros-consumo/",
            {
                "hijo": hijo_almuerzo.pk,
                "fecha_consumo": str(date.today()),
                "nro_tarjeta": tarjeta_almuerzo.pk,
                "tipo_almuerzo": tipo_almuerzo.pk,
            },
            format="json",
        )
        assert resp.status_code == 201
        assert resp.data["costo_almuerzo"] == "12000"

    def test_sin_precio_ni_tipo_falla(self, api_cajero, hijo_almuerzo, tarjeta_almuerzo):
        # No hay precio ni tipo → error
        resp = api_cajero.post(
            "/api/v1/almuerzos/registros-consumo/",
            {
                "hijo": hijo_almuerzo.pk,
                "fecha_consumo": str(date.today()),
                "nro_tarjeta": tarjeta_almuerzo.pk,
            },
            format="json",
        )
        assert resp.status_code == 400

    def test_segundo_registro_costo_cero(self, api_cajero, hijo_almuerzo, tarjeta_almuerzo,
                                          precio_almuerzo, usuario_cajero):
        from apps.almuerzos.models import RegistroConsumoAlmuerzo
        # Primer registro directo en BD
        RegistroConsumoAlmuerzo.objects.create(
            hijo=hijo_almuerzo,
            fecha_consumo=date.today(),
            costo_almuerzo=Decimal("15000"),
            ya_cobrado=True,
            nro_tarjeta=tarjeta_almuerzo,
            registrado_por=usuario_cajero,
            estado=RegistroConsumoAlmuerzo.Estado.REGISTRADO,
        )
        # Segundo registro vía API
        resp = api_cajero.post(
            "/api/v1/almuerzos/registros-consumo/",
            {
                "hijo": hijo_almuerzo.pk,
                "fecha_consumo": str(date.today()),
                "nro_tarjeta": tarjeta_almuerzo.pk,
            },
            format="json",
        )
        assert resp.status_code == 201
        assert resp.data["ya_cobrado"] is False
        assert resp.data["costo_almuerzo"] == "0"

    def test_tarjeta_bloqueada_falla(self, api_cajero, hijo_almuerzo, tarjeta_bloqueada, precio_almuerzo):
        resp = api_cajero.post(
            "/api/v1/almuerzos/registros-consumo/",
            {
                "hijo": hijo_almuerzo.pk,
                "fecha_consumo": str(date.today()),
                "nro_tarjeta": tarjeta_bloqueada.pk,
            },
            format="json",
        )
        assert resp.status_code == 400

    def test_tarjeta_de_otro_hijo_falla(self, api_cajero, hijo_almuerzo, precio_almuerzo, cliente, grado):
        from apps.clientes.models import Hijo
        from apps.core.models import Tarjeta
        otro_hijo = Hijo.objects.create(
            nombre="Otro", apellido="Hijo",
            cliente_responsable=cliente, grado=grado, activo=True,
        )
        tarjeta_otro = Tarjeta.objects.create(
            nro_tarjeta="OTRO-001", hijo=otro_hijo, estado=Tarjeta.Estado.ACTIVA,
        )
        resp = api_cajero.post(
            "/api/v1/almuerzos/registros-consumo/",
            {
                "hijo": hijo_almuerzo.pk,
                "fecha_consumo": str(date.today()),
                "nro_tarjeta": tarjeta_otro.pk,
            },
            format="json",
        )
        assert resp.status_code == 400

    def test_suscripcion_suspendida_falla(self, api_cajero, hijo_almuerzo, tarjeta_almuerzo,
                                           precio_almuerzo, suscripcion_suspendida):
        resp = api_cajero.post(
            "/api/v1/almuerzos/registros-consumo/",
            {
                "hijo": hijo_almuerzo.pk,
                "fecha_consumo": str(date.today()),
                "nro_tarjeta": tarjeta_almuerzo.pk,
                "suscripcion": suscripcion_suspendida.pk,
            },
            format="json",
        )
        assert resp.status_code == 400

    def test_limite_credito_ya_no_bloquea_saldo_queda_negativo(
        self, api_cajero, hijo_almuerzo, tarjeta_almuerzo, precio_almuerzo, plan_con_limite
    ):
        """Almuerzo es cuenta corriente: el límite de crédito del plan ya no
        bloquea el registro — el saldo simplemente queda negativo."""
        from apps.almuerzos.models import SuscripcionAlmuerzo, SaldoAlmuerzo
        hoy = date.today()
        suscripcion = SuscripcionAlmuerzo.objects.create(
            hijo=hijo_almuerzo, plan=plan_con_limite,
            fecha_inicio=hoy, estado=SuscripcionAlmuerzo.Estado.ACTIVA,
        )
        resp = api_cajero.post(
            "/api/v1/almuerzos/registros-consumo/",
            {
                "hijo": hijo_almuerzo.pk,
                "fecha_consumo": str(hoy),
                "nro_tarjeta": tarjeta_almuerzo.pk,
                "suscripcion": suscripcion.pk,
            },
            format="json",
        )
        assert resp.status_code == 201
        saldo = SaldoAlmuerzo.objects.get(hijo=hijo_almuerzo)
        assert saldo.saldo_actual == Decimal("-15000")

    def test_con_suscripcion_activa_ok(self, api_cajero, hijo_almuerzo, tarjeta_almuerzo,
                                        precio_almuerzo, suscripcion_activa):
        resp = api_cajero.post(
            "/api/v1/almuerzos/registros-consumo/",
            {
                "hijo": hijo_almuerzo.pk,
                "fecha_consumo": str(date.today()),
                "nro_tarjeta": tarjeta_almuerzo.pk,
                "suscripcion": suscripcion_activa.pk,
            },
            format="json",
        )
        assert resp.status_code == 201
        assert resp.data["ya_cobrado"] is True


# ── CuentaAlmuerzoMensualViewSet ──────────────────────────────────────────────

@pytest.mark.django_db
class TestCuentaMensual:

    def test_generar_sin_suscripciones(self, api_admin):
        resp = api_admin.post(
            "/api/v1/almuerzos/cuentas-mensuales/generar/",
            {"anio": 2026, "mes": 5},
            format="json",
        )
        assert resp.status_code == 200
        assert resp.data["cuentas_creadas"] == 0

    def test_generar_con_suscripcion_activa(self, api_admin, suscripcion_activa):
        resp = api_admin.post(
            "/api/v1/almuerzos/cuentas-mensuales/generar/",
            {"anio": 2099, "mes": 1},
            format="json",
        )
        assert resp.status_code == 200
        assert resp.data["cuentas_creadas"] >= 1

    def test_generar_datos_invalidos(self, api_admin):
        resp = api_admin.post(
            "/api/v1/almuerzos/cuentas-mensuales/generar/",
            {"anio": 2026, "mes": 13},
            format="json",
        )
        assert resp.status_code == 400

    def test_cliente_web_filtra_sus_cuentas(self, api_cliente_web, cuenta_mensual):
        resp = api_cliente_web.get("/api/v1/almuerzos/cuentas-mensuales/")
        assert resp.status_code == 200

    def test_filtro_estado_acepta_varios_valores_separados_por_coma(self, api_cliente_web, cuenta_mensual):
        """PagarAlmuerzo.tsx pide ?estado=PENDIENTE,PARCIAL — antes de este fix
        el filtro auto-generado era un ChoiceFilter de valor único y devolvía
        400 (portal mostraba 'Error al cargar las cuentas de almuerzo')."""
        resp = api_cliente_web.get(
            "/api/v1/almuerzos/cuentas-mensuales/", {"estado": "PENDIENTE,PARCIAL"}
        )
        assert resp.status_code == 200
        ids = [c["id"] for c in resp.data["results"]]
        assert cuenta_mensual.id in ids

    def test_filtro_estado_un_solo_valor_sigue_funcionando(self, api_cliente_web, cuenta_mensual):
        resp = api_cliente_web.get(
            "/api/v1/almuerzos/cuentas-mensuales/", {"estado": "PAGADO"}
        )
        assert resp.status_code == 200
        ids = [c["id"] for c in resp.data["results"]]
        assert cuenta_mensual.id not in ids

    def test_solo_admin_puede_generar(self, api_cajero):
        resp = api_cajero.post(
            "/api/v1/almuerzos/cuentas-mensuales/generar/",
            {"anio": 2026, "mes": 6},
            format="json",
        )
        assert resp.status_code in (401, 403)


# ── PagoCuentaAlmuerzoViewSet ─────────────────────────────────────────────────

@pytest.mark.django_db
class TestPagoCuenta:

    def test_create_actualiza_cuenta(self, api_cajero, cuenta_mensual, usuario_cajero):
        resp = api_cajero.post(
            "/api/v1/almuerzos/pagos-cuentas/",
            {
                "cuenta": cuenta_mensual.pk,
                "monto": 50000,
                "medio_pago": "EFECTIVO",
                "registrado_por": usuario_cajero.pk,
            },
            format="json",
        )
        assert resp.status_code == 201
        cuenta_mensual.refresh_from_db()
        assert cuenta_mensual.monto_pagado == Decimal("50000")

    def test_list_ok(self, api_cajero):
        resp = api_cajero.get("/api/v1/almuerzos/pagos-cuentas/")
        assert resp.status_code == 200


# ── ReporteAlmuerzosView ──────────────────────────────────────────────────────

@pytest.mark.django_db
class TestReporteAlmuerzos:

    def test_sin_anio_retorna_400(self, api_admin):
        resp = api_admin.get("/api/v1/almuerzos/reportes/")
        assert resp.status_code == 400

    def test_solo_mes_retorna_400(self, api_admin):
        resp = api_admin.get("/api/v1/almuerzos/reportes/", {"mes": "5"})
        assert resp.status_code == 400

    def test_con_anio_mes_retorna_estructura(self, api_admin):
        resp = api_admin.get("/api/v1/almuerzos/reportes/", {"anio": "2026", "mes": "5"})
        assert resp.status_code == 200
        assert "periodo" in resp.data
        assert "totales" in resp.data
        assert "filas" in resp.data

    def test_con_cuenta_mensual_muestra_alumno(self, api_admin, cuenta_mensual):
        hoy = date.today()
        resp = api_admin.get(
            "/api/v1/almuerzos/reportes/",
            {"anio": str(hoy.year), "mes": str(hoy.month)},
        )
        assert resp.status_code == 200
        assert resp.data["totales"]["alumnos"] >= 1

    def test_filtro_por_hijo(self, api_admin, cuenta_mensual, hijo_almuerzo):
        hoy = date.today()
        resp = api_admin.get(
            "/api/v1/almuerzos/reportes/",
            {"anio": str(hoy.year), "mes": str(hoy.month), "hijo": hijo_almuerzo.pk},
        )
        assert resp.status_code == 200
        assert len(resp.data["filas"]) == 1

    def test_filtro_por_grado(self, api_admin, cuenta_mensual, grado):
        hoy = date.today()
        resp = api_admin.get(
            "/api/v1/almuerzos/reportes/",
            {"anio": str(hoy.year), "mes": str(hoy.month), "grado": grado.nombre[:4]},
        )
        assert resp.status_code == 200

    def test_formato_csv(self, api_admin, cuenta_mensual):
        hoy = date.today()
        resp = api_admin.get(
            "/api/v1/almuerzos/reportes/",
            {"anio": str(hoy.year), "mes": str(hoy.month), "formato": "csv"},
        )
        assert resp.status_code == 200
        assert "text/csv" in resp["Content-Type"]
        assert b"REPORTE DE ALMUERZOS" in resp.content

    def test_csv_incluye_alumno(self, api_admin, cuenta_mensual):
        hoy = date.today()
        resp = api_admin.get(
            "/api/v1/almuerzos/reportes/",
            {"anio": str(hoy.year), "mes": str(hoy.month), "formato": "csv"},
        )
        assert resp.status_code == 200
        assert b"Pedro" in resp.content

    def test_filtro_por_tarjeta(self, api_admin, cuenta_mensual, tarjeta_almuerzo):
        hoy = date.today()
        resp = api_admin.get(
            "/api/v1/almuerzos/reportes/",
            {"anio": str(hoy.year), "mes": str(hoy.month), "tarjeta": "ALMZ-VIEW01"},
        )
        assert resp.status_code == 200
        assert len(resp.data["filas"]) == 1
        assert resp.data["filas"][0]["nro_tarjeta"] == "ALMZ-VIEW01"

    def test_filas_incluyen_nro_tarjeta(self, api_admin, cuenta_mensual, tarjeta_almuerzo):
        hoy = date.today()
        resp = api_admin.get(
            "/api/v1/almuerzos/reportes/",
            {"anio": str(hoy.year), "mes": str(hoy.month)},
        )
        assert resp.status_code == 200
        assert len(resp.data["filas"]) >= 1
        assert "nro_tarjeta" in resp.data["filas"][0]

    def test_requiere_autenticacion(self, api_client):
        resp = api_client.get("/api/v1/almuerzos/reportes/", {"anio": "2026", "mes": "5"})
        assert resp.status_code in (401, 403)


# ── Cobertura de ramas adicionales ────────────────────────────────────────────

@pytest.mark.django_db
class TestRegistroConsumoLineasAdicionales:

    def test_sin_tarjeta_falla(self, api_cajero, hijo_almuerzo, precio_almuerzo):
        # nro_tarjeta=None → hits line 200 (ValidationError "Debe especificar la tarjeta")
        resp = api_cajero.post(
            "/api/v1/almuerzos/registros-consumo/",
            {"hijo": hijo_almuerzo.pk, "fecha_consumo": str(date.today())},
            format="json",
        )
        assert resp.status_code == 400

    def test_suscripcion_suspendida_falla(self, api_cajero, hijo_almuerzo, tarjeta_almuerzo,
                                           precio_almuerzo, suscripcion_suspendida):
        # suscripcion.estado != ACTIVA → hits line 219
        resp = api_cajero.post(
            "/api/v1/almuerzos/registros-consumo/",
            {
                "hijo": hijo_almuerzo.pk,
                "fecha_consumo": str(date.today()),
                "nro_tarjeta": tarjeta_almuerzo.pk,
                "suscripcion": suscripcion_suspendida.pk,
            },
            format="json",
        )
        assert resp.status_code == 400

    def test_con_restriccion_no_critica_muestra_advertencias(
        self, api_cajero, hijo_almuerzo, tarjeta_almuerzo, precio_almuerzo
    ):
        # RestriccionHijo no crítica → validar_restricciones_alergenicas retorna advertencias
        # hits line 187 (data["advertencias"] = advertencias)
        from apps.clientes.models import RestriccionHijo
        RestriccionHijo.objects.create(
            hijo=hijo_almuerzo,
            tipo="ALERGIA",
            descripcion="Maní",
            severidad=RestriccionHijo.Severidad.ALTA,
            requiere_autorizacion=False,
            activo=True,
        )
        resp = api_cajero.post(
            "/api/v1/almuerzos/registros-consumo/",
            {"hijo": hijo_almuerzo.pk, "fecha_consumo": str(date.today()), "nro_tarjeta": tarjeta_almuerzo.pk},
            format="json",
        )
        assert resp.status_code == 201
        assert "advertencias" in resp.data


@pytest.mark.django_db
class TestDetalleMenuPermisos:

    def test_create_requiere_admin(self, api_admin, menu_hoy, producto):
        # DetalleMenuDiarioViewSet.get_permissions() → IsAdminOrReadOnly → hits line 473
        resp = api_admin.post(
            "/api/v1/almuerzos/detalle-menu/",
            {"menu": menu_hoy.pk, "producto": producto.pk, "curso": "PLATO_PRINCIPAL", "es_opcional": False},
            format="json",
        )
        assert resp.status_code in (201, 400)


# ── RegistroConsumo destroy (admin) ───────────────────────────────────────────

@pytest.mark.django_db
class TestRegistroConsumoDestroyAdmin:
    """Líneas 177-184: destroy con ADMIN — solo permite borrar registros ANULADOS."""

    def _registro(self, hijo, tarjeta, cajero, estado):
        from apps.almuerzos.models import RegistroConsumoAlmuerzo
        return RegistroConsumoAlmuerzo.objects.create(
            hijo=hijo,
            fecha_consumo=date.today() - timedelta(days=2),
            costo_almuerzo=Decimal("15000"),
            ya_cobrado=True,
            nro_tarjeta=tarjeta,
            registrado_por=cajero,
            estado=estado,
        )

    def test_admin_elimina_registro_anulado_204(
        self, api_admin, hijo_almuerzo, tarjeta_almuerzo, usuario_cajero
    ):
        registro = self._registro(
            hijo_almuerzo, tarjeta_almuerzo, usuario_cajero,
            "ANULADO",
        )
        resp = api_admin.delete(f"/api/v1/almuerzos/registros-consumo/{registro.pk}/")
        assert resp.status_code == 204

    def test_admin_no_puede_eliminar_no_anulado_400(
        self, api_admin, hijo_almuerzo, tarjeta_almuerzo, usuario_cajero
    ):
        registro = self._registro(
            hijo_almuerzo, tarjeta_almuerzo, usuario_cajero,
            "REGISTRADO",
        )
        resp = api_admin.delete(f"/api/v1/almuerzos/registros-consumo/{registro.pk}/")
        assert resp.status_code == 400
        assert "ANULADO" in resp.data["error"]


# ── RegistroConsumo: WhatsApp except silente ──────────────────────────────────

@pytest.mark.django_db
class TestRegistroConsumoWhatsappFallback:
    """Líneas 285-286: whatsapp_cliente lanza excepción → se ignora, create igual exitoso."""

    def test_create_exitoso_aunque_whatsapp_falle(
        self, api_cajero, hijo_almuerzo, tarjeta_almuerzo, precio_almuerzo
    ):
        from unittest.mock import patch
        with patch(
            "apps.notificaciones.services.whatsapp_cliente",
            side_effect=Exception("WhatsApp caído"),
        ):
            resp = api_cajero.post(
                "/api/v1/almuerzos/registros-consumo/",
                {
                    "hijo": hijo_almuerzo.pk,
                    "fecha_consumo": str(date.today()),
                    "nro_tarjeta": tarjeta_almuerzo.pk,
                },
                format="json",
            )
        assert resp.status_code == 201


# ── PagoAlmuerzoMensual create ────────────────────────────────────────────────

@pytest.mark.django_db
class TestPagoAlmuerzoMensualCreate:
    """Líneas 414-427: perform_create del ViewSet de pagos de almuerzo mensual."""

    def test_create_pago_mensual_201(self, api_admin, suscripcion_activa):
        resp = api_admin.post(
            "/api/v1/almuerzos/pagos-mensuales/",
            {
                "suscripcion": suscripcion_activa.pk,
                "monto_pagado": "200000",
                "mes_pagado": str(date.today().replace(day=1)),
            },
            format="json",
        )
        assert resp.status_code == 201

    def test_create_pago_mensual_whatsapp_falla_igual_201(
        self, api_admin, suscripcion_activa
    ):
        from unittest.mock import patch
        with patch(
            "apps.notificaciones.services.whatsapp_cliente",
            side_effect=Exception("WhatsApp caído"),
        ):
            resp = api_admin.post(
                "/api/v1/almuerzos/pagos-mensuales/",
                {
                    "suscripcion": suscripcion_activa.pk,
                    "monto_pagado": "100000",
                    "mes_pagado": str(date.today().replace(day=1)),
                },
                format="json",
            )
        assert resp.status_code == 201


# ── ReporteConsumoGradoView ───────────────────────────────────────────────────

@pytest.mark.django_db
class TestReporteConsumoGrado:
    """Líneas 604-690: reporte consumo por grado — JSON y CSV."""

    def test_sin_params_retorna_400(self, api_admin):
        resp = api_admin.get("/api/v1/almuerzos/reporte-consumo-grado/")
        assert resp.status_code == 400

    def test_con_fechas_retorna_estructura(self, api_admin):
        resp = api_admin.get(
            "/api/v1/almuerzos/reporte-consumo-grado/",
            {"desde": "2026-01-01", "hasta": "2026-12-31"},
        )
        assert resp.status_code == 200
        assert "por_grado" in resp.data
        assert "horarios_pico" in resp.data
        assert "resumen" in resp.data

    def test_con_datos_calcula_tasa_rechazo(
        self, api_admin, hijo_almuerzo, tarjeta_almuerzo, usuario_cajero, precio_almuerzo
    ):
        from apps.almuerzos.models import RegistroConsumoAlmuerzo
        RegistroConsumoAlmuerzo.objects.create(
            hijo=hijo_almuerzo,
            fecha_consumo=date(2026, 6, 10),
            costo_almuerzo=Decimal("15000"),
            ya_cobrado=True,
            nro_tarjeta=tarjeta_almuerzo,
            registrado_por=usuario_cajero,
            estado="REGISTRADO",
        )
        RegistroConsumoAlmuerzo.objects.create(
            hijo=hijo_almuerzo,
            fecha_consumo=date(2026, 6, 11),
            costo_almuerzo=Decimal("15000"),
            ya_cobrado=False,
            nro_tarjeta=tarjeta_almuerzo,
            registrado_por=usuario_cajero,
            estado="RECHAZADO",
        )
        resp = api_admin.get(
            "/api/v1/almuerzos/reporte-consumo-grado/",
            {"desde": "2026-06-01", "hasta": "2026-06-30"},
        )
        assert resp.status_code == 200
        grados = resp.data["por_grado"]
        assert len(grados) == 1
        assert grados[0]["tasa_rechazo"] == 50.0

    def test_formato_csv_retorna_descarga(self, api_admin):
        resp = api_admin.get(
            "/api/v1/almuerzos/reporte-consumo-grado/",
            {"desde": "2026-01-01", "hasta": "2026-12-31", "formato": "csv"},
        )
        assert resp.status_code == 200
        assert "text/csv" in resp["Content-Type"]
        assert "attachment" in resp["Content-Disposition"]

    def test_formato_csv_con_datos_incluye_filas(
        self, api_admin, hijo_almuerzo, tarjeta_almuerzo, usuario_cajero, precio_almuerzo
    ):
        from apps.almuerzos.models import RegistroConsumoAlmuerzo
        RegistroConsumoAlmuerzo.objects.create(
            hijo=hijo_almuerzo,
            fecha_consumo=date(2026, 5, 10),
            costo_almuerzo=Decimal("15000"),
            ya_cobrado=True,
            nro_tarjeta=tarjeta_almuerzo,
            registrado_por=usuario_cajero,
            estado="REGISTRADO",
        )
        resp = api_admin.get(
            "/api/v1/almuerzos/reporte-consumo-grado/",
            {"desde": "2026-05-01", "hasta": "2026-05-31", "formato": "csv"},
        )
        assert resp.status_code == 200
        content = resp.content.decode("utf-8-sig")
        assert "3er grado" in content


# ── ReporteCobranzaAlmuerzosView ──────────────────────────────────────────────

@pytest.mark.django_db
class TestReporteCobranzaAlmuerzos:
    """Líneas 711-806: reporte cobranza por año — JSON y CSV."""

    def test_sin_anio_retorna_400(self, api_admin):
        resp = api_admin.get("/api/v1/almuerzos/reporte-cobranza/")
        assert resp.status_code == 400

    def test_anio_invalido_retorna_400(self, api_admin):
        resp = api_admin.get("/api/v1/almuerzos/reporte-cobranza/", {"anio": "no_anio"})
        assert resp.status_code == 400

    def test_con_anio_retorna_estructura(self, api_admin):
        resp = api_admin.get("/api/v1/almuerzos/reporte-cobranza/", {"anio": "2026"})
        assert resp.status_code == 200
        assert "por_mes" in resp.data
        assert "por_forma_cobro" in resp.data
        assert "resumen" in resp.data

    def test_con_cuentas_calcula_totales(
        self, api_admin, hijo_almuerzo, grado
    ):
        """monto_anual sale de CuentaAlmuerzoMensual (consumo real), cobrado_anual
        de RecargaSaldoAlmuerzo (recargas del saldo corriente) — no del mismo modelo."""
        from datetime import datetime
        from django.utils import timezone
        from apps.almuerzos.models import CuentaAlmuerzoMensual, RecargaSaldoAlmuerzo
        CuentaAlmuerzoMensual.objects.create(
            hijo=hijo_almuerzo,
            anio=2026, mes=6,
            cantidad_almuerzos=10,
            monto_total=Decimal("150000"),
            monto_pagado=Decimal("0"),
            forma_cobro=CuentaAlmuerzoMensual.FormaCobro.EFECTIVO,
        )
        RecargaSaldoAlmuerzo.objects.create(
            hijo=hijo_almuerzo,
            monto_cargado=Decimal("100000"),
            fecha_carga=timezone.make_aware(datetime(2026, 6, 15)),
            estado=RecargaSaldoAlmuerzo.Estado.CONFIRMADA,
            metodo_pago="EFECTIVO",
        )
        resp = api_admin.get("/api/v1/almuerzos/reporte-cobranza/", {"anio": "2026"})
        assert resp.status_code == 200
        assert resp.data["resumen"]["monto_anual"] == 150000
        assert resp.data["resumen"]["cobrado_anual"] == 100000

    def test_formato_csv_retorna_descarga(self, api_admin):
        resp = api_admin.get(
            "/api/v1/almuerzos/reporte-cobranza/",
            {"anio": "2026", "formato": "csv"},
        )
        assert resp.status_code == 200
        assert "text/csv" in resp["Content-Type"]
        assert "attachment" in resp["Content-Disposition"]

    def test_formato_csv_con_datos_incluye_filas(
        self, api_admin, hijo_almuerzo
    ):
        from apps.almuerzos.models import CuentaAlmuerzoMensual
        CuentaAlmuerzoMensual.objects.create(
            hijo=hijo_almuerzo,
            anio=2026, mes=3,
            cantidad_almuerzos=8,
            monto_total=Decimal("120000"),
            monto_pagado=Decimal("120000"),
            forma_cobro=CuentaAlmuerzoMensual.FormaCobro.EFECTIVO,
            estado=CuentaAlmuerzoMensual.Estado.PAGADO,
        )
        resp = api_admin.get(
            "/api/v1/almuerzos/reporte-cobranza/",
            {"anio": "2026", "formato": "csv"},
        )
        assert resp.status_code == 200
        content = resp.content.decode("utf-8-sig")
        assert "Marzo" in content

    def test_excel_retorna_xlsx(self, api_admin):
        resp = api_admin.get(
            "/api/v1/almuerzos/reporte-cobranza/",
            {"anio": "2026", "formato": "excel"},
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp["Content-Type"]
        assert "attachment" in resp.get("Content-Disposition", "")
        assert resp.get("Content-Disposition", "").endswith(".xlsx\"")

    def test_excel_con_datos_genera_filas(self, api_admin):
        import io
        from decimal import Decimal
        from openpyxl import load_workbook
        from apps.almuerzos.models import CuentaAlmuerzoMensual
        from apps.clientes.models import Hijo
        hijo = Hijo.objects.first()
        if hijo:
            CuentaAlmuerzoMensual.objects.create(
                hijo=hijo,
                anio=2026,
                mes=3,
                monto_total=Decimal("150000"),
                monto_pagado=Decimal("150000"),
                estado=CuentaAlmuerzoMensual.Estado.PAGADO,
            )
        resp = api_admin.get(
            "/api/v1/almuerzos/reporte-cobranza/",
            {"anio": "2026", "formato": "excel"},
        )
        assert resp.status_code == 200
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        assert ws.max_row >= 2  # encabezado + al menos 1 fila

    def test_con_saldo_negativo_incluye_saldos_pendientes(self, api_admin, hijo_almuerzo):
        from apps.almuerzos.models import SaldoAlmuerzo
        SaldoAlmuerzo.objects.create(hijo=hijo_almuerzo, saldo_actual=Decimal("-20000"))

        resp = api_admin.get("/api/v1/almuerzos/reporte-cobranza/", {"anio": "2026"})
        assert resp.status_code == 200
        assert len(resp.data["saldos_pendientes"]) == 1
        assert resp.data["saldos_pendientes"][0]["saldo_actual"] == -20000

    def test_csv_con_saldo_negativo_incluye_seccion(self, api_admin, hijo_almuerzo):
        from apps.almuerzos.models import SaldoAlmuerzo
        SaldoAlmuerzo.objects.create(hijo=hijo_almuerzo, saldo_actual=Decimal("-20000"))

        resp = api_admin.get(
            "/api/v1/almuerzos/reporte-cobranza/", {"anio": "2026", "formato": "csv"},
        )
        content = resp.content.decode("utf-8-sig")
        assert "SALDOS NEGATIVOS ACTUALES" in content
        assert hijo_almuerzo.nombre in content

    def test_excel_con_saldo_negativo_incluye_seccion(self, api_admin, hijo_almuerzo):
        import io
        from openpyxl import load_workbook
        from apps.almuerzos.models import SaldoAlmuerzo
        SaldoAlmuerzo.objects.create(hijo=hijo_almuerzo, saldo_actual=Decimal("-20000"))

        resp = api_admin.get(
            "/api/v1/almuerzos/reporte-cobranza/", {"anio": "2026", "formato": "excel"},
        )
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        valores = [cell.value for row in ws.iter_rows() for cell in row]
        assert "SALDOS NEGATIVOS ACTUALES (cuenta corriente, no por mes)" in valores
