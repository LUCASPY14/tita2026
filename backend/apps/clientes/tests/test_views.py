"""
Tests de vistas de clientes.
Cubre: HijoViewSet (CLIENTE_WEB filter),
RestriccionHijoViewSet, AlumnoResponsableViewSet (perform_create, destroy, set_titular),
ReporteCuentaCorrienteView, y ViewSets simples.
"""
import pytest
from decimal import Decimal
from rest_framework.test import APIClient


@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def api_admin(api_client, usuario_admin):
    api_client.force_authenticate(user=usuario_admin)
    return api_client


@pytest.fixture
def api_cajero(api_client, usuario_cajero):
    api_client.force_authenticate(user=usuario_cajero)
    return api_client


@pytest.fixture
def hijo_fixture(db, cliente):
    from apps.clientes.models import Hijo
    return Hijo.objects.create(
        nombre="Pedro",
        apellido="López",
        cliente_responsable=cliente,
    )


@pytest.fixture
def alumno_responsable(db, hijo_fixture, cliente, usuario_admin):
    from apps.clientes.models import AlumnoResponsable
    return AlumnoResponsable.objects.create(
        hijo=hijo_fixture,
        cliente=cliente,
        parentesco="PADRE",
        es_titular=True,
        orden_cobro=1,
        agregado_por=usuario_admin,
    )


@pytest.fixture
def usuario_cliente_web(db, cliente):
    from apps.usuarios.models import Usuario
    return Usuario.objects.create_user(
        email="web@clientes.test",
        password="test1234",
        nombre="Web",
        apellido="Cliente",
        rol=Usuario.Rol.CLIENTE_WEB,
        cliente=cliente,
    )


@pytest.fixture
def api_cliente_web(api_client, usuario_cliente_web):
    api_client.force_authenticate(user=usuario_cliente_web)
    return api_client


@pytest.fixture
def cuenta_con_deuda(db, cliente, usuario_cajero):
    from apps.clientes.models import CuentaCorrienteCliente
    return CuentaCorrienteCliente.objects.create(
        cliente=cliente,
        tipo=CuentaCorrienteCliente.Tipo.DEBITO,
        monto=Decimal("50000"),
        saldo_anterior=Decimal("0"),
        saldo_resultante=Decimal("50000"),
        descripcion="Venta fiado test",
        creado_por=usuario_cajero,
    )


# ── ViewSets simples ──────────────────────────────────────────────────────────

@pytest.mark.django_db
class TestViewSetsSimples:

    def test_clientes_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/clientes/clientes/")
        assert resp.status_code == 200

    def test_tipos_cliente_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/clientes/tipos-cliente/")
        assert resp.status_code == 200

    def test_hijos_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/clientes/hijos/")
        assert resp.status_code == 200

    def test_grados_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/clientes/grados/")
        assert resp.status_code == 200

    def test_historial_grados_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/clientes/historial-grados/")
        assert resp.status_code == 200

    def test_restricciones_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/clientes/restricciones/")
        assert resp.status_code == 200

    def test_autorizaciones_saldo_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/clientes/autorizaciones-saldo/")
        assert resp.status_code == 200

    def test_paises_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/clientes/paises/")
        assert resp.status_code == 200

    def test_ciudades_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/clientes/ciudades/")
        assert resp.status_code == 200

    def test_cuentas_corrientes_list(self, api_cajero):
        resp = api_cajero.get("/api/v1/clientes/cuentas-corrientes/")
        assert resp.status_code == 200

    def test_cuentas_corrientes_rol_sin_permiso_de_cobranza_falla(self, api_client, db):
        from apps.usuarios.models import Usuario
        usuario_cocina = Usuario.objects.create_user(
            email="cocina@test.com",
            password="test1234",
            nombre="Cocina",
            apellido="Test",
            rol=Usuario.Rol.COCINA,
        )
        api_client.force_authenticate(user=usuario_cocina)
        resp = api_client.get("/api/v1/clientes/cuentas-corrientes/")
        assert resp.status_code == 403

    def test_cuentas_corrientes_supervisor_puede_cobrar(self, api_client, db):
        from apps.usuarios.models import Usuario
        usuario_supervisor = Usuario.objects.create_user(
            email="supervisor@test.com",
            password="test1234",
            nombre="Supervisor",
            apellido="Test",
            rol=Usuario.Rol.SUPERVISOR,
        )
        api_client.force_authenticate(user=usuario_supervisor)
        resp = api_client.get("/api/v1/clientes/cuentas-corrientes/")
        assert resp.status_code == 200

    def test_responsables_list(self, api_admin):
        resp = api_admin.get("/api/v1/clientes/responsables/")
        assert resp.status_code == 200

    def test_requiere_autenticacion(self, api_client):
        resp = api_client.get("/api/v1/clientes/clientes/")
        assert resp.status_code in (401, 403)


# ── HijoViewSet — CLIENTE_WEB filter ─────────────────────────────────────────

@pytest.mark.django_db
class TestHijoViewSetClienteWeb:

    def test_cliente_web_ve_solo_sus_hijos(self, api_cliente_web, hijo_fixture):
        resp = api_cliente_web.get("/api/v1/clientes/hijos/")
        assert resp.status_code == 200
        ids = [h["id_hijo"] for h in resp.data["results"]] if "results" in resp.data else [h["id_hijo"] for h in resp.data]
        assert hijo_fixture.pk in ids

    def test_cliente_web_sin_cliente_ve_nada(self, api_client, db):
        from apps.usuarios.models import Usuario
        user = Usuario.objects.create_user(
            email="sincliente@test.com", password="x",
            nombre="Sin", apellido="Cliente",
            rol=Usuario.Rol.CLIENTE_WEB,
        )
        api_client.force_authenticate(user=user)
        resp = api_client.get("/api/v1/clientes/hijos/")
        assert resp.status_code == 200
        data = resp.data.get("results", resp.data)
        assert len(data) == 0


# ── RestriccionHijoViewSet — CLIENTE_WEB filter ───────────────────────────────

@pytest.mark.django_db
class TestRestriccionClienteWeb:

    def test_cliente_web_sin_cliente_ve_nada(self, api_client, db):
        from apps.usuarios.models import Usuario
        user = Usuario.objects.create_user(
            email="sincliente2@test.com", password="x",
            nombre="Sin", apellido="Cliente",
            rol=Usuario.Rol.CLIENTE_WEB,
        )
        api_client.force_authenticate(user=user)
        resp = api_client.get("/api/v1/clientes/restricciones/")
        assert resp.status_code == 200
        data = resp.data.get("results", resp.data)
        assert len(data) == 0

    def test_cliente_web_con_cliente_lista_sus_restricciones(self, api_cliente_web):
        resp = api_cliente_web.get("/api/v1/clientes/restricciones/")
        assert resp.status_code == 200


# ── AlumnoResponsableViewSet ──────────────────────────────────────────────────

@pytest.mark.django_db
class TestAlumnoResponsable:

    def test_create_asigna_agregado_por(self, api_admin, hijo_fixture, cliente):
        resp = api_admin.post(
            "/api/v1/clientes/responsables/",
            {
                "hijo": hijo_fixture.pk,
                "cliente": cliente.pk,
                "parentesco": "MADRE",
                "es_titular": False,
                "orden_cobro": 2,
            },
            format="json",
        )
        assert resp.status_code == 201

    def test_destroy_ultimo_titular_falla(self, api_admin, alumno_responsable):
        resp = api_admin.delete(f"/api/v1/clientes/responsables/{alumno_responsable.pk}/")
        assert resp.status_code == 400

    def test_destroy_no_unico_ok(self, api_admin, hijo_fixture, cliente, usuario_admin):
        from apps.clientes.models import AlumnoResponsable
        # Crear un segundo responsable para que el primero no sea el único
        otro = AlumnoResponsable.objects.create(
            hijo=hijo_fixture, cliente=cliente, parentesco="MADRE",
            es_titular=False, orden_cobro=2, agregado_por=usuario_admin,
        )
        resp = api_admin.delete(f"/api/v1/clientes/responsables/{otro.pk}/")
        assert resp.status_code == 204

    def test_set_titular_ok(self, api_admin, alumno_responsable):
        resp = api_admin.post(
            f"/api/v1/clientes/responsables/{alumno_responsable.pk}/set_titular/"
        )
        assert resp.status_code == 200


# ── ReporteCuentaCorrienteView ────────────────────────────────────────────────

@pytest.mark.django_db
class TestReporteCuentaCorriente:

    def test_json_sin_deudas(self, api_admin):
        resp = api_admin.get("/api/v1/clientes/reporte-cuenta-corriente/")
        assert resp.status_code == 200
        assert "resumen" in resp.data
        assert "detalle" in resp.data

    def test_json_con_deuda(self, api_admin, cuenta_con_deuda):
        resp = api_admin.get("/api/v1/clientes/reporte-cuenta-corriente/")
        assert resp.status_code == 200
        assert resp.data["resumen"]["clientes_con_deuda"] >= 1

    def test_csv_format(self, api_admin):
        resp = api_admin.get(
            "/api/v1/clientes/reporte-cuenta-corriente/",
            {"formato": "csv"},
        )
        assert resp.status_code == 200
        assert "text/csv" in resp["Content-Type"]
        assert b"REPORTE CUENTA CORRIENTE" in resp.content

    def test_csv_con_deuda_incluye_cliente(self, api_admin, cuenta_con_deuda):
        resp = api_admin.get(
            "/api/v1/clientes/reporte-cuenta-corriente/",
            {"formato": "csv"},
        )
        assert resp.status_code == 200
        assert b"Juan" in resp.content

    def test_aging_buckets_31_60(self, api_admin, cliente, usuario_cajero):
        from apps.clientes.models import CuentaCorrienteCliente
        from django.utils import timezone
        import datetime
        fecha_vieja = timezone.now() - datetime.timedelta(days=45)
        CuentaCorrienteCliente.objects.create(
            cliente=cliente, tipo=CuentaCorrienteCliente.Tipo.DEBITO,
            monto=Decimal("10000"), saldo_anterior=Decimal("0"),
            saldo_resultante=Decimal("10000"), descripcion="Deuda 45d",
            creado_por=usuario_cajero, fecha=fecha_vieja,
        )
        resp = api_admin.get("/api/v1/clientes/reporte-cuenta-corriente/")
        assert resp.status_code == 200
        detalle = resp.data.get("detalle", [])
        assert any(d.get("aging") == "31-60" for d in detalle)

    def test_aging_buckets_61_90(self, api_admin, cliente, usuario_cajero):
        from apps.clientes.models import CuentaCorrienteCliente
        from django.utils import timezone
        import datetime
        fecha_vieja = timezone.now() - datetime.timedelta(days=75)
        CuentaCorrienteCliente.objects.create(
            cliente=cliente, tipo=CuentaCorrienteCliente.Tipo.DEBITO,
            monto=Decimal("10000"), saldo_anterior=Decimal("0"),
            saldo_resultante=Decimal("10000"), descripcion="Deuda 75d",
            creado_por=usuario_cajero, fecha=fecha_vieja,
        )
        resp = api_admin.get("/api/v1/clientes/reporte-cuenta-corriente/")
        assert resp.status_code == 200
        detalle = resp.data.get("detalle", [])
        assert any(d.get("aging") == "61-90" for d in detalle)

    def test_aging_buckets_90_plus(self, api_admin, cliente, usuario_cajero):
        from apps.clientes.models import CuentaCorrienteCliente
        from django.utils import timezone
        import datetime
        fecha_vieja = timezone.now() - datetime.timedelta(days=120)
        CuentaCorrienteCliente.objects.create(
            cliente=cliente, tipo=CuentaCorrienteCliente.Tipo.DEBITO,
            monto=Decimal("10000"), saldo_anterior=Decimal("0"),
            saldo_resultante=Decimal("10000"), descripcion="Deuda 120d",
            creado_por=usuario_cajero, fecha=fecha_vieja,
        )
        resp = api_admin.get("/api/v1/clientes/reporte-cuenta-corriente/")
        assert resp.status_code == 200
        detalle = resp.data.get("detalle", [])
        assert any(d.get("aging") == "90+" for d in detalle)

    def test_requiere_autenticacion(self, api_client):
        resp = api_client.get("/api/v1/clientes/reporte-cuenta-corriente/")
        assert resp.status_code in (401, 403)

    def test_excel_retorna_xlsx(self, api_admin):
        resp = api_admin.get(
            "/api/v1/clientes/reporte-cuenta-corriente/",
            {"formato": "excel"},
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp["Content-Type"]
        assert "attachment" in resp.get("Content-Disposition", "")
        assert resp.get("Content-Disposition", "").endswith(".xlsx\"")

    def test_excel_con_datos_genera_filas(self, api_admin, cuenta_con_deuda):
        import io
        from openpyxl import load_workbook
        resp = api_admin.get(
            "/api/v1/clientes/reporte-cuenta-corriente/",
            {"formato": "excel"},
        )
        assert resp.status_code == 200
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        assert ws.max_row >= 2  # encabezado + al menos 1 fila de datos


# ── CuentaCorrienteClienteViewSet.create ─────────────────────────────────────

@pytest.mark.django_db
class TestCuentaCorrienteCreate:

    def test_sin_cliente_retorna_400(self, api_cajero):
        resp = api_cajero.post(
            "/api/v1/clientes/cuentas-corrientes/",
            {"monto": "10000"},
            format="json",
        )
        assert resp.status_code == 400
        assert "cliente" in str(resp.data)

    def test_monto_cero_retorna_400(self, api_cajero, cliente):
        resp = api_cajero.post(
            "/api/v1/clientes/cuentas-corrientes/",
            {"cliente": cliente.pk, "monto": "0"},
            format="json",
        )
        assert resp.status_code == 400

    def test_monto_negativo_retorna_400(self, api_cajero, cliente):
        resp = api_cajero.post(
            "/api/v1/clientes/cuentas-corrientes/",
            {"cliente": cliente.pk, "monto": "-5000"},
            format="json",
        )
        assert resp.status_code == 400

    def test_monto_no_numerico_retorna_400(self, api_cajero, cliente):
        resp = api_cajero.post(
            "/api/v1/clientes/cuentas-corrientes/",
            {"cliente": cliente.pk, "monto": "no_es_numero"},
            format="json",
        )
        assert resp.status_code == 400

    def test_cliente_inexistente_retorna_404(self, api_cajero):
        resp = api_cajero.post(
            "/api/v1/clientes/cuentas-corrientes/",
            {"cliente": 999_999, "monto": "10000"},
            format="json",
        )
        assert resp.status_code == 404

    def test_happy_path_crea_movimiento_credito(self, api_cajero, cliente, usuario_cajero):
        from apps.clientes.models import CuentaCorrienteCliente
        # Deuda previa para que haya saldo_anterior
        CuentaCorrienteCliente.objects.create(
            cliente=cliente,
            tipo=CuentaCorrienteCliente.Tipo.DEBITO,
            monto=Decimal("30000"),
            saldo_anterior=Decimal("0"),
            saldo_resultante=Decimal("30000"),
            creado_por=usuario_cajero,
        )
        resp = api_cajero.post(
            "/api/v1/clientes/cuentas-corrientes/",
            {"cliente": cliente.pk, "monto": "20000", "descripcion": "Pago parcial"},
            format="json",
        )
        assert resp.status_code == 201
        assert resp.data["tipo"] == "CREDITO"
        assert Decimal(resp.data["monto"]) == Decimal("20000")

    def test_medio_pago_pos_debito_se_resuelve_correctamente(self, api_cajero, usuario_cajero, cliente):
        """Regresión: 'POS DEBITO' (código fijo del frontend) no matcheaba
        contra 'POS Bancario debito' (nombre real del catálogo) y el
        MovimientoCaja quedaba sin medio_pago — se contaba como Prepago en
        el arqueo de caja aunque fuera un cobro de cuenta corriente."""
        from apps.contabilidad.models import Caja, CierreCaja, MovimientoCaja
        from apps.core.models import MedioPago

        pos_debito = MedioPago.objects.create(descripcion="POS Bancario debito", activo=True)
        caja = Caja.objects.create(nombre="Caja CC Test", activo=True)
        cierre = CierreCaja.objects.create(
            caja=caja, empleado=usuario_cajero, monto_inicial=Decimal("0"),
            estado=CierreCaja.Estado.ABIERTO,
        )

        resp = api_cajero.post(
            "/api/v1/clientes/cuentas-corrientes/",
            {"cliente": cliente.pk, "monto": "262500", "medio_pago": "POS DEBITO"},
            format="json",
        )
        assert resp.status_code == 201

        mov = MovimientoCaja.objects.get(cierre=cierre, tipo=MovimientoCaja.Tipo.INGRESO)
        assert mov.medio_pago_id == pos_debito.id_medio_pago

    def test_ambas_categorias_sin_origen_retorna_400(self, api_cajero, cliente, usuario_cajero):
        from apps.clientes.models import CuentaCorrienteCliente
        CuentaCorrienteCliente.objects.create(
            cliente=cliente, tipo=CuentaCorrienteCliente.Tipo.DEBITO,
            monto=Decimal("100000"), saldo_anterior=Decimal("0"), saldo_resultante=Decimal("100000"),
            creado_por=usuario_cajero, origen=CuentaCorrienteCliente.Origen.CANTINA,
        )
        CuentaCorrienteCliente.objects.create(
            cliente=cliente, tipo=CuentaCorrienteCliente.Tipo.DEBITO,
            monto=Decimal("50000"), saldo_anterior=Decimal("100000"), saldo_resultante=Decimal("150000"),
            creado_por=usuario_cajero, origen=CuentaCorrienteCliente.Origen.ALMUERZO,
        )
        resp = api_cajero.post(
            "/api/v1/clientes/cuentas-corrientes/",
            {"cliente": cliente.pk, "monto": "30000"},
            format="json",
        )
        assert resp.status_code == 400
        assert "origen" in str(resp.data).lower()

    def test_ambas_categorias_con_origen_crea_movimiento_tageado(self, api_cajero, cliente, usuario_cajero):
        from apps.clientes.models import CuentaCorrienteCliente
        CuentaCorrienteCliente.objects.create(
            cliente=cliente, tipo=CuentaCorrienteCliente.Tipo.DEBITO,
            monto=Decimal("100000"), saldo_anterior=Decimal("0"), saldo_resultante=Decimal("100000"),
            creado_por=usuario_cajero, origen=CuentaCorrienteCliente.Origen.CANTINA,
        )
        CuentaCorrienteCliente.objects.create(
            cliente=cliente, tipo=CuentaCorrienteCliente.Tipo.DEBITO,
            monto=Decimal("50000"), saldo_anterior=Decimal("100000"), saldo_resultante=Decimal("150000"),
            creado_por=usuario_cajero, origen=CuentaCorrienteCliente.Origen.ALMUERZO,
        )
        resp = api_cajero.post(
            "/api/v1/clientes/cuentas-corrientes/",
            {"cliente": cliente.pk, "monto": "30000", "origen": "ALMUERZO"},
            format="json",
        )
        assert resp.status_code == 201
        assert resp.data["origen"] == "ALMUERZO"

    def test_monto_supera_deuda_de_la_categoria_retorna_400(self, api_cajero, cliente, usuario_cajero):
        from apps.clientes.models import CuentaCorrienteCliente
        CuentaCorrienteCliente.objects.create(
            cliente=cliente, tipo=CuentaCorrienteCliente.Tipo.DEBITO,
            monto=Decimal("100000"), saldo_anterior=Decimal("0"), saldo_resultante=Decimal("100000"),
            creado_por=usuario_cajero, origen=CuentaCorrienteCliente.Origen.CANTINA,
        )
        CuentaCorrienteCliente.objects.create(
            cliente=cliente, tipo=CuentaCorrienteCliente.Tipo.DEBITO,
            monto=Decimal("50000"), saldo_anterior=Decimal("100000"), saldo_resultante=Decimal("150000"),
            creado_por=usuario_cajero, origen=CuentaCorrienteCliente.Origen.ALMUERZO,
        )
        resp = api_cajero.post(
            "/api/v1/clientes/cuentas-corrientes/",
            {"cliente": cliente.pk, "monto": "60000", "origen": "ALMUERZO"},
            format="json",
        )
        assert resp.status_code == 400

    def test_solo_una_categoria_con_deuda_no_requiere_origen(self, api_cajero, cliente, usuario_cajero):
        from apps.clientes.models import CuentaCorrienteCliente
        CuentaCorrienteCliente.objects.create(
            cliente=cliente, tipo=CuentaCorrienteCliente.Tipo.DEBITO,
            monto=Decimal("100000"), saldo_anterior=Decimal("0"), saldo_resultante=Decimal("100000"),
            creado_por=usuario_cajero, origen=CuentaCorrienteCliente.Origen.CANTINA,
        )
        resp = api_cajero.post(
            "/api/v1/clientes/cuentas-corrientes/",
            {"cliente": cliente.pk, "monto": "30000"},
            format="json",
        )
        assert resp.status_code == 201
        assert resp.data["origen"] == "CANTINA"


# ── ClienteViewSet.perform_create → _crear_usuario_portal ────────────────────

@pytest.mark.django_db
class TestCrearUsuarioPortal:

    def test_post_cliente_crea_usuario_portal_con_email_sintetico(
        self, api_admin, tipo_cliente, lista_precio,
    ):
        """Crear un cliente sin email vía API genera un CLIENTE_WEB con email sintético."""
        from apps.usuarios.models import Usuario
        resp = api_admin.post(
            "/api/v1/clientes/clientes/",
            {
                "nombres": "Portal",
                "apellidos": "Nuevo",
                "ruc_ci": "PORTAL99",
                "tipo_cliente": tipo_cliente.pk,
                "lista_precio": lista_precio.pk,
            },
            format="json",
        )
        assert resp.status_code == 201
        assert Usuario.objects.filter(
            rol=Usuario.Rol.CLIENTE_WEB,
            email__contains="portal.tita.local",
        ).exists()

    def test_post_cliente_con_email_existente_sin_cliente_vincula_usuario(
        self, api_admin, tipo_cliente, lista_precio,
    ):
        """Si ya existe un Usuario con ese email y sin cliente, se lo vincula en lugar de crear uno nuevo."""
        from apps.usuarios.models import Usuario
        usuario_previo = Usuario.objects.create_user(
            email="padre@escuela.com",
            password="x",
            nombre="Pre",
            apellido="Existente",
            rol=Usuario.Rol.CLIENTE_WEB,
        )
        resp = api_admin.post(
            "/api/v1/clientes/clientes/",
            {
                "nombres": "Pre",
                "apellidos": "Existente",
                "ruc_ci": "PRE001",
                "email": "padre@escuela.com",
                "tipo_cliente": tipo_cliente.pk,
                "lista_precio": lista_precio.pk,
            },
            format="json",
        )
        assert resp.status_code == 201
        usuario_previo.refresh_from_db()
        assert usuario_previo.cliente_id is not None
        # No se creó un segundo usuario con ese email
        assert Usuario.objects.filter(email="padre@escuela.com").count() == 1

    def test_cliente_ya_con_portal_no_crea_segundo_usuario(
        self, api_admin, tipo_cliente, lista_precio,
    ):
        """Si el cliente ya tiene usuario_portal, _crear_usuario_portal retorna sin crear otro."""
        from apps.clientes.models import Cliente
        from apps.usuarios.models import Usuario
        from apps.clientes.views import _crear_usuario_portal
        cliente = Cliente.objects.create(
            nombres="Ya",
            apellidos="Tiene",
            ruc_ci="YATIENE01",
            tipo_cliente=tipo_cliente,
            lista_precio=lista_precio,
        )
        Usuario.objects.create_user(
            email="yatiene@test.com",
            password="x",
            nombre="Ya",
            apellido="Tiene",
            rol=Usuario.Rol.CLIENTE_WEB,
            cliente=cliente,
        )
        before = Usuario.objects.count()
        _crear_usuario_portal(cliente)  # debe ser no-op
        assert Usuario.objects.count() == before


# ── AlumnoResponsableViewSet.set_titular — ValueError ────────────────────────

@pytest.mark.django_db
class TestSetTitularError:

    def test_responsable_inactivo_retorna_400(
        self, api_admin, hijo_fixture, cliente, usuario_admin,
    ):
        """Intentar poner como titular a un responsable inactivo retorna 400."""
        from apps.clientes.models import AlumnoResponsable
        responsable_inactivo = AlumnoResponsable.objects.create(
            hijo=hijo_fixture,
            cliente=cliente,
            parentesco="TIO",
            es_titular=False,
            orden_cobro=2,
            activo=False,
            agregado_por=usuario_admin,
        )
        resp = api_admin.post(
            f"/api/v1/clientes/responsables/{responsable_inactivo.pk}/set_titular/"
        )
        assert resp.status_code == 400
        assert "inactivo" in str(resp.data).lower()


# ── ClienteViewSet / HijoViewSet — guarda de hard-delete ───────────────────────

@pytest.mark.django_db
class TestClienteDestroyGuard:

    def test_no_se_puede_eliminar_cliente_activo(self, api_admin, cliente):
        from apps.clientes.models import Cliente
        assert cliente.activo is True
        resp = api_admin.delete(f"/api/v1/clientes/clientes/{cliente.pk}/")
        assert resp.status_code == 400
        assert "inactivo" in str(resp.data).lower()
        assert Cliente.objects.filter(pk=cliente.pk).exists()

    def test_elimina_cliente_inactivo_sin_historial(self, api_admin, cliente):
        from apps.clientes.models import Cliente
        cliente.activo = False
        cliente.save(update_fields=["activo"])
        resp = api_admin.delete(f"/api/v1/clientes/clientes/{cliente.pk}/")
        assert resp.status_code == 204
        assert not Cliente.objects.filter(pk=cliente.pk).exists()

    def test_no_se_puede_eliminar_cliente_con_ventas(self, api_admin, cliente, usuario_admin):
        from apps.clientes.models import Cliente
        from apps.ventas.models import Venta
        cliente.activo = False
        cliente.save(update_fields=["activo"])
        Venta.objects.create(cliente=cliente, cajero=usuario_admin, monto_total="10000")
        resp = api_admin.delete(f"/api/v1/clientes/clientes/{cliente.pk}/")
        assert resp.status_code == 400
        assert "asociados" in str(resp.data).lower()
        assert Cliente.objects.filter(pk=cliente.pk).exists()


@pytest.mark.django_db
class TestHijoDestroyGuard:

    def test_no_se_puede_eliminar_hijo_activo(self, api_admin, hijo_fixture):
        assert hijo_fixture.activo is True
        resp = api_admin.delete(f"/api/v1/clientes/hijos/{hijo_fixture.pk}/")
        assert resp.status_code == 400
        assert "inactivo" in str(resp.data).lower()

    def test_elimina_hijo_inactivo_sin_historial(self, api_admin, hijo_fixture):
        from apps.clientes.models import Hijo
        hijo_fixture.activo = False
        hijo_fixture.save(update_fields=["activo"])
        resp = api_admin.delete(f"/api/v1/clientes/hijos/{hijo_fixture.pk}/")
        assert resp.status_code == 204
        assert not Hijo.objects.filter(pk=hijo_fixture.pk).exists()
