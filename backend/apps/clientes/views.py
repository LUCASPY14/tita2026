"""
Views para la app clientes
"""

import csv
from datetime import date
from decimal import Decimal

from django.db.models import DecimalField, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.http import FileResponse, Http404, HttpResponse
from django.utils import timezone

from rest_framework import viewsets, status
from rest_framework.filters import SearchFilter
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from common.permissions import IsAdmin, IsAdminOrReadOnly, IsCajeroOrAdmin, IsStaffOrClienteWeb, IsStaffUser
from common.utils.medios_pago import resolver_medio_pago
from apps.usuarios.auditoria import registrar_auditoria

from django_filters.rest_framework import DjangoFilterBackend

from rest_framework.decorators import action

from .models import (
    AlumnoResponsable,
    AutorizacionSaldoNegativo,
    Ciudad,
    Cliente,
    CuentaCorrienteCliente,
    Grado,
    HistorialGrado,
    Hijo,
    Pais,
    RestriccionHijo,
    TipoCliente,
)
from .serializers import (
    AlumnoResponsableSerializer,
    AutorizacionSaldoNegativoSerializer,
    CiudadSerializer,
    ClienteSerializer,
    CuentaCorrienteClienteSerializer,
    GradoSerializer,
    HistorialGradoSerializer,
    HijoSerializer,
    PaisSerializer,
    RestriccionHijoSerializer,
    TipoClienteSerializer,
)
from .services import cambiar_titular, purgar_alumno


def _crear_usuario_portal(cliente):
    """Crea (o vincula) un usuario CLIENTE_WEB para el cliente dado.

    - Si ya tiene usuario portal, no hace nada.
    - Usa el email del cliente como identificador; si no tiene email, genera
      uno sintético: <ruc_ci_limpio>@portal.tita.local
    - La contraseña inicial es el RUC/CI tal como está almacenado.
    - El usuario queda marcado con debe_cambiar_contrasena=True.
    """
    from apps.usuarios.models import Usuario

    if hasattr(cliente, "usuario_portal"):
        return

    ruc_ci_limpio = cliente.ruc_ci.strip()
    email = (cliente.email or "").strip()
    if not email:
        sufijo = ruc_ci_limpio.replace("-", "").replace(".", "")
        email = f"{sufijo}@portal.tita.local"

    if Usuario.objects.filter(email=email).exists():
        # Ya existe un usuario con ese email: vincular si no tiene cliente
        usuario = Usuario.objects.get(email=email)
        if not usuario.cliente_id:
            usuario.cliente = cliente
            usuario.save(update_fields=["cliente"])
        return

    Usuario.objects.create_user(
        email=email,
        password=ruc_ci_limpio,
        nombre=cliente.nombres,
        apellido=cliente.apellidos,
        rol=Usuario.Rol.CLIENTE_WEB,
        is_active=True,
        email_verificado=bool(cliente.email),
        debe_cambiar_contrasena=True,
        cliente=cliente,
    )


class ClienteViewSet(viewsets.ModelViewSet):
    queryset = Cliente.objects.select_related("tipo_cliente", "lista_precio").annotate(
        saldo_negativo_tarjetas=Coalesce(
            Sum(
                "hijos__tarjeta__saldo_actual",
                filter=Q(hijos__tarjeta__saldo_actual__lt=0),
            ),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
        )
    ).order_by("apellidos", "nombres")
    serializer_class = ClienteSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ["activo", "tipo_cliente"]
    search_fields = ["ruc_ci", "nombres", "apellidos"]

    def perform_create(self, serializer):
        cliente = serializer.save()
        _crear_usuario_portal(cliente)
        registrar_auditoria(
            request=self.request,
            operacion="CREAR_CLIENTE",
            tabla="clientes_cliente",
            id_registro=cliente.id,
            descripcion=f"Cliente creado: {cliente.nombres} {cliente.apellidos} RUC/CI={cliente.ruc_ci}",
        )

    def perform_update(self, serializer):
        cliente = serializer.save()
        registrar_auditoria(
            request=self.request,
            operacion="MODIFICAR_CLIENTE",
            tabla="clientes_cliente",
            id_registro=cliente.id,
            descripcion=f"Cliente modificado: {cliente.nombres} {cliente.apellidos} RUC/CI={cliente.ruc_ci}",
        )

    def perform_destroy(self, instance):
        from django.db.models import ProtectedError
        from rest_framework.exceptions import ValidationError

        if instance.activo:
            raise ValidationError(
                "Solo se pueden eliminar clientes inactivos. Desactivalo primero."
            )
        # instance.delete() vacía instance.pk/id — capturar antes para la auditoría.
        cliente_id = instance.id
        descripcion = f"Cliente eliminado: {instance.nombres} {instance.apellidos} RUC/CI={instance.ruc_ci}"
        try:
            instance.delete()
        except ProtectedError:
            raise ValidationError(
                "No se puede eliminar: tiene ventas, tarjetas u otros registros asociados. "
                "El historial no se puede borrar — dejalo desactivado."
            )
        registrar_auditoria(
            request=self.request,
            operacion="ELIMINAR_CLIENTE",
            tabla="clientes_cliente",
            id_registro=cliente_id,
            descripcion=descripcion,
        )

    @action(detail=True, methods=["post"], url_path="reset-pin",
            permission_classes=[IsAdminOrReadOnly])
    def reset_pin(self, request, pk=None):
        """Solo ADMIN: resetea el PIN del cliente a '0000'."""
        if request.user.rol != "ADMIN":
            return Response({"error": "Solo el administrador puede resetear el PIN."},
                            status=status.HTTP_403_FORBIDDEN)
        cliente = self.get_object()
        cliente.set_pin("0000")
        return Response({"ok": True, "mensaje": "PIN reseteado a 0000."})

    @action(detail=True, methods=["post"], url_path="cambiar-pin",
            permission_classes=[IsAuthenticated])
    def cambiar_pin(self, request, pk=None):
        """El padre (CLIENTE_WEB) cambia su propio PIN. El admin puede cambiar cualquiera."""
        cliente = self.get_object()
        user = request.user

        # Verificar que el cliente web solo modifique su propio PIN
        if user.rol == "CLIENTE_WEB":
            if not hasattr(user, "cliente") or user.cliente_id != cliente.pk:
                return Response({"error": "No podés modificar el PIN de otro cliente."},
                                status=status.HTTP_403_FORBIDDEN)

        pin_actual = request.data.get("pin_actual", "")
        pin_nuevo = request.data.get("pin_nuevo", "")

        if not pin_nuevo or len(pin_nuevo) != 4 or not pin_nuevo.isdigit():
            return Response({"error": "El PIN nuevo debe ser exactamente 4 dígitos."},
                            status=status.HTTP_400_BAD_REQUEST)

        # Admin no necesita confirmar el PIN actual
        if user.rol != "ADMIN":
            if not cliente.check_pin(pin_actual):
                return Response({"error": "PIN actual incorrecto."},
                                status=status.HTTP_400_BAD_REQUEST)

        cliente.set_pin(pin_nuevo)
        return Response({"ok": True, "mensaje": "PIN actualizado correctamente."})


class CuentaCorrienteClienteViewSet(viewsets.ModelViewSet):
    queryset = CuentaCorrienteCliente.objects.select_related("cliente").all()
    serializer_class = CuentaCorrienteClienteSerializer
    permission_classes = [IsStaffUser]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["cliente", "tipo"]
    http_method_names = ["get", "post", "head", "options"]

    def create(self, request, *args, **kwargs):
        from decimal import Decimal, InvalidOperation
        from django.db import transaction
        from apps.contabilidad.models import CierreCaja, MovimientoCaja

        cliente_id = request.data.get("cliente")
        monto_raw = request.data.get("monto")
        descripcion = request.data.get("descripcion") or "Pago de cuenta corriente"
        metodo_pago = (request.data.get("medio_pago") or "").strip()

        if not cliente_id:
            return Response({"error": "El campo 'cliente' es obligatorio."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            monto = Decimal(str(monto_raw))
            if monto <= 0:
                raise ValueError
        except (InvalidOperation, ValueError, TypeError):
            return Response({"error": "El monto debe ser un número mayor a 0."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            cliente = Cliente.objects.get(pk=cliente_id)
        except Cliente.DoesNotExist:
            return Response({"error": "Cliente no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        genera_factura_legal = bool(request.data.get("genera_factura_legal", False))
        nro_factura = str(request.data.get("nro_factura", "") or "").strip()

        with transaction.atomic():
            mov = CuentaCorrienteCliente.objects.create(
                cliente=cliente,
                tipo=CuentaCorrienteCliente.Tipo.CREDITO,
                monto=monto,
                descripcion=descripcion,
                creado_por=request.user,
            )
            # Registrar ingreso en caja si el usuario tiene un turno abierto
            cierre = CierreCaja.objects.filter(
                empleado=request.user, estado=CierreCaja.Estado.ABIERTO
            ).first()
            if cierre:
                medio_pago_obj = resolver_medio_pago(metodo_pago)
                MovimientoCaja.objects.create(
                    cierre=cierre,
                    tipo=MovimientoCaja.Tipo.INGRESO,
                    monto=monto,
                    descripcion=f"Cobro CC — {cliente.nombres} {cliente.apellidos}",
                    medio_pago=medio_pago_obj,
                )

            # Facturación
            if genera_factura_legal:
                if nro_factura:
                    from apps.contabilidad.services import FacturacionService
                    factura = FacturacionService.emitir_factura(
                        cliente=cliente,
                        nro_factura=nro_factura,
                        monto_total=monto,
                        **FacturacionService._calcular_iva_10(monto),
                    )
                    mov.factura = factura
                mov.genera_factura_legal = True
                mov.save(update_fields=["factura", "genera_factura_legal"])

        return Response(self.get_serializer(mov).data, status=status.HTTP_201_CREATED)


class TipoClienteViewSet(viewsets.ModelViewSet):
    queryset = TipoCliente.objects.all()
    serializer_class = TipoClienteSerializer
    permission_classes = [IsAdminOrReadOnly]


class HijoViewSet(viewsets.ModelViewSet):
    serializer_class = HijoSerializer
    permission_classes = [IsStaffOrClienteWeb]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ["activo", "cliente_responsable"]
    search_fields = ["nombre", "apellido"]

    def get_permissions(self):
        if self.action == "foto":
            return [IsCajeroOrAdmin()]
        if self.action in ("pendientes_purga", "aprobar_purga"):
            return [IsAdmin()]
        return super().get_permissions()

    def get_queryset(self):
        from django.db.models import Prefetch
        qs = Hijo.objects.select_related("cliente_responsable", "grado").prefetch_related(
            Prefetch(
                "responsables",
                queryset=AlumnoResponsable.objects.filter(activo=True)
                    .select_related("cliente")
                    .order_by("orden_cobro"),
            )
        )
        if self.request.user.es_cliente_web:
            cliente = getattr(self.request.user, "cliente", None)
            if cliente is None:
                return qs.none()
            qs = qs.filter(cliente_responsable=cliente)
        return qs

    def perform_create(self, serializer):
        hijo = serializer.save()
        registrar_auditoria(
            request=self.request,
            operacion="CREAR_HIJO",
            tabla="clientes_hijo",
            id_registro=hijo.id,
            descripcion=f"Alumno creado: {hijo.nombre_completo} (cliente={hijo.cliente_responsable_id})",
        )

    def perform_update(self, serializer):
        instance = serializer.instance
        era_activo = instance.activo
        hijo = serializer.save()
        # Baja manual: si se acaba de desactivar y no tenía fecha_baja, se
        # completa sola — de acá arranca el reloj de 1 año hasta la purga.
        if era_activo and not hijo.activo and not hijo.fecha_baja:
            hijo.fecha_baja = timezone.now()
            hijo.save(update_fields=["fecha_baja"])

    def perform_destroy(self, instance):
        from django.db.models import ProtectedError
        from rest_framework.exceptions import ValidationError

        if instance.activo:
            raise ValidationError(
                "Solo se pueden eliminar alumnos inactivos. Desactivalo primero."
            )
        hijo_id = instance.id
        descripcion = f"Alumno eliminado: {instance.nombre_completo} (cliente={instance.cliente_responsable_id})"
        try:
            instance.delete()
        except ProtectedError:
            raise ValidationError(
                "No se puede eliminar: tiene consumos, tarjeta u otros registros asociados. "
                "El historial no se puede borrar — dejalo desactivado (o usá la purga de datos tras 1 año de baja)."
            )
        registrar_auditoria(
            request=self.request,
            operacion="ELIMINAR_HIJO",
            tabla="clientes_hijo",
            id_registro=hijo_id,
            descripcion=descripcion,
        )

    @action(detail=True, methods=["get"], url_path="foto")
    def foto(self, request, pk=None):
        """Sirve la foto del alumno — solo ADMIN/CAJERO (ver get_permissions)."""
        hijo = self.get_object()
        if not hijo.foto_perfil:
            raise Http404("Este alumno no tiene foto cargada.")
        return FileResponse(hijo.foto_perfil.open("rb"), content_type="image/jpeg")

    @action(detail=False, methods=["get"], url_path="pendientes-purga")
    def pendientes_purga(self, request):
        """Alumnos dados de baja hace más de 1 año, pendientes de aprobación de purga."""
        qs = Hijo.objects.filter(
            purga_solicitada_en__isnull=False,
            datos_purgados=False,
        ).select_related("cliente_responsable", "grado").order_by("purga_solicitada_en")
        return Response(HijoSerializer(qs, many=True, context={"request": request}).data)

    @action(detail=True, methods=["post"], url_path="aprobar-purga")
    def aprobar_purga(self, request, pk=None):
        """Ejecuta la anonimización de datos sensibles — requiere aprobación explícita de un ADMIN."""
        hijo = self.get_object()
        if not hijo.purga_solicitada_en:
            return Response(
                {"error": "Este alumno no está marcado como pendiente de purga."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if hijo.datos_purgados:
            return Response(
                {"error": "Este alumno ya fue purgado."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        hijo = purgar_alumno(hijo, aprobado_por=request.user)
        registrar_auditoria(
            request=self.request,
            operacion="PURGAR_DATOS_ALUMNO",
            tabla="clientes_hijo",
            id_registro=hijo.id,
            descripcion=f"Datos sensibles anonimizados: {hijo.nombre_completo}",
        )
        return Response(HijoSerializer(hijo, context={"request": request}).data)


class GradoViewSet(viewsets.ModelViewSet):
    queryset = Grado.objects.all()
    serializer_class = GradoSerializer
    permission_classes = [IsAdminOrReadOnly]


class HistorialGradoViewSet(viewsets.ModelViewSet):
    queryset = HistorialGrado.objects.select_related("hijo").all()
    serializer_class = HistorialGradoSerializer
    permission_classes = [IsStaffUser]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["hijo", "anio_escolar"]


class RestriccionHijoViewSet(viewsets.ModelViewSet):
    serializer_class = RestriccionHijoSerializer
    permission_classes = [IsStaffOrClienteWeb]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["hijo", "severidad", "activo"]

    def get_queryset(self):
        qs = RestriccionHijo.objects.select_related("hijo__cliente_responsable")
        if self.request.user.es_cliente_web:
            cliente = getattr(self.request.user, "cliente", None)
            if cliente is None:
                return qs.none()
            qs = qs.filter(hijo__cliente_responsable=cliente)
        return qs


class AutorizacionSaldoNegativoViewSet(viewsets.ModelViewSet):
    queryset = AutorizacionSaldoNegativo.objects.select_related("cliente", "venta").all()
    serializer_class = AutorizacionSaldoNegativoSerializer
    permission_classes = [IsCajeroOrAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["cliente", "estado"]


class PaisViewSet(viewsets.ModelViewSet):
    queryset = Pais.objects.all()
    serializer_class = PaisSerializer
    permission_classes = [IsAdminOrReadOnly]


class CiudadViewSet(viewsets.ModelViewSet):
    queryset = Ciudad.objects.select_related("pais").all()
    serializer_class = CiudadSerializer
    permission_classes = [IsAdminOrReadOnly]


# ==============================================================================
# REPORTE CUENTA CORRIENTE
# ==============================================================================

class ReporteCuentaCorrienteView(APIView):
    """
    GET /api/clientes/reporte-cuenta-corriente/
    Clientes con saldo pendiente y distribución por aging (30/60/90/90+ días).
    Opcional: ?formato=csv
    """
    permission_classes = [IsStaffUser]

    def get(self, request):
        hoy = date.today()

        # Último movimiento por cliente para conocer saldo actual
        from django.db.models import OuterRef, Subquery

        ultimo_mov = (
            CuentaCorrienteCliente.objects
            .filter(cliente=OuterRef("pk"))
            .order_by("-id")
            .values("saldo_resultante")[:1]
        )
        clientes_con_saldo = (
            Cliente.objects
            .filter(activo=True)
            .annotate(saldo_deuda=Subquery(ultimo_mov))
            .filter(saldo_deuda__gt=0)
            .order_by("-saldo_deuda")
        )

        filas = []
        for cliente in clientes_con_saldo:
            saldo = Decimal(str(cliente.saldo_deuda or 0))

            # Aging = antigüedad del ciclo de deuda actual.
            # Buscamos el último movimiento donde el saldo volvió a 0 (o quedó negativo),
            # y contamos desde el primer movimiento posterior a ese punto.
            ultimo_saldo_cero = (
                CuentaCorrienteCliente.objects
                .filter(cliente=cliente, saldo_resultante__lte=0)
                .order_by("-fecha", "-id")
                .values("id", "fecha")
                .first()
            )
            dias_atraso = 0
            if ultimo_saldo_cero:
                # Primer mov con saldo > 0 DESPUÉS del último cierre
                inicio_ciclo = (
                    CuentaCorrienteCliente.objects
                    .filter(cliente=cliente, id__gt=ultimo_saldo_cero["id"], saldo_resultante__gt=0)
                    .order_by("fecha", "id")
                    .values("fecha")
                    .first()
                )
            else:
                # El saldo nunca fue 0: tomar el movimiento más antiguo con saldo > 0
                inicio_ciclo = (
                    CuentaCorrienteCliente.objects
                    .filter(cliente=cliente, saldo_resultante__gt=0)
                    .order_by("fecha", "id")
                    .values("fecha")
                    .first()
                )
            if inicio_ciclo:
                dias_atraso = (hoy - inicio_ciclo["fecha"].date()).days

            if dias_atraso <= 30:
                bucket = "0-30"
            elif dias_atraso <= 60:
                bucket = "31-60"
            elif dias_atraso <= 90:
                bucket = "61-90"
            else:
                bucket = "90+"

            filas.append({
                "cliente_id": cliente.pk,
                "cliente": cliente.nombre_completo,
                "ruc_ci": cliente.ruc_ci,
                "telefono": cliente.telefono or "",
                "email": cliente.email or "",
                "saldo_deuda": int(saldo),
                "dias_atraso": dias_atraso,
                "aging": bucket,
            })

        # Totales por bucket
        aging_totales = {"0-30": 0, "31-60": 0, "61-90": 0, "90+": 0}
        for f in filas:
            aging_totales[f["aging"]] += f["saldo_deuda"]

        total_deuda = sum(f["saldo_deuda"] for f in filas)

        formato = request.query_params.get("formato")

        if formato == "csv":
            resp = HttpResponse(content_type="text/csv; charset=utf-8-sig")
            resp["Content-Disposition"] = (
                f'attachment; filename="cuenta_corriente_{hoy}.csv"'
            )
            writer = csv.writer(resp)
            writer.writerow(["REPORTE CUENTA CORRIENTE", str(hoy)])
            writer.writerow([])
            writer.writerow(["Cliente", "RUC/CI", "Teléfono", "Email", "Saldo (Gs)", "Días atraso", "Aging"])
            for f in filas:
                writer.writerow([f["cliente"], f["ruc_ci"], f["telefono"],
                                  f["email"], f["saldo_deuda"], f["dias_atraso"], f["aging"]])
            writer.writerow([])
            writer.writerow(["TOTALES POR AGING"])
            for bucket, total in aging_totales.items():
                writer.writerow([bucket, total])
            return resp

        if formato == "excel":
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Cuenta Corriente"

            header_fill = PatternFill("solid", fgColor="1E3A5F")
            header_font = Font(bold=True, color="FFFFFF")
            total_font = Font(bold=True)
            totals_fill = PatternFill("solid", fgColor="E8F0FE")

            ws.append([f"REPORTE CUENTA CORRIENTE — {hoy}"])
            ws["A1"].font = Font(bold=True, size=13)
            ws.append([])

            headers = ["Cliente", "RUC/CI", "Teléfono", "Email", "Saldo (Gs)", "Días atraso", "Aging"]
            ws.append(headers)
            for cell in ws[ws.max_row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for f in filas:
                ws.append([f["cliente"], f["ruc_ci"], f["telefono"],
                            f["email"], f["saldo_deuda"], f["dias_atraso"], f["aging"]])

            ws.append([])
            ws.append(["TOTALES POR AGING", "", "", "", "", "", ""])
            for cell in ws[ws.max_row]:
                cell.font = total_font
            for bucket, total in aging_totales.items():
                row = [bucket, "", "", "", total, "", ""]
                ws.append(row)
                for cell in ws[ws.max_row]:
                    cell.fill = totals_fill

            ws.append([])
            ws.append(["TOTAL DEUDA", "", "", "", total_deuda, "", ""])
            for cell in ws[ws.max_row]:
                cell.font = total_font

            col_widths = [35, 14, 14, 28, 14, 14, 10]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            resp = HttpResponse(
                buf.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            resp["Content-Disposition"] = f'attachment; filename="cuenta_corriente_{hoy}.xlsx"'
            return resp

        return Response({
            "fecha": str(hoy),
            "resumen": {
                "clientes_con_deuda": len(filas),
                "total_deuda": total_deuda,
                "aging": aging_totales,
            },
            "detalle": filas,
        })


# ==============================================================================
# ALUMNO RESPONSABLE
# ==============================================================================

class AlumnoResponsableViewSet(viewsets.ModelViewSet):
    """
    CRUD de responsables de un alumno.

    Rutas automáticas (router):
      GET    /clientes/responsables/          → lista (filtrable por hijo, cliente, activo)
      POST   /clientes/responsables/          → crear responsable
      GET    /clientes/responsables/{id}/     → detalle
      PATCH  /clientes/responsables/{id}/     → editar
      DELETE /clientes/responsables/{id}/     → eliminar (valida que no sea el último titular)

    Acciones custom:
      POST /clientes/responsables/{id}/set_titular/ → designar como titular
    """

    serializer_class = AlumnoResponsableSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["hijo", "cliente", "activo", "es_titular"]

    def get_queryset(self):
        return (
            AlumnoResponsable.objects.select_related("hijo", "cliente", "agregado_por")
            .order_by("hijo_id", "orden_cobro")
        )

    def perform_create(self, serializer):
        serializer.save(agregado_por=self.request.user)

    def perform_destroy(self, instance):
        # Impedir eliminar el único titular activo
        if instance.es_titular and instance.activo:
            otros_activos = AlumnoResponsable.objects.filter(
                hijo=instance.hijo, activo=True
            ).exclude(pk=instance.pk).count()
            if otros_activos == 0:
                from rest_framework.exceptions import ValidationError
                raise ValidationError(
                    "No se puede eliminar al único responsable activo del alumno. "
                    "Agregue otro responsable antes de eliminar este."
                )
        instance.delete()

    @action(detail=True, methods=["post"], url_path="set_titular")
    def set_titular(self, request, pk=None):
        """
        POST /clientes/responsables/{id}/set_titular/
        Designa este responsable como titular del alumno.
        Sincroniza Hijo.cliente_responsable automáticamente.
        """
        responsable = self.get_object()
        try:
            actualizado = cambiar_titular(
                hijo=responsable.hijo,
                nuevo_cliente_id=responsable.cliente_id,
                changed_by=request.user,
            )
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            AlumnoResponsableSerializer(actualizado).data,
            status=status.HTTP_200_OK,
        )
