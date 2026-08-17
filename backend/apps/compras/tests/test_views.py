"""Tests para apps.compras.views — ViewSets de proveedores, compras, pagos, etc."""
import pytest
from decimal import Decimal
from rest_framework.test import APIClient


@pytest.fixture
def api(usuario_cajero):
    client = APIClient()
    client.force_authenticate(user=usuario_cajero)
    return client


@pytest.fixture
def proveedor(db):
    from apps.compras.models import Proveedor
    return Proveedor.objects.create(ruc="80002001-2", razon_social="Prov. View Test")


@pytest.fixture
def compra_contado(db, proveedor, usuario_cajero, producto):
    from apps.compras.services import CompraService
    return CompraService.registrar_compra(
        proveedor=proveedor,
        creado_por=usuario_cajero,
        tipo_pago="CONTADO",
        items=[{
            "producto": producto,
            "cantidad": Decimal("2"),
            "costo_unitario": Decimal("5000"),
        }],
    )


@pytest.fixture
def compra_credito(db, proveedor, usuario_cajero, producto):
    from apps.compras.services import CompraService
    return CompraService.registrar_compra(
        proveedor=proveedor,
        creado_por=usuario_cajero,
        tipo_pago="CREDITO",
        items=[{
            "producto": producto,
            "cantidad": Decimal("3"),
            "costo_unitario": Decimal("8000"),
        }],
    )


# ── ProveedorViewSet ───────────────────────────────────────────────────────────

@pytest.mark.django_db
class TestProveedorViewSet:
    def test_list(self, api, proveedor):
        resp = api.get("/api/v1/compras/proveedores/")
        assert resp.status_code == 200
        assert resp.data["count"] >= 1

    def test_create(self, api):
        resp = api.post("/api/v1/compras/proveedores/", {
            "ruc": "80099001-9",
            "razon_social": "Nuevo Proveedor",
        })
        assert resp.status_code == 201

    def test_retrieve(self, api, proveedor):
        resp = api.get(f"/api/v1/compras/proveedores/{proveedor.pk}/")
        assert resp.status_code == 200
        assert resp.data["razon_social"] == "Prov. View Test"

    def test_sin_auth_retorna_401(self, proveedor):
        resp = APIClient().get("/api/v1/compras/proveedores/")
        assert resp.status_code == 401


# ── CuentaCorrienteProveedorViewSet ────────────────────────────────────────────

@pytest.mark.django_db
class TestCuentaCorrienteViewSet:
    def test_list(self, api, proveedor, usuario_cajero):
        from apps.compras.models import CuentaCorrienteProveedor
        CuentaCorrienteProveedor.objects.create(
            proveedor=proveedor,
            tipo=CuentaCorrienteProveedor.Tipo.DEBITO,
            monto=Decimal("100000"),
            saldo_anterior=Decimal("0"),
            saldo_resultante=Decimal("0"),
            creado_por=usuario_cajero,
        )
        resp = api.get("/api/v1/compras/cuentas-corrientes/")
        assert resp.status_code == 200


# ── CompraViewSet ──────────────────────────────────────────────────────────────

@pytest.mark.django_db
class TestCompraViewSet:
    def test_list_con_queryset_anotado(self, api, compra_contado):
        resp = api.get("/api/v1/compras/compras/")
        assert resp.status_code == 200
        assert resp.data["count"] >= 1

    def test_retrieve(self, api, compra_contado):
        resp = api.get(f"/api/v1/compras/compras/{compra_contado.pk}/")
        assert resp.status_code == 200
        assert "saldo_pendiente" in resp.data

    def test_confirmar_entrega_contado_retorna_400(self, api, compra_contado):
        resp = api.post(f"/api/v1/compras/compras/{compra_contado.pk}/confirmar-entrega/")
        assert resp.status_code == 400
        assert "crédito" in resp.data["error"]

    def test_confirmar_entrega_credito_ok(self, api, compra_credito):
        from apps.compras.models import Compra
        resp = api.post(f"/api/v1/compras/compras/{compra_credito.pk}/confirmar-entrega/")
        assert resp.status_code == 200
        compra_credito.refresh_from_db()
        assert compra_credito.estado_entrega == Compra.EstadoEntrega.RECIBIDA

    def test_confirmar_entrega_ya_recibida_retorna_400(self, api, compra_credito):
        api.post(f"/api/v1/compras/compras/{compra_credito.pk}/confirmar-entrega/")
        resp = api.post(f"/api/v1/compras/compras/{compra_credito.pk}/confirmar-entrega/")
        assert resp.status_code == 400


# ── Otros ViewSets ─────────────────────────────────────────────────────────────

@pytest.mark.django_db
class TestOtrosViewSets:
    def test_detalles_compra_list(self, api, compra_contado):
        resp = api.get("/api/v1/compras/detalles-compra/")
        assert resp.status_code == 200

    def test_aplicaciones_pago_list(self, api):
        resp = api.get("/api/v1/compras/aplicaciones-pago/")
        assert resp.status_code == 200

    # Pagos y notas de crédito requieren ADMIN o SUPERVISOR
    def test_pagos_list(self, api_admin):
        resp = api_admin.get("/api/v1/compras/pagos/")
        assert resp.status_code == 200

    def test_pagos_list_cajero_retorna_403(self, api):
        resp = api.get("/api/v1/compras/pagos/")
        assert resp.status_code == 403

    def test_notas_credito_list(self, api_admin):
        resp = api_admin.get("/api/v1/compras/notas-credito/")
        assert resp.status_code == 200

    def test_detalles_nc_list(self, api_admin):
        resp = api_admin.get("/api/v1/compras/detalles-nc/")
        assert resp.status_code == 200


# ── OrdenCompraViewSet ─────────────────────────────────────────────────────────

@pytest.fixture
def api_admin(usuario_admin):
    client = APIClient()
    client.force_authenticate(user=usuario_admin)
    return client


@pytest.fixture
def orden_borrador(db, proveedor, usuario_cajero, producto):
    from apps.compras.models import DetalleOrdenCompra, OrdenCompra
    orden = OrdenCompra.objects.create(
        proveedor=proveedor,
        creado_por=usuario_cajero,
        tipo_pago=OrdenCompra.TipoPago.CONTADO,
        estado=OrdenCompra.Estado.BORRADOR,
        monto_total=Decimal("10000"),
    )
    DetalleOrdenCompra.objects.create(
        orden=orden,
        producto=producto,
        cantidad=Decimal("2"),
        costo_unitario=Decimal("5000"),
        subtotal=Decimal("10000"),
    )
    return orden


@pytest.fixture
def orden_pendiente(db, orden_borrador, api_admin):
    api_admin.post(f"/api/v1/compras/ordenes/{orden_borrador.pk}/submit/")
    orden_borrador.refresh_from_db()
    return orden_borrador


@pytest.fixture
def orden_aprobada(db, orden_pendiente, api_admin):
    api_admin.post(f"/api/v1/compras/ordenes/{orden_pendiente.pk}/aprobar/")
    orden_pendiente.refresh_from_db()
    return orden_pendiente


@pytest.mark.django_db
class TestOrdenCompraViewSet:

    def test_crear_orden_borrador(self, api_admin, proveedor, producto):
        resp = api_admin.post("/api/v1/compras/ordenes/", {
            "proveedor": proveedor.pk,
            "tipo_pago": "CONTADO",
            "items": [{"producto": producto.pk, "cantidad": "1", "costo_unitario": "5000"}],
        }, format="json")
        assert resp.status_code == 201
        assert resp.data["estado"] == "BORRADOR"

    def test_list_ordenes(self, api_admin, orden_borrador):
        resp = api_admin.get("/api/v1/compras/ordenes/")
        assert resp.status_code == 200
        assert resp.data["count"] >= 1

    def test_submit_borrador_a_pendiente(self, api, orden_borrador):
        resp = api.post(f"/api/v1/compras/ordenes/{orden_borrador.pk}/submit/")
        assert resp.status_code == 200
        assert resp.data["estado"] == "PENDIENTE"

    def test_submit_sin_detalles_retorna_400(self, api, proveedor, usuario_cajero):
        from apps.compras.models import OrdenCompra
        orden_vacia = OrdenCompra.objects.create(
            proveedor=proveedor, creado_por=usuario_cajero,
            tipo_pago=OrdenCompra.TipoPago.CONTADO,
            estado=OrdenCompra.Estado.BORRADOR, monto_total=Decimal("0"),
        )
        resp = api.post(f"/api/v1/compras/ordenes/{orden_vacia.pk}/submit/")
        assert resp.status_code == 400
        assert "producto" in resp.data["error"].lower()

    def test_submit_no_borrador_retorna_400(self, api, orden_pendiente):
        resp = api.post(f"/api/v1/compras/ordenes/{orden_pendiente.pk}/submit/")
        assert resp.status_code == 400

    def test_aprobar_pendiente_admin(self, api_admin, orden_pendiente):
        resp = api_admin.post(f"/api/v1/compras/ordenes/{orden_pendiente.pk}/aprobar/")
        assert resp.status_code == 200
        assert resp.data["estado"] == "APROBADA"

    def test_aprobar_cajero_retorna_403(self, api, orden_pendiente):
        resp = api.post(f"/api/v1/compras/ordenes/{orden_pendiente.pk}/aprobar/")
        assert resp.status_code == 403

    def test_rechazar_pendiente_con_motivo(self, api_admin, orden_pendiente):
        resp = api_admin.post(
            f"/api/v1/compras/ordenes/{orden_pendiente.pk}/rechazar/",
            {"motivo": "Precio fuera de rango"},
            format="json",
        )
        assert resp.status_code == 200
        assert resp.data["estado"] == "RECHAZADA"
        assert resp.data["motivo_rechazo"] == "Precio fuera de rango"

    def test_rechazar_sin_motivo_retorna_400(self, api_admin, orden_pendiente):
        resp = api_admin.post(
            f"/api/v1/compras/ordenes/{orden_pendiente.pk}/rechazar/",
            {"motivo": ""},
            format="json",
        )
        assert resp.status_code == 400

    def test_convertir_aprobada_crea_compra(self, api_admin, orden_aprobada):
        from apps.compras.models import Compra
        resp = api_admin.post(f"/api/v1/compras/ordenes/{orden_aprobada.pk}/convertir/")
        assert resp.status_code == 200
        assert resp.data["orden"]["estado"] == "CONVERTIDA"
        compra_id = resp.data["compra_id"]
        assert Compra.objects.filter(pk=compra_id).exists()

    def test_convertir_cajero_retorna_403(self, api, orden_aprobada):
        resp = api.post(f"/api/v1/compras/ordenes/{orden_aprobada.pk}/convertir/")
        assert resp.status_code == 403

    def test_aprobar_no_pendiente_retorna_400(self, api_admin, orden_aprobada):
        resp = api_admin.post(f"/api/v1/compras/ordenes/{orden_aprobada.pk}/aprobar/")
        assert resp.status_code == 400
        assert "Pendiente" in resp.data["error"]

    def test_rechazar_no_pendiente_retorna_400(self, api_admin, orden_aprobada):
        resp = api_admin.post(
            f"/api/v1/compras/ordenes/{orden_aprobada.pk}/rechazar/",
            {"motivo": "test"},
            format="json",
        )
        assert resp.status_code == 400
        assert "Pendiente" in resp.data["error"]

    def test_convertir_no_aprobada_retorna_400(self, api_admin, orden_pendiente):
        resp = api_admin.post(f"/api/v1/compras/ordenes/{orden_pendiente.pk}/convertir/")
        assert resp.status_code == 400
        assert "Aprobada" in resp.data["error"]

    def test_convertir_validation_error_retorna_400(self, api_admin, orden_aprobada, producto):
        from unittest.mock import patch
        from rest_framework.exceptions import ValidationError
        with patch(
            "apps.compras.views.CompraService.registrar_compra",
            side_effect=ValidationError({"items": "Error de validación"}),
        ):
            resp = api_admin.post(f"/api/v1/compras/ordenes/{orden_aprobada.pk}/convertir/")
        assert resp.status_code == 400

    def test_patch_borrador_reemplaza_items(self, api_admin, orden_borrador, producto):
        """PATCH en BORRADOR: reemplaza los items y recalcula monto_total."""
        from apps.compras.models import DetalleOrdenCompra
        nuevo_costo = Decimal("8000")
        resp = api_admin.patch(
            f"/api/v1/compras/ordenes/{orden_borrador.pk}/",
            {"items": [{"producto": producto.pk, "cantidad": "3", "costo_unitario": str(nuevo_costo)}]},
            format="json",
        )
        assert resp.status_code == 200
        orden_borrador.refresh_from_db()
        assert orden_borrador.monto_total == Decimal("24000")
        assert DetalleOrdenCompra.objects.filter(orden=orden_borrador).count() == 1

    def test_patch_no_borrador_retorna_400(self, api_admin, orden_pendiente, producto):
        """PATCH en estado distinto de BORRADOR → 400."""
        resp = api_admin.patch(
            f"/api/v1/compras/ordenes/{orden_pendiente.pk}/",
            {"items": [{"producto": producto.pk, "cantidad": "1", "costo_unitario": "5000"}]},
            format="json",
        )
        assert resp.status_code == 400
        assert "Borrador" in str(resp.data)


# ── NotaCreditoProveedor con observacion ──────────────────────────────────────

@pytest.mark.django_db
class TestNotaCreditoConObservacion:
    """Línea 318: desc incluye observacion cuando se provee."""

    def test_create_nc_con_observacion_crea_cc_con_desc(self, api_admin, proveedor):
        resp = api_admin.post(
            "/api/v1/compras/notas-credito/",
            {
                "proveedor": proveedor.pk,
                "monto_total": "50000",
                "observacion": "Devolución de mercadería dañada",
            },
            format="json",
        )
        assert resp.status_code == 201
        from apps.compras.models import CuentaCorrienteProveedor
        cc = CuentaCorrienteProveedor.objects.filter(proveedor=proveedor).first()
        assert "Devolución" in cc.descripcion


# ── ReporteComprasProveedoresView ─────────────────────────────────────────────

@pytest.mark.django_db
class TestReporteComprasProveedores:
    """Líneas 555 (funnel loop) y 571 (CSV loop con datos)."""

    def test_sin_params_retorna_400(self, api_admin):
        resp = api_admin.get("/api/v1/compras/reporte-compras/")
        assert resp.status_code == 400

    def test_con_fechas_retorna_estructura(self, api_admin):
        resp = api_admin.get(
            "/api/v1/compras/reporte-compras/",
            {"desde": "2026-01-01", "hasta": "2026-12-31"},
        )
        assert resp.status_code == 200
        assert "por_proveedor" in resp.data
        assert "funnel_oc" in resp.data

    def test_con_ordenes_llena_funnel(self, api_admin, orden_pendiente):
        from django.utils import timezone
        hoy = timezone.now().date().isoformat()
        resp = api_admin.get(
            "/api/v1/compras/reporte-compras/",
            {"desde": hoy, "hasta": hoy},
        )
        assert resp.status_code == 200
        assert resp.data["funnel_oc"]["PENDIENTE"] >= 1

    def test_formato_csv_con_datos_incluye_filas(
        self, api_admin, compra_contado, proveedor
    ):
        from django.utils import timezone
        hoy = timezone.now().date().isoformat()
        resp = api_admin.get(
            "/api/v1/compras/reporte-compras/",
            {"desde": hoy, "hasta": hoy, "formato": "csv"},
        )
        assert resp.status_code == 200
        assert "text/csv" in resp["Content-Type"]
        content = resp.content.decode("utf-8-sig")
        assert proveedor.razon_social in content


# ── ReporteNotasCreditoCompraView ─────────────────────────────────────────────

@pytest.mark.django_db
class TestReporteNotasCreditoCompra:
    """Líneas 600-675: reporte de notas de crédito — JSON y CSV."""

    def test_sin_params_retorna_400(self, api_admin):
        resp = api_admin.get("/api/v1/compras/reporte-notas-credito/")
        assert resp.status_code == 400

    def test_con_fechas_retorna_estructura(self, api_admin):
        resp = api_admin.get(
            "/api/v1/compras/reporte-notas-credito/",
            {"desde": "2026-01-01", "hasta": "2026-12-31"},
        )
        assert resp.status_code == 200
        assert "detalle" in resp.data
        assert "resumen" in resp.data

    def test_con_datos_muestra_nc(self, api_admin, proveedor, usuario_admin):
        from django.utils import timezone
        from apps.compras.models import NotaCreditoProveedor
        NotaCreditoProveedor.objects.create(
            proveedor=proveedor,
            monto_total=Decimal("30000"),
            estado=NotaCreditoProveedor.Estado.EMITIDA,
            creado_por=usuario_admin,
        )
        hoy = timezone.now().date().isoformat()
        resp = api_admin.get(
            "/api/v1/compras/reporte-notas-credito/",
            {"desde": hoy, "hasta": hoy},
        )
        assert resp.status_code == 200
        assert len(resp.data["detalle"]) == 1
        assert resp.data["resumen"]["total_emitidas"] == 1

    def test_filtro_por_estado(self, api_admin, proveedor, usuario_admin):
        from django.utils import timezone
        from apps.compras.models import NotaCreditoProveedor
        NotaCreditoProveedor.objects.create(
            proveedor=proveedor, monto_total=Decimal("10000"),
            estado=NotaCreditoProveedor.Estado.ANULADA, creado_por=usuario_admin,
        )
        hoy = timezone.now().date().isoformat()
        resp = api_admin.get(
            "/api/v1/compras/reporte-notas-credito/",
            {"desde": hoy, "hasta": hoy, "estado": "EMITIDA"},
        )
        assert resp.status_code == 200
        assert len(resp.data["detalle"]) == 0

    def test_formato_csv_con_datos_incluye_filas(self, api_admin, proveedor, usuario_admin):
        from django.utils import timezone
        from apps.compras.models import NotaCreditoProveedor
        NotaCreditoProveedor.objects.create(
            proveedor=proveedor, monto_total=Decimal("25000"),
            estado=NotaCreditoProveedor.Estado.EMITIDA, creado_por=usuario_admin,
        )
        hoy = timezone.now().date().isoformat()
        resp = api_admin.get(
            "/api/v1/compras/reporte-notas-credito/",
            {"desde": hoy, "hasta": hoy, "formato": "csv"},
        )
        assert resp.status_code == 200
        assert "text/csv" in resp["Content-Type"]
        content = resp.content.decode("utf-8-sig")
        assert proveedor.razon_social in content


# ── ReporteAgingProveedoresView ───────────────────────────────────────────────

@pytest.mark.django_db
class TestReporteAgingProveedores:
    """Líneas 691-774: aging cuentas a pagar — JSON y CSV."""

    def test_retorna_estructura(self, api_admin):
        resp = api_admin.get("/api/v1/compras/reporte-aging-proveedores/")
        assert resp.status_code == 200
        assert "resumen" in resp.data
        assert "detalle" in resp.data

    def test_con_deuda_muestra_proveedor(
        self, api_admin, proveedor, usuario_cajero, producto
    ):
        from apps.compras.services import CompraService
        CompraService.registrar_compra(
            proveedor=proveedor,
            creado_por=usuario_cajero,
            tipo_pago="CREDITO",
            items=[{"producto": producto, "cantidad": Decimal("1"), "costo_unitario": Decimal("100000")}],
        )
        resp = api_admin.get("/api/v1/compras/reporte-aging-proveedores/")
        assert resp.status_code == 200
        assert resp.data["resumen"]["proveedores_con_deuda"] >= 1

    def test_formato_csv_sin_datos_retorna_csv(self, api_admin):
        resp = api_admin.get(
            "/api/v1/compras/reporte-aging-proveedores/",
            {"formato": "csv"},
        )
        assert resp.status_code == 200
        assert "text/csv" in resp["Content-Type"]

    def test_formato_csv_con_datos_incluye_proveedor(
        self, api_admin, proveedor, usuario_cajero, producto
    ):
        from apps.compras.services import CompraService
        CompraService.registrar_compra(
            proveedor=proveedor,
            creado_por=usuario_cajero,
            tipo_pago="CREDITO",
            items=[{"producto": producto, "cantidad": Decimal("2"), "costo_unitario": Decimal("50000")}],
        )
        resp = api_admin.get(
            "/api/v1/compras/reporte-aging-proveedores/",
            {"formato": "csv"},
        )
        assert resp.status_code == 200
        content = resp.content.decode("utf-8-sig")
        assert proveedor.razon_social in content

    def test_deuda_antigua_cae_en_bucket_31_60(
        self, api_admin, proveedor, usuario_cajero
    ):
        """Deuda con 40 días de antigüedad → bucket 31-60 (líneas 730-731)."""
        from apps.compras.models import CuentaCorrienteProveedor
        from django.utils import timezone
        from datetime import timedelta
        cc = CuentaCorrienteProveedor.objects.create(
            proveedor=proveedor,
            tipo=CuentaCorrienteProveedor.Tipo.DEBITO,
            monto=Decimal("200000"),
            descripcion="Deuda antigua",
            creado_por=usuario_cajero,
        )
        # Retroceder fecha para que caiga en el bucket 31-60
        CuentaCorrienteProveedor.objects.filter(pk=cc.pk).update(
            fecha=timezone.now() - timedelta(days=40)
        )
        resp = api_admin.get("/api/v1/compras/reporte-aging-proveedores/")
        assert resp.status_code == 200
        detalle = resp.data["detalle"]
        assert len(detalle) >= 1
        assert detalle[0]["aging"] == "31-60"

    def test_deuda_70_dias_cae_en_bucket_61_90(
        self, api_admin, proveedor, usuario_cajero
    ):
        """Deuda con 70 días → bucket 61-90 (línea 733)."""
        from apps.compras.models import CuentaCorrienteProveedor
        from django.utils import timezone
        from datetime import timedelta
        cc = CuentaCorrienteProveedor.objects.create(
            proveedor=proveedor,
            tipo=CuentaCorrienteProveedor.Tipo.DEBITO,
            monto=Decimal("250000"),
            descripcion="Deuda 70 días",
            creado_por=usuario_cajero,
        )
        CuentaCorrienteProveedor.objects.filter(pk=cc.pk).update(
            fecha=timezone.now() - timedelta(days=70)
        )
        resp = api_admin.get("/api/v1/compras/reporte-aging-proveedores/")
        assert resp.status_code == 200
        assert any(d["aging"] == "61-90" for d in resp.data["detalle"])

    def test_deuda_muy_antigua_cae_en_bucket_90_mas(
        self, api_admin, proveedor, usuario_cajero
    ):
        """Deuda con 100 días → bucket 90+ (líneas 734-735)."""
        from apps.compras.models import CuentaCorrienteProveedor
        from django.utils import timezone
        from datetime import timedelta
        cc = CuentaCorrienteProveedor.objects.create(
            proveedor=proveedor,
            tipo=CuentaCorrienteProveedor.Tipo.DEBITO,
            monto=Decimal("300000"),
            descripcion="Deuda muy antigua",
            creado_por=usuario_cajero,
        )
        CuentaCorrienteProveedor.objects.filter(pk=cc.pk).update(
            fecha=timezone.now() - timedelta(days=100)
        )
        resp = api_admin.get("/api/v1/compras/reporte-aging-proveedores/")
        assert resp.status_code == 200
        detalle = resp.data["detalle"]
        assert any(d["aging"] == "90+" for d in detalle)

    def test_excel_retorna_xlsx(self, api_admin):
        resp = api_admin.get(
            "/api/v1/compras/reporte-aging-proveedores/",
            {"formato": "excel"},
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp["Content-Type"]
        assert "attachment" in resp.get("Content-Disposition", "")
        assert resp.get("Content-Disposition", "").endswith(".xlsx\"")

    def test_excel_con_datos_genera_filas(self, api_admin, proveedor, usuario_cajero):
        import io
        from decimal import Decimal
        from openpyxl import load_workbook
        from apps.compras.models import CuentaCorrienteProveedor
        CuentaCorrienteProveedor.objects.create(
            proveedor=proveedor,
            tipo=CuentaCorrienteProveedor.Tipo.DEBITO,
            monto=Decimal("100000"),
            descripcion="Deuda test excel",
            creado_por=usuario_cajero,
        )
        resp = api_admin.get(
            "/api/v1/compras/reporte-aging-proveedores/",
            {"formato": "excel"},
        )
        assert resp.status_code == 200
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        assert ws.max_row >= 2  # encabezado + al menos 1 fila de datos


# ── CompraViewSet.create y update (líneas 107-174) ────────────────────────────

@pytest.mark.django_db
class TestCompraViewSetCreateUpdate:
    """Cubre CompraViewSet.create, _resolve_items y update."""

    def test_create_compra_contado(self, api, proveedor, producto):
        resp = api.post(
            "/api/v1/compras/compras/",
            {
                "proveedor": proveedor.pk,
                "tipo_pago": "CONTADO",
                "items": [{"producto": producto.pk, "cantidad": "2", "costo_unitario": "5000"}],
            },
            format="json",
        )
        assert resp.status_code == 201
        assert resp.data["monto_total"] == "10000"

    def test_create_compra_credito(self, api, proveedor, producto):
        resp = api.post(
            "/api/v1/compras/compras/",
            {
                "proveedor": proveedor.pk,
                "tipo_pago": "CREDITO",
                "nro_factura_proveedor": "F-001",
                "observaciones": "Entrega pendiente",
                "items": [{"producto": producto.pk, "cantidad": "1", "costo_unitario": "15000"}],
            },
            format="json",
        )
        assert resp.status_code == 201
        assert resp.data["tipo_pago"] == "CREDITO"

    def test_create_proveedor_no_encontrado_retorna_400(self, api, producto):
        resp = api.post(
            "/api/v1/compras/compras/",
            {
                "proveedor": 99999,
                "tipo_pago": "CONTADO",
                "items": [{"producto": producto.pk, "cantidad": "1", "costo_unitario": "5000"}],
            },
            format="json",
        )
        assert resp.status_code == 400
        assert "proveedor" in resp.data

    def test_update_compra_cambia_proveedor_e_items(self, api, compra_contado, proveedor, producto):
        from apps.compras.models import Proveedor
        otro = Proveedor.objects.create(ruc="80003001-3", razon_social="Otro Proveedor")
        resp = api.patch(
            f"/api/v1/compras/compras/{compra_contado.pk}/",
            {
                "proveedor": otro.pk,
                "tipo_pago": "CONTADO",
                "items": [{"producto": producto.pk, "cantidad": "5", "costo_unitario": "2000"}],
            },
            format="json",
        )
        assert resp.status_code == 200
        assert resp.data["monto_total"] == "10000"

    def test_update_proveedor_no_encontrado_retorna_400(self, api, compra_contado, producto):
        resp = api.patch(
            f"/api/v1/compras/compras/{compra_contado.pk}/",
            {
                "proveedor": 99999,
                "tipo_pago": "CONTADO",
                "items": [{"producto": producto.pk, "cantidad": "1", "costo_unitario": "5000"}],
            },
            format="json",
        )
        assert resp.status_code == 400
        assert "proveedor" in resp.data


# ── PagoProveedorViewSet.create (líneas 209-258) ──────────────────────────────

@pytest.mark.django_db
class TestPagoProveedorCreate:
    """Cubre PagoProveedorViewSet.create y sus ramas de error."""

    def test_create_pago_actualiza_estado_pagado(
        self, api_admin, compra_credito, medio_pago_efectivo
    ):
        resp = api_admin.post(
            "/api/v1/compras/pagos/",
            {
                "compra": compra_credito.pk,
                "monto": str(compra_credito.monto_total),
                "medio_pago": medio_pago_efectivo.pk,
            },
            format="json",
        )
        assert resp.status_code == 201
        compra_credito.refresh_from_db()
        assert compra_credito.estado_pago == "PAGADO"

    def test_create_pago_parcial_actualiza_estado_parcial(
        self, api_admin, compra_credito, medio_pago_efectivo
    ):
        monto_parcial = compra_credito.monto_total / 2
        resp = api_admin.post(
            "/api/v1/compras/pagos/",
            {
                "compra": compra_credito.pk,
                "monto": str(monto_parcial),
                "medio_pago": medio_pago_efectivo.pk,
            },
            format="json",
        )
        assert resp.status_code == 201
        compra_credito.refresh_from_db()
        assert compra_credito.estado_pago == "PARCIAL"

    def test_create_pago_registra_movimiento_cc(
        self, api_admin, compra_credito, medio_pago_efectivo
    ):
        from apps.compras.models import CuentaCorrienteProveedor
        api_admin.post(
            "/api/v1/compras/pagos/",
            {
                "compra": compra_credito.pk,
                "monto": "5000",
                "medio_pago": medio_pago_efectivo.pk,
            },
            format="json",
        )
        assert CuentaCorrienteProveedor.objects.filter(
            proveedor=compra_credito.proveedor,
            tipo=CuentaCorrienteProveedor.Tipo.CREDITO,
        ).exists()

    def test_create_pago_compra_no_encontrada_retorna_400(
        self, api_admin, medio_pago_efectivo
    ):
        resp = api_admin.post(
            "/api/v1/compras/pagos/",
            {"compra": 99999, "monto": "5000", "medio_pago": medio_pago_efectivo.pk},
            format="json",
        )
        assert resp.status_code == 400
        assert "compra" in resp.data

    def test_create_pago_medio_pago_no_encontrado_retorna_400(
        self, api_admin, compra_credito
    ):
        resp = api_admin.post(
            "/api/v1/compras/pagos/",
            {"compra": compra_credito.pk, "monto": "5000", "medio_pago": 99999},
            format="json",
        )
        assert resp.status_code == 400
        assert "medio_pago" in resp.data


# ── NotaCreditoProveedorViewSet.create — ramas de validación (líneas 283-316) ─

@pytest.mark.django_db
class TestNotaCreditoCreateValidaciones:
    """Cubre las ramas de validación del create de NC que faltaban."""

    def test_sin_proveedor_retorna_400(self, api_admin):
        resp = api_admin.post(
            "/api/v1/compras/notas-credito/",
            {"monto_total": "10000"},
            format="json",
        )
        assert resp.status_code == 400
        assert "proveedor" in resp.data["error"].lower()

    def test_monto_cero_retorna_400(self, api_admin, proveedor):
        resp = api_admin.post(
            "/api/v1/compras/notas-credito/",
            {"proveedor": proveedor.pk, "monto_total": "0"},
            format="json",
        )
        assert resp.status_code == 400
        assert "monto" in resp.data["error"].lower()

    def test_monto_texto_invalido_retorna_400(self, api_admin, proveedor):
        resp = api_admin.post(
            "/api/v1/compras/notas-credito/",
            {"proveedor": proveedor.pk, "monto_total": "no-es-numero"},
            format="json",
        )
        assert resp.status_code == 400
        assert "monto" in resp.data["error"].lower()

    def test_proveedor_no_encontrado_retorna_404(self, api_admin):
        resp = api_admin.post(
            "/api/v1/compras/notas-credito/",
            {"proveedor": 99999, "monto_total": "10000"},
            format="json",
        )
        assert resp.status_code == 404
        assert "proveedor" in resp.data["error"].lower()

    def test_compra_de_otro_proveedor_retorna_404(
        self, api_admin, proveedor, compra_contado
    ):
        otro = __import__("apps.compras.models", fromlist=["Proveedor"]).Proveedor.objects.create(
            ruc="80004001-4", razon_social="Otro Prov"
        )
        resp = api_admin.post(
            "/api/v1/compras/notas-credito/",
            {
                "proveedor": otro.pk,
                "monto_total": "5000",
                "compra_original": compra_contado.pk,
            },
            format="json",
        )
        assert resp.status_code == 404
        assert "compra" in resp.data["error"].lower()

    def test_create_con_compra_original_incluye_id_en_descripcion(
        self, api_admin, proveedor, compra_contado
    ):
        from apps.compras.models import CuentaCorrienteProveedor
        resp = api_admin.post(
            "/api/v1/compras/notas-credito/",
            {
                "proveedor": proveedor.pk,
                "monto_total": "5000",
                "compra_original": compra_contado.pk,
            },
            format="json",
        )
        assert resp.status_code == 201
        cc = CuentaCorrienteProveedor.objects.filter(
            proveedor=proveedor,
            tipo=CuentaCorrienteProveedor.Tipo.NOTA_CREDITO,
        ).last()
        assert f"Compra #{compra_contado.pk}" in cc.descripcion


# ── NotaCreditoProveedorViewSet.anular (líneas 333-352) ───────────────────────

@pytest.mark.django_db
class TestNotaCreditoAnular:
    """Cubre la action anular de NotaCreditoProveedorViewSet."""

    @pytest.fixture
    def nc(self, db, proveedor, usuario_admin, api_admin):
        resp = api_admin.post(
            "/api/v1/compras/notas-credito/",
            {"proveedor": proveedor.pk, "monto_total": "20000"},
            format="json",
        )
        assert resp.status_code == 201
        return resp.data

    def test_anular_nc_emitida_retorna_200_y_estado_anulada(self, api_admin, nc):
        resp = api_admin.post(f"/api/v1/compras/notas-credito/{nc['id']}/anular/")
        assert resp.status_code == 200
        assert resp.data["estado"] == "ANULADA"

    def test_anular_nc_emitida_crea_debito_en_cc(self, api_admin, nc, proveedor):
        from apps.compras.models import CuentaCorrienteProveedor
        api_admin.post(f"/api/v1/compras/notas-credito/{nc['id']}/anular/")
        assert CuentaCorrienteProveedor.objects.filter(
            proveedor=proveedor,
            tipo=CuentaCorrienteProveedor.Tipo.DEBITO,
        ).exists()

    def test_anular_dos_veces_retorna_400(self, api_admin, nc):
        api_admin.post(f"/api/v1/compras/notas-credito/{nc['id']}/anular/")
        resp = api_admin.post(f"/api/v1/compras/notas-credito/{nc['id']}/anular/")
        assert resp.status_code == 400
        assert "anulada" in resp.data["error"].lower()


# ── CompraViewSet.anular ────────────────────────────────────────────────────────

@pytest.fixture
def usuario_supervisor(db):
    from apps.usuarios.models import Usuario
    return Usuario.objects.create_user(
        email="supervisor_compras@test.com", password="test1234",
        nombre="Sup", apellido="Compras", rol="SUPERVISOR",
    )


@pytest.fixture
def api_supervisor(usuario_supervisor):
    client = APIClient()
    client.force_authenticate(user=usuario_supervisor)
    return client


@pytest.mark.django_db
class TestCompraAnular:
    """Cubre la action anular de CompraViewSet."""

    def test_sin_autenticacion_falla(self, compra_contado):
        client = APIClient()
        resp = client.post(f"/api/v1/compras/compras/{compra_contado.pk}/anular/")
        assert resp.status_code in (401, 403)

    def test_cajero_no_puede_anular(self, api, compra_contado):
        resp = api.post(f"/api/v1/compras/compras/{compra_contado.pk}/anular/")
        assert resp.status_code == 403

    def test_supervisor_puede_anular(self, api_supervisor, compra_contado):
        resp = api_supervisor.post(f"/api/v1/compras/compras/{compra_contado.pk}/anular/")
        assert resp.status_code == 200
        assert resp.data["estado_pago"] == "ANULADA"

    def test_anula_contado_revierte_stock(self, api_admin, compra_contado, producto):
        from apps.inventario.models import Stock
        stock_previo = Stock.objects.get(producto=producto).cantidad

        resp = api_admin.post(f"/api/v1/compras/compras/{compra_contado.pk}/anular/")

        assert resp.status_code == 200
        stock = Stock.objects.get(producto=producto)
        assert stock.cantidad == stock_previo - Decimal("2")

    def test_anula_contado_crea_movimiento_egreso_correccion(self, api_admin, compra_contado):
        from apps.inventario.models import MovimientoStock
        api_admin.post(f"/api/v1/compras/compras/{compra_contado.pk}/anular/")
        assert MovimientoStock.objects.filter(
            compra=compra_contado,
            tipo=MovimientoStock.Tipo.EGRESO,
            motivo=MovimientoStock.Motivo.CORRECCION,
        ).exists()

    def test_anula_credito_no_recibida_no_revierte_stock_pero_si_cc(
        self, api_admin, compra_credito, producto
    ):
        from apps.inventario.models import Stock, MovimientoStock
        from apps.compras.models import CuentaCorrienteProveedor
        # CREDITO recién registrada: aún no se confirmó entrega, no tocó stock.
        assert not Stock.objects.filter(producto=producto).exists()

        resp = api_admin.post(f"/api/v1/compras/compras/{compra_credito.pk}/anular/")

        assert resp.status_code == 200
        assert not MovimientoStock.objects.filter(compra=compra_credito).exists()
        assert CuentaCorrienteProveedor.objects.filter(
            compra=compra_credito,
            tipo=CuentaCorrienteProveedor.Tipo.CREDITO,
        ).exists()

    def test_anula_credito_recibida_revierte_stock_y_cc(
        self, api_admin, compra_credito, producto
    ):
        from apps.inventario.models import Stock
        from apps.compras.models import CuentaCorrienteProveedor
        api_admin.post(f"/api/v1/compras/compras/{compra_credito.pk}/confirmar-entrega/")
        stock_previo = Stock.objects.get(producto=producto).cantidad

        resp = api_admin.post(f"/api/v1/compras/compras/{compra_credito.pk}/anular/")

        assert resp.status_code == 200
        stock = Stock.objects.get(producto=producto)
        assert stock.cantidad == stock_previo - Decimal("3")
        assert CuentaCorrienteProveedor.objects.filter(
            compra=compra_credito,
            tipo=CuentaCorrienteProveedor.Tipo.CREDITO,
        ).exists()

    def test_anular_dos_veces_retorna_400(self, api_admin, compra_contado):
        api_admin.post(f"/api/v1/compras/compras/{compra_contado.pk}/anular/")
        resp = api_admin.post(f"/api/v1/compras/compras/{compra_contado.pk}/anular/")
        assert resp.status_code == 400
        assert "anulada" in resp.data["error"].lower()

    def test_con_pago_aplicado_no_se_puede_anular(
        self, api_admin, compra_credito, medio_pago_efectivo
    ):
        api_admin.post(
            "/api/v1/compras/pagos/",
            {
                "compra": compra_credito.pk,
                "monto": str(compra_credito.monto_total),
                "medio_pago": medio_pago_efectivo.pk,
            },
            format="json",
        )
        resp = api_admin.post(f"/api/v1/compras/compras/{compra_credito.pk}/anular/")
        assert resp.status_code == 400
        assert "pagos" in resp.data["error"].lower()
        compra_credito.refresh_from_db()
        assert compra_credito.estado_pago == "PAGADO"
