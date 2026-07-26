"""
Views para la app compras
"""

from django.db import models, transaction
from django.db.models import DecimalField, Sum, Value
from django.db.models.functions import Coalesce
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.views import APIView

from common.permissions import IsCajeroOrAdmin, IsAdminOrSupervisor, IsStaffUser
from common.mixins import ExportCSVMixin
from .filters import CompraFilter

from django_filters.rest_framework import DjangoFilterBackend

from .models import (
    Proveedor,
    CuentaCorrienteProveedor,
    Compra,
    DetalleCompra,
    PagoProveedor,
    AplicacionPagoCompra,
    NotaCreditoProveedor,
    DetalleNotaCreditoProveedor,
    OrdenCompra,
    ProductoProveedor,
)
from .services import CompraService
from .serializers import (
    ProveedorSerializer,
    CuentaCorrienteProveedorSerializer,
    CompraSerializer,
    CompraWriteSerializer,
    DetalleCompraSerializer,
    PagoProveedorSerializer,
    PagoProveedorWriteSerializer,
    AplicacionPagoCompraSerializer,
    NotaCreditoProveedorSerializer,
    DetalleNotaCreditoProveedorSerializer,
    OrdenCompraSerializer,
    ProductoProveedorSerializer,
)


class ProveedorViewSet(ExportCSVMixin, viewsets.ModelViewSet):
    queryset = Proveedor.objects.all()
    serializer_class = ProveedorSerializer
    permission_classes = [IsCajeroOrAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["activo"]
    search_fields = ["ruc", "razon_social"]
    ordering_fields = ["razon_social", "ruc"]
    ordering = ["razon_social"]
    export_filename = "proveedores"
    export_fields = [
        ("RUC", "ruc"),
        ("Razón Social", "razon_social"),
        ("Teléfono", "telefono"),
        ("Email", "email"),
        ("Activo", "activo"),
    ]


class CuentaCorrienteProveedorViewSet(viewsets.ModelViewSet):
    queryset = CuentaCorrienteProveedor.objects.select_related("proveedor").all()
    serializer_class = CuentaCorrienteProveedorSerializer
    permission_classes = [IsCajeroOrAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["proveedor", "tipo"]


class CompraViewSet(ExportCSVMixin, viewsets.ModelViewSet):
    serializer_class = CompraSerializer
    permission_classes = [IsStaffUser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = CompraFilter
    search_fields = ["proveedor__razon_social", "nro_factura_proveedor"]
    ordering_fields = ["fecha", "monto_total"]
    ordering = ["-fecha"]
    export_filename = "compras"
    export_fields = [
        ("Fecha", lambda o: str(o.fecha)[:10]),
        ("Proveedor", "proveedor__razon_social"),
        ("Nro. Factura", "nro_factura_proveedor"),
        ("Monto Total", "monto_total"),
        ("Estado Pago", "estado_pago"),
    ]

    def get_queryset(self):
        return (
            Compra.objects
            .select_related("proveedor")
            .prefetch_related("detalles")
            .annotate(
                _total_pagado=Coalesce(
                    Sum("aplicaciones_pago__monto_aplicado"),
                    Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
                )
            )
        )

    def _resolve_items(self, items_data):
        from apps.productos.models import Producto
        resolved = []
        for item in items_data:
            producto = Producto.objects.get(pk=item["producto"])
            resolved.append({
                "producto": producto,
                "cantidad": item["cantidad"],
                "costo_unitario": item["costo_unitario"],
            })
        return resolved

    def create(self, request, *args, **kwargs):
        write_ser = CompraWriteSerializer(data=request.data)
        write_ser.is_valid(raise_exception=True)
        data = write_ser.validated_data
        try:
            proveedor = Proveedor.objects.get(pk=data["proveedor"])
        except Proveedor.DoesNotExist:
            return Response({"proveedor": ["Proveedor no encontrado."]}, status=status.HTTP_400_BAD_REQUEST)
        try:
            items = self._resolve_items(data["items"])
            compra = CompraService.registrar_compra(
                proveedor=proveedor,
                creado_por=request.user,
                tipo_pago=data["tipo_pago"],
                items=items,
                nro_factura_proveedor=data.get("nro_factura_proveedor", ""),
                observaciones=data.get("observaciones", ""),
            )
        except ValidationError as exc:
            return Response(exc.detail, status=status.HTTP_400_BAD_REQUEST)
        compra_data = self.get_queryset().get(pk=compra.pk)
        return Response(CompraSerializer(compra_data).data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        write_ser = CompraWriteSerializer(data=request.data, partial=partial)
        write_ser.is_valid(raise_exception=True)
        data = write_ser.validated_data
        try:
            proveedor = Proveedor.objects.get(pk=data["proveedor"])
        except Proveedor.DoesNotExist:
            return Response({"proveedor": ["Proveedor no encontrado."]}, status=status.HTTP_400_BAD_REQUEST)
        from decimal import Decimal
        items = self._resolve_items(data.get("items", []))
        instance.proveedor = proveedor
        instance.tipo_pago = data.get("tipo_pago", instance.tipo_pago)
        instance.nro_factura_proveedor = data.get("nro_factura_proveedor", instance.nro_factura_proveedor)
        instance.observaciones = data.get("observaciones", instance.observaciones)
        if items:
            instance.detalles.all().delete()
            monto_total = Decimal("0")
            for item in items:
                subtotal = item["cantidad"] * item["costo_unitario"]
                monto_total += subtotal
                DetalleCompra.objects.create(
                    compra=instance,
                    producto=item["producto"],
                    cantidad=item["cantidad"],
                    costo_unitario=item["costo_unitario"],
                    subtotal=subtotal,
                )
            instance.monto_total = monto_total
        instance.save()
        compra_data = self.get_queryset().get(pk=instance.pk)
        return Response(CompraSerializer(compra_data).data)

    @action(detail=True, methods=["post"], url_path="confirmar-entrega")
    def confirmar_entrega(self, request, pk=None):
        """Confirma la recepción de mercadería de una compra a crédito pendiente."""
        compra = self.get_object()
        if compra.tipo_pago != Compra.TipoPago.CREDITO:
            return Response(
                {"error": "Solo se pueden confirmar entregas de compras a crédito."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            compra = CompraService.confirmar_compra(compra=compra, autorizado_por=request.user)
        except ValidationError as exc:
            return Response(exc.detail, status=status.HTTP_400_BAD_REQUEST)
        return Response(CompraSerializer(compra).data)


class DetalleCompraViewSet(viewsets.ModelViewSet):
    queryset = DetalleCompra.objects.select_related("compra", "producto").order_by("-compra__fecha", "-id")
    serializer_class = DetalleCompraSerializer
    permission_classes = [IsCajeroOrAdmin]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ["producto", "compra"]
    ordering_fields = ["compra__fecha"]


class PagoProveedorViewSet(viewsets.ModelViewSet):
    queryset = PagoProveedor.objects.select_related("proveedor", "medio_pago").prefetch_related("aplicaciones").all()
    serializer_class = PagoProveedorSerializer
    permission_classes = [IsAdminOrSupervisor]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["proveedor", "estado", "medio_pago"]

    def create(self, request, *args, **kwargs):
        from decimal import Decimal
        from apps.core.models import MedioPago
        write_ser = PagoProveedorWriteSerializer(data=request.data)
        write_ser.is_valid(raise_exception=True)
        data = write_ser.validated_data
        try:
            compra = Compra.objects.get(pk=data["compra"])
        except Compra.DoesNotExist:
            return Response({"compra": ["Compra no encontrada."]}, status=status.HTTP_400_BAD_REQUEST)
        try:
            medio_pago = MedioPago.objects.get(pk=data["medio_pago"])
        except MedioPago.DoesNotExist:
            return Response({"medio_pago": ["Medio de pago no encontrado."]}, status=status.HTTP_400_BAD_REQUEST)
        monto = Decimal(str(data["monto"]))
        with transaction.atomic():
            pago = PagoProveedor.objects.create(
                proveedor=compra.proveedor,
                monto_total=monto,
                medio_pago=medio_pago,
                observaciones=data.get("observaciones") or "",
                creado_por=request.user,
                estado=PagoProveedor.Estado.CONCILIADO,
            )
            AplicacionPagoCompra.objects.create(
                pago=pago,
                compra=compra,
                monto_aplicado=monto,
            )
            # Recalcular estado_pago de la compra
            total_pagado = AplicacionPagoCompra.objects.filter(compra=compra).aggregate(
                total=models.Sum("monto_aplicado")
            )["total"] or Decimal("0")
            if total_pagado >= compra.monto_total:
                compra.estado_pago = "PAGADO"
            elif total_pagado > 0:
                compra.estado_pago = "PARCIAL"
            compra.save(update_fields=["estado_pago"])

            # Registrar movimiento CRÉDITO en cuenta corriente del proveedor
            CuentaCorrienteProveedor.objects.create(
                proveedor=compra.proveedor,
                tipo=CuentaCorrienteProveedor.Tipo.CREDITO,
                monto=monto,
                pago=pago,
                compra=compra,
                descripcion=f"Pago compra #{compra.pk} - {medio_pago.descripcion}",
                creado_por=request.user,
            )

        return Response(PagoProveedorSerializer(pago).data, status=status.HTTP_201_CREATED)


class AplicacionPagoCompraViewSet(viewsets.ModelViewSet):
    queryset = AplicacionPagoCompra.objects.select_related("pago", "compra").all()
    serializer_class = AplicacionPagoCompraSerializer
    permission_classes = [IsCajeroOrAdmin]


class NotaCreditoProveedorViewSet(viewsets.ModelViewSet):
    queryset = (
        NotaCreditoProveedor.objects
        .select_related("proveedor", "compra_original")
        .prefetch_related("detalles")
        .all()
    )
    serializer_class = NotaCreditoProveedorSerializer
    permission_classes = [IsAdminOrSupervisor]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["proveedor", "estado"]

    def create(self, request, *args, **kwargs):
        from decimal import Decimal, InvalidOperation
        from apps.inventario.models import Stock, MovimientoStock

        proveedor_id = request.data.get("proveedor")
        monto_raw = request.data.get("monto_total")
        compra_id = request.data.get("compra_original")
        nro_factura = request.data.get("nro_factura_compra") or ""
        observacion = request.data.get("observacion") or ""
        tipo_nc = request.data.get("tipo_nc") or NotaCreditoProveedor.TipoNC.AJUSTE_PRECIO
        detalles_raw = request.data.get("detalles") or []

        if tipo_nc not in NotaCreditoProveedor.TipoNC.values:
            return Response({"error": "tipo_nc inválido."}, status=status.HTTP_400_BAD_REQUEST)

        if not proveedor_id:
            return Response({"error": "El campo 'proveedor' es obligatorio."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            monto = Decimal(str(monto_raw))
            if monto <= 0:
                raise ValueError
        except (InvalidOperation, ValueError, TypeError):
            return Response({"error": "El monto debe ser un número mayor a 0."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            proveedor = Proveedor.objects.get(pk=proveedor_id)
        except Proveedor.DoesNotExist:
            return Response({"error": "Proveedor no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        compra = None
        if compra_id:
            try:
                compra = Compra.objects.get(pk=compra_id, proveedor=proveedor)
            except Compra.DoesNotExist:
                return Response(
                    {"error": "Compra no encontrada para este proveedor."},
                    status=status.HTTP_404_NOT_FOUND,
                )

        # Validate detalles if DEVOLUCION
        detalles_validados = []
        if tipo_nc == NotaCreditoProveedor.TipoNC.DEVOLUCION and detalles_raw:
            from apps.productos.models import Producto
            for i, det in enumerate(detalles_raw):
                try:
                    prod = Producto.objects.get(pk=det["producto"])
                except (Producto.DoesNotExist, KeyError):
                    return Response(
                        {"error": f"Detalle {i+1}: producto no encontrado."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                try:
                    cant = Decimal(str(det["cantidad"]))
                    if cant <= 0:
                        raise ValueError
                except (InvalidOperation, ValueError, KeyError):
                    return Response({"error": f"Detalle {i+1}: cantidad inválida."}, status=status.HTTP_400_BAD_REQUEST)
                try:
                    pu = Decimal(str(det.get("precio_unitario", 0)))
                except (InvalidOperation, TypeError):
                    pu = Decimal("0")
                detalles_validados.append({"producto": prod, "cantidad": cant, "precio_unitario": pu})

        with transaction.atomic():
            nc = NotaCreditoProveedor.objects.create(
                proveedor=proveedor,
                compra_original=compra,
                monto_total=monto,
                nro_factura_compra=nro_factura,
                observacion=observacion,
                tipo_nc=tipo_nc,
                estado=NotaCreditoProveedor.Estado.EMITIDA,
                creado_por=request.user,
            )
            desc = f"NC #{nc.pk}"
            if compra:
                desc += f" - Compra #{compra.pk}"
            if observacion:
                desc += f" - {observacion}"
            CuentaCorrienteProveedor.objects.create(
                proveedor=proveedor,
                tipo=CuentaCorrienteProveedor.Tipo.NOTA_CREDITO,
                monto=monto,
                nota_credito=nc,
                compra=compra,
                descripcion=desc,
                creado_por=request.user,
            )

            # Generate stock EGRESO movements for devolución type
            if tipo_nc == NotaCreditoProveedor.TipoNC.DEVOLUCION and detalles_validados:
                for det in detalles_validados:
                    prod = det["producto"]
                    cant = det["cantidad"]
                    pu = det["precio_unitario"]
                    subtotal = (cant * pu).quantize(Decimal("1"))

                    DetalleNotaCreditoProveedor.objects.create(
                        nota_credito=nc,
                        producto=prod,
                        cantidad=cant,
                        precio_unitario=pu,
                        subtotal=subtotal,
                    )

                    if prod.requiere_stock and not prod.es_servicio:
                        stock, _ = Stock.objects.get_or_create(
                            producto=prod,
                            defaults={"cantidad": Decimal("0")},
                        )
                        stock = Stock.objects.select_for_update().get(pk=stock.pk)
                        stock.cantidad -= cant
                        stock.save()
                        MovimientoStock.objects.create(
                            producto=prod,
                            tipo=MovimientoStock.Tipo.EGRESO,
                            motivo=MovimientoStock.Motivo.DEVOLUCION_PROVEEDOR,
                            cantidad=cant,
                            stock_resultante=stock.cantidad,
                            nota_credito=nc,
                            autorizado_por=request.user,
                            observaciones=f"Devolución NC #{nc.pk}",
                        )

        return Response(self.get_serializer(nc).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["post"], url_path="anular")
    def anular(self, request, pk=None):
        from decimal import Decimal
        from apps.inventario.models import Stock, MovimientoStock

        nc = self.get_object()
        if nc.estado == NotaCreditoProveedor.Estado.ANULADA:
            return Response({"error": "La nota de crédito ya está anulada."}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            tiene_cc = CuentaCorrienteProveedor.objects.filter(nota_credito=nc).exists()
            if tiene_cc:
                # Revertir: la NC redujo la deuda → un DEBITO la restaura
                CuentaCorrienteProveedor.objects.create(
                    proveedor=nc.proveedor,
                    tipo=CuentaCorrienteProveedor.Tipo.DEBITO,
                    monto=nc.monto_total,
                    nota_credito=nc,
                    descripcion=f"Reversión por anulación NC #{nc.pk}",
                    creado_por=request.user,
                )

            # Reverse stock movements generated by this NC
            movimientos_egreso = MovimientoStock.objects.select_for_update().filter(
                nota_credito=nc, tipo=MovimientoStock.Tipo.EGRESO
            ).select_related("producto")
            for mov in movimientos_egreso:
                prod = mov.producto
                if prod.requiere_stock and not prod.es_servicio:
                    stock, _ = Stock.objects.get_or_create(
                        producto=prod,
                        defaults={"cantidad": Decimal("0")},
                    )
                    stock = Stock.objects.select_for_update().get(pk=stock.pk)
                    stock.cantidad += mov.cantidad
                    stock.save()
                    MovimientoStock.objects.create(
                        producto=prod,
                        tipo=MovimientoStock.Tipo.INGRESO,
                        motivo=MovimientoStock.Motivo.DEVOLUCION_PROVEEDOR,
                        cantidad=mov.cantidad,
                        stock_resultante=stock.cantidad,
                        nota_credito=nc,
                        autorizado_por=request.user,
                        observaciones=f"Reversión por anulación NC #{nc.pk}",
                    )

            nc.estado = NotaCreditoProveedor.Estado.ANULADA
            nc.save(update_fields=["estado"])

        return Response(self.get_serializer(nc).data)


class DetalleNotaCreditoProveedorViewSet(viewsets.ModelViewSet):
    queryset = DetalleNotaCreditoProveedor.objects.select_related("nota_credito", "producto").all()
    serializer_class = DetalleNotaCreditoProveedorSerializer
    permission_classes = [IsAdminOrSupervisor]


class ProductoProveedorViewSet(viewsets.ModelViewSet):
    queryset = ProductoProveedor.objects.select_related("proveedor", "producto").all()
    serializer_class = ProductoProveedorSerializer
    permission_classes = [IsStaffUser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["proveedor", "producto"]
    search_fields = ["producto__descripcion"]
    ordering_fields = ["producto__descripcion", "precio_compra", "fecha_ultima_compra"]
    ordering = ["producto__descripcion"]


class OrdenCompraViewSet(viewsets.ModelViewSet):
    serializer_class = OrdenCompraSerializer
    permission_classes = [IsCajeroOrAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["estado", "tipo_pago", "proveedor"]
    search_fields = ["proveedor__razon_social", "nro_factura_esperada"]
    ordering_fields = ["fecha_creacion", "monto_total"]
    ordering = ["-fecha_creacion"]

    def get_queryset(self):
        return (
            OrdenCompra.objects
            .select_related("proveedor", "creado_por", "aprobado_por")
            .prefetch_related("detalles__producto")
        )

    def _require_approver(self, request):
        """Sólo ADMIN y SUPERVISOR pueden aprobar/rechazar/convertir."""
        rol = getattr(request.user, "rol", None)
        if rol not in ("ADMIN", "SUPERVISOR"):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Solo ADMIN o SUPERVISOR pueden realizar esta acción.")

    @action(detail=True, methods=["post"])
    def submit(self, request, pk=None):
        """Envía la OC a revisión: BORRADOR → PENDIENTE."""
        orden = self.get_object()
        if orden.estado != OrdenCompra.Estado.BORRADOR:
            return Response(
                {"error": "Solo se puede enviar a revisión una OC en estado Borrador."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not orden.detalles.exists():
            return Response(
                {"error": "La OC debe tener al menos un producto."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        orden.estado = OrdenCompra.Estado.PENDIENTE
        orden.save(update_fields=["estado", "fecha_actualizacion"])
        return Response(OrdenCompraSerializer(orden).data)

    @action(detail=True, methods=["post"])
    def aprobar(self, request, pk=None):
        """Aprueba la OC: PENDIENTE → APROBADA."""
        self._require_approver(request)
        orden = self.get_object()
        if orden.estado != OrdenCompra.Estado.PENDIENTE:
            return Response(
                {"error": "Solo se puede aprobar una OC en estado Pendiente."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        from django.utils import timezone
        orden.estado = OrdenCompra.Estado.APROBADA
        orden.aprobado_por = request.user
        orden.fecha_aprobacion = timezone.now()
        orden.motivo_rechazo = None
        orden.save(update_fields=[
            "estado", "aprobado_por", "fecha_aprobacion",
            "motivo_rechazo", "fecha_actualizacion",
        ])
        return Response(OrdenCompraSerializer(orden).data)

    @action(detail=True, methods=["post"])
    def rechazar(self, request, pk=None):
        """Rechaza la OC: PENDIENTE → RECHAZADA."""
        self._require_approver(request)
        orden = self.get_object()
        if orden.estado != OrdenCompra.Estado.PENDIENTE:
            return Response(
                {"error": "Solo se puede rechazar una OC en estado Pendiente."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        motivo = request.data.get("motivo", "").strip()
        if not motivo:
            return Response({"error": "Debe indicar el motivo del rechazo."}, status=status.HTTP_400_BAD_REQUEST)
        orden.estado = OrdenCompra.Estado.RECHAZADA
        orden.motivo_rechazo = motivo
        orden.save(update_fields=["estado", "motivo_rechazo", "fecha_actualizacion"])
        return Response(OrdenCompraSerializer(orden).data)

    @action(detail=True, methods=["post"])
    def convertir(self, request, pk=None):
        """Convierte la OC aprobada en una Compra real: APROBADA → CONVERTIDA."""
        self._require_approver(request)
        orden = self.get_object()
        if orden.estado != OrdenCompra.Estado.APROBADA:
            return Response(
                {"error": "Solo se puede convertir una OC en estado Aprobada."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        items = [
            {
                "producto": det.producto,
                "cantidad": det.cantidad,
                "costo_unitario": det.costo_unitario,
            }
            for det in orden.detalles.select_related("producto").all()
        ]
        try:
            compra = CompraService.registrar_compra(
                proveedor=orden.proveedor,
                creado_por=request.user,
                tipo_pago=orden.tipo_pago,
                items=items,
                nro_factura_proveedor=orden.nro_factura_esperada or "",
                observaciones=orden.observaciones or "",
            )
        except ValidationError as exc:
            return Response(exc.detail, status=status.HTTP_400_BAD_REQUEST)

        orden.estado = OrdenCompra.Estado.CONVERTIDA
        orden.compra_generada = compra
        orden.save(update_fields=["estado", "compra_generada", "fecha_actualizacion"])
        return Response({
            "orden": OrdenCompraSerializer(orden).data,
            "compra_id": compra.pk,
        })


class ReporteComprasProveedoresView(APIView):
    """
    GET /api/compras/reporte-compras/?desde=YYYY-MM-DD&hasta=YYYY-MM-DD
    Gasto por proveedor, tasa de entrega y funnel de órdenes de compra.
    Opcional: ?formato=csv
    """
    permission_classes = [IsStaffUser]

    def get(self, request):
        import csv as csv_module
        from django.http import HttpResponse as HR
        from django.db.models import Count, Sum, Q

        desde = request.query_params.get("desde")
        hasta = request.query_params.get("hasta")

        if not desde or not hasta:
            return Response(
                {"error": "Se requieren los parámetros desde y hasta (YYYY-MM-DD)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Gasto y fulfilment por proveedor
        qs = (
            Compra.objects
            .filter(fecha__date__gte=desde, fecha__date__lte=hasta)
            .values("proveedor__id", "proveedor__razon_social", "proveedor__ruc")
            .annotate(
                n_compras=Count("id"),
                monto_total=Sum("monto_total"),
                entregadas=Count("id", filter=Q(estado_entrega="RECIBIDA")),
                entrega_parcial=Count("id", filter=Q(estado_entrega="PARCIAL")),
                entrega_pendiente=Count("id", filter=Q(estado_entrega="PENDIENTE")),
                pagadas=Count("id", filter=Q(estado_pago="PAGADO")),
                pago_parcial=Count("id", filter=Q(estado_pago="PARCIAL")),
                pago_pendiente=Count("id", filter=Q(estado_pago="PENDIENTE")),
            )
            .order_by("-monto_total")
        )

        por_proveedor = []
        for r in qs:
            n = r["n_compras"] or 1
            por_proveedor.append({
                "proveedor_id": r["proveedor__id"],
                "proveedor": r["proveedor__razon_social"] or "",
                "ruc": r["proveedor__ruc"] or "",
                "n_compras": r["n_compras"],
                "monto_total": int(r["monto_total"] or 0),
                "entregadas": r["entregadas"],
                "entrega_parcial": r["entrega_parcial"],
                "entrega_pendiente": r["entrega_pendiente"],
                "tasa_entrega": round(r["entregadas"] / n * 100, 1),
                "pagadas": r["pagadas"],
                "pago_parcial": r["pago_parcial"],
                "pago_pendiente": r["pago_pendiente"],
            })

        # Funnel de órdenes de compra en el período
        oc_qs = (
            OrdenCompra.objects
            .filter(fecha_creacion__date__gte=desde, fecha_creacion__date__lte=hasta)
            .values("estado")
            .annotate(n=Count("id"))
        )
        funnel = {e: 0 for e in ["BORRADOR", "PENDIENTE", "APROBADA", "RECHAZADA", "CONVERTIDA"]}
        for r in oc_qs:
            funnel[r["estado"]] = r["n"]
        funnel["total"] = sum(funnel.values())

        total_monto = sum(p["monto_total"] for p in por_proveedor)
        total_compras = sum(p["n_compras"] for p in por_proveedor)

        if request.query_params.get("formato") == "csv":
            response = HR(content_type="text/csv; charset=utf-8-sig")
            response["Content-Disposition"] = (
                f'attachment; filename="compras_proveedores_{desde}_{hasta}.csv"'
            )
            writer = csv_module.writer(response)
            writer.writerow(["REPORTE COMPRAS POR PROVEEDOR", f"{desde} al {hasta}"])
            writer.writerow([])
            writer.writerow([
                "Proveedor", "RUC", "N° Compras", "Monto Total (Gs)",
                "Entregadas", "% Entrega", "Pagadas",
            ])
            for p in por_proveedor:
                writer.writerow([p["proveedor"], p["ruc"], p["n_compras"], p["monto_total"],
                                  p["entregadas"], p["tasa_entrega"], p["pagadas"]])
            writer.writerow([])
            writer.writerow(["FUNNEL ÓRDENES DE COMPRA"])
            for estado, n in funnel.items():
                writer.writerow([estado, n])
            return response

        return Response({
            "periodo": {"desde": desde, "hasta": hasta},
            "resumen": {
                "total_compras": total_compras,
                "monto_total": total_monto,
                "n_proveedores": len(por_proveedor),
            },
            "por_proveedor": por_proveedor,
            "funnel_oc": funnel,
        })


class ReporteNotasCreditoCompraView(APIView):
    """
    GET /api/compras/reporte-notas-credito/?desde=YYYY-MM-DD&hasta=YYYY-MM-DD[&estado=]
    Notas de crédito emitidas por proveedores: volumen, montos y quién las registra.
    Opcional: ?formato=csv
    """
    permission_classes = [IsStaffUser]

    def get(self, request):
        import csv as csv_module
        from django.http import HttpResponse as HR

        desde = request.query_params.get("desde")
        hasta = request.query_params.get("hasta")
        estado_filtro = request.query_params.get("estado")

        if not desde or not hasta:
            return Response(
                {"error": "Se requieren los parámetros desde y hasta (YYYY-MM-DD)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = NotaCreditoProveedor.objects.filter(
            fecha__date__gte=desde,
            fecha__date__lte=hasta,
        ).select_related("proveedor", "compra_original", "creado_por")

        if estado_filtro:
            qs = qs.filter(estado=estado_filtro)

        filas = []
        for nc in qs.order_by("-fecha"):
            proveedor = nc.proveedor
            registrado_por = nc.creado_por
            filas.append({
                "id": nc.pk,
                "fecha": nc.fecha.strftime("%Y-%m-%d %H:%M"),
                "proveedor_id": proveedor.pk if proveedor else None,
                "proveedor": proveedor.razon_social if proveedor else "",
                "ruc": getattr(proveedor, "ruc", "") or "",
                "estado": nc.estado,
                "observacion": nc.observacion or "",
                "nro_factura_compra": nc.nro_factura_compra or "",
                "monto_total": int(nc.monto_total),
                "creado_por_id": registrado_por.pk if registrado_por else None,
                "creado_por": (
                    f"{registrado_por.nombre} {registrado_por.apellido}".strip()
                    if registrado_por else ""
                ),
                "compra_original_id": nc.compra_original_id,
            })

        resumen = {
            "total_emitidas": sum(1 for f in filas if f["estado"] == "EMITIDA"),
            "total_aplicadas": sum(1 for f in filas if f["estado"] == "APLICADA"),
            "total_anuladas": sum(1 for f in filas if f["estado"] == "ANULADA"),
            "monto_emitidas": sum(f["monto_total"] for f in filas if f["estado"] == "EMITIDA"),
            "monto_aplicadas": sum(f["monto_total"] for f in filas if f["estado"] == "APLICADA"),
            "monto_anuladas": sum(f["monto_total"] for f in filas if f["estado"] == "ANULADA"),
            "monto_total": sum(f["monto_total"] for f in filas if f["estado"] != "ANULADA"),
        }

        if request.query_params.get("formato") == "csv":
            response = HR(content_type="text/csv; charset=utf-8-sig")
            response["Content-Disposition"] = (
                f'attachment; filename="notas_credito_compras_{desde}_{hasta}.csv"'
            )
            writer = csv_module.writer(response)
            writer.writerow(["NOTAS DE CRÉDITO - COMPRAS", f"{desde} al {hasta}"])
            writer.writerow([])
            writer.writerow([
                "ID", "Fecha", "Proveedor", "RUC", "Estado",
                "Observación", "N° Factura", "Monto (Gs)", "Registrado por",
            ])
            for f in filas:
                writer.writerow([
                    f["id"], f["fecha"], f["proveedor"], f["ruc"],
                    f["estado"], f["observacion"], f["nro_factura_compra"],
                    f["monto_total"], f["creado_por"],
                ])
            writer.writerow([])
            writer.writerow(["RESUMEN"])
            writer.writerow(["Emitidas", resumen["total_emitidas"], resumen["monto_emitidas"]])
            writer.writerow(["Aplicadas", resumen["total_aplicadas"], resumen["monto_aplicadas"]])
            writer.writerow(["Anuladas", resumen["total_anuladas"], resumen["monto_anuladas"]])
            return response

        return Response({
            "periodo": {"desde": desde, "hasta": hasta},
            "resumen": resumen,
            "detalle": filas,
        })


class ReporteAgingProveedoresView(APIView):
    """
    GET /api/compras/reporte-aging-proveedores/
    Proveedores con saldo pendiente y distribución por aging (30/60/90/90+ días).
    Opcional: ?formato=csv
    """
    permission_classes = [IsStaffUser]

    def get(self, request):
        import csv as csv_module
        from datetime import date
        from django.http import HttpResponse as HR
        from django.db.models import OuterRef, Subquery

        hoy = date.today()

        ultimo_mov = (
            CuentaCorrienteProveedor.objects
            .filter(proveedor=OuterRef("pk"))
            .order_by("-id")
            .values("saldo_resultante")[:1]
        )
        proveedores_con_saldo = (
            Proveedor.objects
            .filter(activo=True)
            .annotate(saldo_deuda=Subquery(ultimo_mov))
            .filter(saldo_deuda__gt=0)
            .order_by("-saldo_deuda")
        )

        filas = []
        for proveedor in proveedores_con_saldo:
            from decimal import Decimal
            saldo = Decimal(str(proveedor.saldo_deuda or 0))

            ultimo_saldo_cero = (
                CuentaCorrienteProveedor.objects
                .filter(proveedor=proveedor, saldo_resultante__lte=0)
                .order_by("-fecha", "-id")
                .values("id", "fecha")
                .first()
            )
            dias_atraso = 0
            if ultimo_saldo_cero:
                inicio_ciclo = (
                    CuentaCorrienteProveedor.objects
                    .filter(proveedor=proveedor, id__gt=ultimo_saldo_cero["id"], saldo_resultante__gt=0)
                    .order_by("fecha", "id")
                    .values("fecha")
                    .first()
                )
            else:
                inicio_ciclo = (
                    CuentaCorrienteProveedor.objects
                    .filter(proveedor=proveedor, saldo_resultante__gt=0)
                    .order_by("fecha", "id")
                    .values("fecha")
                    .first()
                )
            if inicio_ciclo:
                dias_atraso = (hoy - inicio_ciclo["fecha"].date()).days

            if dias_atraso <= 30:
                bucket = "0-30"
            elif dias_atraso <= 60:
                bucket = "31-60"
            elif dias_atraso <= 90:
                bucket = "61-90"
            else:
                bucket = "90+"

            filas.append({
                "proveedor_id": proveedor.pk,
                "proveedor": proveedor.razon_social,
                "ruc": proveedor.ruc or "",
                "telefono": proveedor.telefono or "",
                "email": proveedor.email or "",
                "saldo_deuda": int(saldo),
                "dias_atraso": dias_atraso,
                "aging": bucket,
            })

        aging_totales = {"0-30": 0, "31-60": 0, "61-90": 0, "90+": 0}
        for f in filas:
            aging_totales[f["aging"]] += f["saldo_deuda"]

        total_deuda = sum(f["saldo_deuda"] for f in filas)

        formato = request.query_params.get("formato")

        if formato == "csv":
            response = HR(content_type="text/csv; charset=utf-8-sig")
            response["Content-Disposition"] = (
                f'attachment; filename="aging_proveedores_{hoy}.csv"'
            )
            writer = csv_module.writer(response)
            writer.writerow(["AGING CUENTAS A PAGAR - PROVEEDORES", str(hoy)])
            writer.writerow([])
            writer.writerow(["Proveedor", "RUC", "Teléfono", "Email", "Saldo (Gs)", "Días atraso", "Aging"])
            for f in filas:
                writer.writerow([
                    f["proveedor"], f["ruc"], f["telefono"],
                    f["email"], f["saldo_deuda"], f["dias_atraso"], f["aging"],
                ])
            writer.writerow([])
            writer.writerow(["TOTALES POR AGING"])
            for bucket, total in aging_totales.items():
                writer.writerow([bucket, total])
            return response

        if formato == "excel":
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Aging Proveedores"

            header_fill = PatternFill("solid", fgColor="1E3A5F")
            header_font = Font(bold=True, color="FFFFFF")
            total_font = Font(bold=True)
            totals_fill = PatternFill("solid", fgColor="E8F0FE")

            ws.append([f"AGING CUENTAS A PAGAR — PROVEEDORES — {hoy}"])
            ws["A1"].font = Font(bold=True, size=13)
            ws.append([])

            headers = ["Proveedor", "RUC", "Teléfono", "Email", "Saldo (Gs)", "Días atraso", "Aging"]
            ws.append(headers)
            for cell in ws[ws.max_row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for f in filas:
                ws.append([f["proveedor"], f["ruc"], f["telefono"],
                            f["email"], f["saldo_deuda"], f["dias_atraso"], f["aging"]])

            ws.append([])
            ws.append(["TOTALES POR AGING", "", "", "", "", "", ""])
            for cell in ws[ws.max_row]:
                cell.font = total_font
            for bucket, total in aging_totales.items():
                row = [bucket, "", "", "", total, "", ""]
                ws.append(row)
                for cell in ws[ws.max_row]:
                    cell.fill = totals_fill

            ws.append([])
            ws.append(["TOTAL DEUDA", "", "", "", total_deuda, "", ""])
            for cell in ws[ws.max_row]:
                cell.font = total_font

            col_widths = [35, 14, 14, 28, 14, 14, 10]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            response = HR(
                buf.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="aging_proveedores_{hoy}.xlsx"'
            return response

        return Response({
            "fecha": str(hoy),
            "resumen": {
                "proveedores_con_deuda": len(filas),
                "total_deuda": total_deuda,
                "aging": aging_totales,
            },
            "detalle": filas,
        })
